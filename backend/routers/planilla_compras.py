"""Planilla de Compras -- replica del Excel "PLANILLA DE COMPRAS OFICIAL
2026": ledger mensual de las facturas de proveedor ya ingresadas en Odoo,
con una categoria (Tipo) POR PROVEEDOR que solo vive en nuestra base --
nunca se escribe en Odoo, se asigna una vez y se reusa siempre.

Tipos: AL=Alimentos, BA=Barra, GF=Gastos Fijos, OT=Otros, AS=Aseo.

Por ahora escaneado solo para Doña Delfina (company_id=2 en Odoo) -- el
Excel original es de ese local. Se puede generalizar a los demas locales
mas adelante agregando el company_id correspondiente."""
import os
import sys
from calendar import monthrange
from datetime import date, timedelta
from pathlib import Path

from fastapi import APIRouter, Depends, HTTPException, Response, status

from ..db import get_db
from ..deps import get_current_claims
from ..excel_exporter_compras import exportar_mes

_REPO_ROOT = Path(__file__).resolve().parents[2]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))
from odoo_connector import OdooClient  # noqa: E402
from tcpos_connector import TcposWebReportSession, construir_parametros  # noqa: E402

from ..schemas import (PlanillaComprasItem, PlanillaComprasOut, PlanillaComprasResumen, ProveedorTipoIn,
                       ProveedorTipoOut, VentaPeriodoIn, VentaPeriodoTcposOut)
from ..tcpos_report_parser import parsear_financial_overview_cash_to_deposit

_MESES_ES = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO",
             "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]

router = APIRouter(prefix="/planilla-compras", tags=["planilla-compras"])

COMPANY_ID_DONA_DELFINA = 2
TIPOS_VALIDOS = {"AL", "BA", "GF", "OT", "AS"}
TIPOS_COSTO_VENTA = {"AL", "BA"}  # igual que el Excel real: Costo Venta = N6+O6 (solo Alimentos + Barra)

# Reporte "Financial overview" de TCPOS -- confirmados via el endpoint de
# descubrimiento (ya borrado). Mismo outlet que el import diario de ventas
# (planilla.py) -- "1001 Margo Isidora" == Doña Delfina en este sistema.
_TCPOS_REPORT_FORM_NAME = "FinancialOverview1Form"
_TCPOS_REPORT_ASSEMBLY_NAME = "Report.FinancialOverview1"
_TCPOS_OUTLET_ID_MARGO_ISIDORA = 13


def _require_admin(claims: dict):
    if claims["rol"] != "administrador":
        raise HTTPException(status.HTTP_403_FORBIDDEN, "Solo un administrador puede ver la Planilla de Compras")


def _odoo() -> OdooClient:
    try:
        cliente = OdooClient(os.environ["ODOO_URL"], os.environ["ODOO_DB"],
                              os.environ["ODOO_FACTURAS_USER"], os.environ["ODOO_FACTURAS_PASSWORD"])
    except KeyError as e:
        raise HTTPException(status.HTTP_500_INTERNAL_SERVER_ERROR, f"Falta configurar la variable de entorno {e}")
    ok, msg = cliente.connect()
    if not ok:
        raise HTTPException(status.HTTP_502_BAD_GATEWAY, f"No se pudo conectar a Odoo: {msg}")
    return cliente


def _obtener_items_y_resumen(anio: int, mes: int) -> PlanillaComprasOut:
    cliente = _odoo()
    ultimo_dia = monthrange(anio, mes)[1]
    desde = f"{anio:04d}-{mes:02d}-01"
    hasta = f"{anio:04d}-{mes:02d}-{ultimo_dia:02d}"

    # invoice_origin != False -- solo facturas que vienen de una Orden de
    # Compra (el flujo OC -> recepcion -> factura). Sin eso, "Facturas de
    # proveedores" tambien trae bancos, seguros, arriendos (inmobiliarias),
    # telefonia, etc. -- gastos administrativos que nunca pasan por una OC
    # y que esta planilla NO debe mostrar (solo compras de mercaderia).
    moves = cliente._call('account.move', 'search_read',
        [[['move_type', '=', 'in_invoice'], ['company_id', '=', COMPANY_ID_DONA_DELFINA],
          ['invoice_date', '>=', desde], ['invoice_date', '<=', hasta], ['state', '!=', 'cancel'],
          ['invoice_origin', '!=', False]]],
        {'fields': ['id', 'partner_id', 'l10n_latam_document_number', 'invoice_date', 'amount_untaxed', 'amount_total'],
         'order': 'invoice_date'})

    db = get_db()
    mapeos = db.table("planilla_compras_proveedor_tipo").select("odoo_partner_id,tipo").execute().data or []
    tipo_por_partner = {m["odoo_partner_id"]: m["tipo"] for m in mapeos}

    items = []
    for m in moves:
        partner_id, partner_nombre = m.get('partner_id') or [None, '']
        subtotal = m.get('amount_untaxed') or 0
        total = m.get('amount_total') or 0
        items.append(PlanillaComprasItem(
            factura_id=m['id'], proveedor_id=partner_id, proveedor_nombre=partner_nombre,
            num_factura=m.get('l10n_latam_document_number') or '', fecha=m.get('invoice_date'),
            subtotal=subtotal, iva=total - subtotal, total=total,
            tipo=tipo_por_partner.get(partner_id),
        ))

    # % Costo Venta -- misma formula del Excel real: Costo Venta = compras
    # Tipo AL+BA del mes (subtotal, sin IVA); Venta Neta = Venta del Periodo
    # (ingresada a mano, igual que en el Excel) / 1.19.
    costo_venta = sum(it.subtotal for it in items if it.tipo in TIPOS_COSTO_VENTA)
    fila_venta = db.table("planilla_compras_venta_periodo").select("venta_periodo") \
        .eq("anio", anio).eq("mes", mes).execute().data
    venta_periodo = fila_venta[0]["venta_periodo"] if fila_venta else None
    venta_neta = venta_periodo / 1.19 if venta_periodo else None
    pct_costo_venta = costo_venta / venta_neta if venta_neta else None

    resumen = PlanillaComprasResumen(
        venta_periodo=venta_periodo, venta_neta=venta_neta,
        costo_venta=costo_venta, pct_costo_venta=pct_costo_venta,
    )
    return PlanillaComprasOut(items=items, resumen=resumen)


@router.get("", response_model=PlanillaComprasOut)
def listar(anio: int, mes: int, claims: dict = Depends(get_current_claims)):
    """Todas las facturas de proveedor del mes en Odoo (Doña Delfina), con
    el Tipo de cada una resuelto por su proveedor -- null si ese proveedor
    todavia no tiene Tipo asignado (hay que clasificarlo en /proveedores)."""
    _require_admin(claims)
    return _obtener_items_y_resumen(anio, mes)


@router.get("/exportar")
def exportar(anio: int, mes: int, claims: dict = Depends(get_current_claims)):
    """Genera el Excel real "PLANILLA DE COMPRAS OFICIAL", ya lleno con las
    facturas del mes -- misma plantilla original, con sus formulas intactas.
    Escribe Tipo/Proveedor/N Factura/IVA/Total tal cual vienen de Odoo (sin
    intentar matchear el nombre corto de las columnas de desglose por
    proveedor -- esas quedan en 0 hasta que se ajusten a mano, igual que
    cualquier fila nueva en el Excel real)."""
    _require_admin(claims)
    if not (1 <= mes <= 12):
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Mes inválido")
    datos = _obtener_items_y_resumen(anio, mes)
    contenido = exportar_mes(anio, mes, [it.model_dump() for it in datos.items], datos.resumen.venta_periodo)
    nombre_mes = _MESES_ES[mes - 1].capitalize()
    return Response(
        content=contenido,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="Planilla de Compras {nombre_mes} {anio}.xlsx"'},
    )


@router.get("/_debug-verificar-export")
def _debug_verificar_export(anio: int, mes: int, claims: dict = Depends(get_current_claims)):
    """TEMPORAL -- confirma que la celda G (SUB TOTAL) quede con un valor fijo
    (no una formula que dependa de que el visor recalcule al abrir)."""
    _require_admin(claims)
    import openpyxl
    from io import BytesIO
    datos = _obtener_items_y_resumen(anio, mes)
    contenido = exportar_mes(anio, mes, [it.model_dump() for it in datos.items], datos.resumen.venta_periodo)
    wb = openpyxl.load_workbook(BytesIO(contenido))
    ws = wb[_MESES_ES[mes - 1]]
    filas = []
    for r in range(9, 9 + min(len(datos.items), 5)):
        filas.append({
            "fila": r,
            "tipo": ws.cell(row=r, column=4).value,
            "proveedor": ws.cell(row=r, column=5).value,
            "num_factura": ws.cell(row=r, column=6).value,
            "subtotal_G": ws.cell(row=r, column=7).value,
            "iva_H": ws.cell(row=r, column=8).value,
            "total_I": ws.cell(row=r, column=9).value,
        })
    return {"total_items": len(datos.items), "filas_muestra": filas}


@router.put("/venta-periodo", status_code=status.HTTP_204_NO_CONTENT)
def fijar_venta_periodo(body: VentaPeriodoIn, claims: dict = Depends(get_current_claims)):
    """Venta del periodo ($) del mes -- editable a mano (igual que en el
    Excel real), y tambien se puede precargar desde TCPOS con
    GET /venta-periodo/tcpos antes de guardar."""
    _require_admin(claims)
    db = get_db()
    db.table("planilla_compras_venta_periodo").upsert({
        "anio": body.anio, "mes": body.mes, "venta_periodo": body.venta_periodo,
        "actualizado_por": claims["sub"],
    }, on_conflict="anio,mes").execute()


@router.get("/venta-periodo/tcpos", response_model=VentaPeriodoTcposOut)
def obtener_venta_periodo_tcpos(anio: int, mes: int, claims: dict = Depends(get_current_claims)):
    """Trae la Venta del Periodo real desde TCPOS -- reporte "Financial
    overview", "Cash to deposit" de la fila Total (asi lo definio el
    usuario: es el monto vendido). Rango: dia 1 del mes hasta AYER como
    maximo -- nunca hasta hoy, el dia en curso todavia no cierra (mismo
    criterio que el import diario de ventas en planilla.py). Solo trae el
    valor, no lo guarda -- el admin confirma con PUT /venta-periodo."""
    _require_admin(claims)
    ultimo_dia_mes = monthrange(anio, mes)[1]
    hasta = date(anio, mes, ultimo_dia_mes)
    ayer = date.today() - timedelta(days=1)
    if hasta > ayer:
        hasta = ayer
    desde = date(anio, mes, 1)
    if desde > hasta:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, "Todavía no hay ventas cerradas para ese mes")

    try:
        session = TcposWebReportSession(
            os.environ["TCPOS_URL"], os.environ["TCPOS_OPERATOR_CODE"], os.environ["TCPOS_PASSWORD"],
        )
    except KeyError as e:
        raise HTTPException(status.HTTP_500_INTERNAL_SERVER_ERROR, f"Falta configurar la variable de entorno {e}")

    try:
        formulario = session.formulario_de_parametros(_TCPOS_REPORT_FORM_NAME, _TCPOS_REPORT_ASSEMBLY_NAME)
        overrides = {
            "edDateFrom": f"{desde.isoformat()}T00:00:00", "edDateTo": f"{hasta.isoformat()}T00:00:00",
            "rbSolarDate": True, "clbShops": [_TCPOS_OUTLET_ID_MARGO_ISIDORA], "ckWithdrawalDeposit": True,
        }
        parametros = construir_parametros(formulario, overrides)
        resultado = session.ejecutar_reporte(_TCPOS_REPORT_FORM_NAME, _TCPOS_REPORT_ASSEMBLY_NAME, parametros)
        pdf_url = resultado.get("pdfUrl") if isinstance(resultado, dict) else None
        if not pdf_url:
            raise RuntimeError(f"TCPOS no devolvió un pdfUrl válido: {resultado}")
        pdf_bytes = session.descargar_archivo(pdf_url)
        venta_periodo = parsear_financial_overview_cash_to_deposit(pdf_bytes)
    except Exception as e:
        raise HTTPException(status.HTTP_502_BAD_GATEWAY, f"Error al traer la venta desde TCPOS: {e}")

    return VentaPeriodoTcposOut(anio=anio, mes=mes, venta_periodo=venta_periodo,
                                 desde=desde.isoformat(), hasta=hasta.isoformat())


@router.get("/proveedores", response_model=list[ProveedorTipoOut])
def listar_proveedores(claims: dict = Depends(get_current_claims)):
    """Catalogo completo proveedor -> Tipo, visible y editable."""
    _require_admin(claims)
    db = get_db()
    rows = db.table("planilla_compras_proveedor_tipo").select("*").order("proveedor_nombre").execute().data or []
    return [ProveedorTipoOut(**r) for r in rows]


@router.put("/proveedores", status_code=status.HTTP_204_NO_CONTENT)
def asignar_tipo(body: ProveedorTipoIn, claims: dict = Depends(get_current_claims)):
    _require_admin(claims)
    if body.tipo not in TIPOS_VALIDOS:
        raise HTTPException(status.HTTP_400_BAD_REQUEST, f"Tipo inválido -- debe ser uno de {sorted(TIPOS_VALIDOS)}")
    db = get_db()
    db.table("planilla_compras_proveedor_tipo").upsert({
        "odoo_partner_id": body.odoo_partner_id, "proveedor_nombre": body.proveedor_nombre,
        "tipo": body.tipo, "actualizado_por": claims["sub"],
    }, on_conflict="odoo_partner_id").execute()
