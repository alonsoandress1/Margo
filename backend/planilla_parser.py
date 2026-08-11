"""Parser de la planilla semanal "Inventario Cocina Semanal" (Excel).

Cada hoja del dia (ej. "S1 Lunes") tiene 3 bloques que nos interesan:

1. Bloque de insumos (filas 4-57): Stock Inicial/Entregas/Ventas/Mermas
   por insumo. Las columnas F (Entregas) y J (Stock Informado) son
   formulas que apuntan a una fila especifica de los bloques 2 y 3
   (el offset no es constante entre insumos, asi que se lee la formula
   en vez de asumir una posicion fija).
2. Bloque "MERMAS" (filas ~157-216, en dos tramos: kilogramos y
   unidades): Stock Informado + desglose de mermas por motivo
   (Produccion/Defectuosos/Clientes/Cortesia/Reutilizar) para cada
   insumo, en la fila que el bloque 1 referencia.
3. Bloque "VENTAS" (filas 66-153): Codigo/Item (SKU del POS) + Vendido
   -- un plato puede aparecer en varias filas (una por insumo de su
   receta), asi que se toma una sola vez por Codigo.
4. Bloque "ENTREGAS A COCINA" (filas ~221+): cantidad entregada desde
   Bodega a Cocina ese dia, en la fila que el bloque 1 referencia via
   la formula de la columna F.

Solo lectura -- no modifica el archivo. Las filas exactas de los
bloques 2/3/4 se resuelven dinamicamente a partir de las formulas del
bloque 1, no estan hardcodeadas, porque no son un offset constante.
"""
import re
from typing import BinaryIO

import openpyxl

BLOQUE_INSUMOS = range(4, 58)
VENTAS_DESDE = 66
VENTAS_HASTA = 153
MERMA_COLUMNAS = [
    ("produccion", 6), ("defectuosos", 7), ("clientes", 8), ("cortesia", 9), ("reutilizar", 10),
]

_RE_REF = re.compile(r"=E(\d+)")


def _ref_fila(formula) -> int | None:
    if not formula:
        return None
    m = _RE_REF.match(str(formula).strip())
    return int(m.group(1)) if m else None


def _num(v) -> float:
    try:
        return float(v) if v is not None else 0.0
    except (TypeError, ValueError):
        return 0.0


def hojas_disponibles(archivo: str | BinaryIO) -> list[str]:
    """Nombres de las hojas de dia (excluye la hoja resumen 'MERMAS S1')."""
    wb = openpyxl.load_workbook(archivo, data_only=True, read_only=True)
    return [n for n in wb.sheetnames if "MERMAS" not in n.upper()]


def parsear_dia(archivo: str | BinaryIO, hoja: str) -> dict:
    """Retorna {"ventas": [...], "insumos": [...]} para la hoja de un dia especifico."""
    wb_formulas = openpyxl.load_workbook(archivo, data_only=False)
    wb_valores = openpyxl.load_workbook(archivo, data_only=True)
    if hoja not in wb_valores.sheetnames:
        raise ValueError(f"La hoja '{hoja}' no existe en el archivo. Hojas disponibles: {wb_valores.sheetnames}")

    ws_f = wb_formulas[hoja]
    ws_v = wb_valores[hoja]

    # -- ventas por plato (una fila por Codigo, aunque se repita) --
    ventas: dict[str, dict] = {}
    for r in range(VENTAS_DESDE, VENTAS_HASTA + 1):
        codigo = ws_v.cell(row=r, column=3).value
        if not codigo:
            continue
        codigo = str(codigo).strip()
        if codigo in ventas:
            continue
        item = ws_v.cell(row=r, column=4).value
        vendido = ws_v.cell(row=r, column=5).value
        ventas[codigo] = {"codigo": codigo, "nombre": str(item or codigo).strip(), "cantidad": _num(vendido)}

    # -- insumos: stock informado + mermas + entregas --
    insumos = []
    for r in BLOQUE_INSUMOS:
        nombre = ws_v.cell(row=r, column=4).value
        if not nombre or str(nombre).strip().upper() == "PRODUCTOS":
            continue

        merma_ref = _ref_fila(ws_f.cell(row=r, column=10).value)  # J
        entrega_ref = _ref_fila(ws_f.cell(row=r, column=6).value)  # F
        if merma_ref is None and entrega_ref is None:
            continue

        stock_informado = None
        desglose = {}
        if merma_ref:
            stock_informado = ws_v.cell(row=merma_ref, column=5).value
            for etiqueta, col in MERMA_COLUMNAS:
                v = ws_v.cell(row=merma_ref, column=col).value
                if v:
                    desglose[etiqueta] = _num(v)

        entrega_cantidad = _num(ws_v.cell(row=entrega_ref, column=5).value) if entrega_ref else 0.0

        insumos.append({
            "nombre": str(nombre).strip(),
            "stock_informado": _num(stock_informado) if stock_informado is not None else None,
            "mermas_desglose": desglose,
            "entrega_cantidad": entrega_cantidad,
        })

    return {"ventas": list(ventas.values()), "insumos": insumos}
