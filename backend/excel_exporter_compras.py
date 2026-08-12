# -*- coding: utf-8 -*-
"""Exporta la Planilla de Compras de la web al Excel real "PLANILLA DE COMPRAS
OFICIAL" -- usa la plantilla real (backend/templates/planilla_compras_meses/,
un archivo POR MES, limpio de datos pero con TODAS las formulas originales
intactas, incluido el desglose por Tipo y por columna-de-proveedor arrastrado
hasta el final de la hoja) y solo escribe en las celdas de entrada
(Tipo/Proveedor/N Factura/IVA/Total por fila, mas Venta Periodo/Venta Actual
del resumen). El resto lo calcula el propio Excel al abrirlo.

Un archivo por mes (no un unico workbook de 12 hojas) a propósito -- el
archivo real completo tiene ~650 filas x 106 columnas x 12 hojas con formula
en casi todas las celdas, y openpyxl tarda ~25s solo en CARGARLO (confirmado
midiendo en local) -- eso hacia expirar el request en Render (502 a los
~45s). Cargar solo la hoja del mes pedido baja eso a ~2-3s.

Las columnas y la fila de Total fueron confirmadas leyendo el archivo real mes
por mes (no son un formato inventado) -- Tipo/Proveedor/N Factura/IVA/Total
caen siempre en las mismas columnas (D/E/F/H/I) en los 12 meses, pero la fila
de Total varia por mes (637-657 segun el mes), por eso se ubica en runtime
buscando la celda "Total" en la columna B en vez de asumir una fila fija."""
from io import BytesIO
from pathlib import Path

import openpyxl

_TEMPLATES_DIR = Path(__file__).resolve().parent / "templates" / "planilla_compras_meses"

_MESES = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO", "JULIO", "AGOSTO",
          "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]

_COL_TIPO, _COL_PROVEEDOR, _COL_FACTURA, _COL_SUBTOTAL, _COL_IVA, _COL_TOTAL = 4, 5, 6, 7, 8, 9
_FILA_DATOS_INICIO = 9
_COL_VENTA_PERIODO = 5  # E3
_COL_VENTA_ACTUAL = 4   # D4


def _fila_total(ws) -> int:
    for r in range(_FILA_DATOS_INICIO, ws.max_row + 1):
        v = ws.cell(row=r, column=2).value
        if v and str(v).strip().upper() == "TOTAL":
            return r
    raise ValueError(f"No se encontró la fila 'Total' en la hoja {ws.title}")


def exportar_mes(anio: int, mes: int, items: list[dict], resumen: dict) -> bytes:
    """Genera el Excel real, ya lleno, para un mes.

    items: [{proveedor_nombre, num_factura, subtotal, iva, total, tipo}], en el orden en
    que se quieren escribir (ya vienen ordenados por fecha desde Odoo).
    resumen: {venta_periodo, venta_neta, costo_venta, pct_costo_venta} -- los
    mismos 4 numeros que ya se calculan para la pantalla web (Planilla de
    Compras -- % Costo Venta). Se escriben directo en las celdas del resumen
    (E3/D4/F3/D5/G3/D6) y en el bloque "TOTAL ITEM" (G6/H6/I6, sumando todos
    los items) -- NINGUNA de estas celdas se deja como formula: dependen de
    que Excel recalcule al abrir (row9:row_total, N6/O6, etc.) y varios
    visores no lo hacen automaticamente pese a fullCalcOnLoad, dejando el
    resumen entero en blanco (bug real reportado por el usuario).
    """
    hoja = _MESES[mes - 1]
    wb = openpyxl.load_workbook(_TEMPLATES_DIR / f"{hoja.lower()}.xlsx")
    ws = wb[hoja]

    fila_total = _fila_total(ws)
    filas_disponibles = fila_total - _FILA_DATOS_INICIO
    if len(items) > filas_disponibles:
        raise ValueError(
            f"{hoja} {anio} tiene {len(items)} facturas pero la plantilla solo tiene "
            f"{filas_disponibles} filas disponibles antes de la fila de Total")

    for i, it in enumerate(items):
        fila = _FILA_DATOS_INICIO + i
        ws.cell(row=fila, column=_COL_TIPO).value = it.get("tipo")
        ws.cell(row=fila, column=_COL_PROVEEDOR).value = it.get("proveedor_nombre")
        ws.cell(row=fila, column=_COL_FACTURA).value = it.get("num_factura")
        ws.cell(row=fila, column=_COL_SUBTOTAL).value = it.get("subtotal")
        ws.cell(row=fila, column=_COL_IVA).value = it.get("iva")
        ws.cell(row=fila, column=_COL_TOTAL).value = it.get("total")

    venta_periodo = resumen.get("venta_periodo")
    if venta_periodo is not None:
        ws.cell(row=3, column=_COL_VENTA_PERIODO).value = venta_periodo  # E3 -- VENTA PERIODO
        ws.cell(row=4, column=_COL_VENTA_ACTUAL).value = venta_periodo   # D4 -- VENTA ACTUAL

    venta_neta = resumen.get("venta_neta")
    costo_venta = resumen.get("costo_venta")
    pct_costo_venta = resumen.get("pct_costo_venta")
    if venta_neta is not None:
        ws.cell(row=3, column=6).value = venta_neta      # F3 -- NETO (=E3/1.19)
    if costo_venta is not None:
        ws.cell(row=5, column=4).value = costo_venta     # D5 -- COSTO VENTA (=N6+O6)
    if pct_costo_venta is not None:
        ws.cell(row=3, column=7).value = pct_costo_venta  # G3 -- % COSTO VENTA
        ws.cell(row=6, column=4).value = pct_costo_venta  # D6 -- % COSTO VENTA

    # TOTAL ITEM (G6/H6/I6) -- suma de TODAS las facturas del mes (no solo
    # AL+BA como Costo Venta), misma cifra que la fila "Total" de la tabla.
    ws.cell(row=6, column=7).value = sum(it.get("subtotal") or 0 for it in items)  # G6
    ws.cell(row=6, column=8).value = sum(it.get("iva") or 0 for it in items)       # H6
    ws.cell(row=6, column=9).value = sum(it.get("total") or 0 for it in items)     # I6

    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()
