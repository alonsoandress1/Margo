#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
app.py — Dashboard de Pronóstico · Margo / Nelí
Ejecutar: streamlit run app.py
"""

import os, io, json, html as _html, warnings
from datetime import datetime, timedelta
from collections import defaultdict

warnings.filterwarnings("ignore")

# ── Config persistente ────────────────────────────────────────────────────────

_BASE_DIR    = os.path.dirname(os.path.abspath(__file__))
_CONFIG_PATH = os.path.join(_BASE_DIR, "config.json")
_DEFAULT_FOLDER = os.path.join(_BASE_DIR, "Reportes")

def _load_config() -> dict:
    try:
        with open(_CONFIG_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def _save_config(cfg: dict) -> None:
    with open(_CONFIG_PATH, "w", encoding="utf-8") as f:
        json.dump(cfg, f, indent=2, ensure_ascii=False)

def _get_folder() -> str:
    cfg = _load_config()
    folder = cfg.get("reportes_folder", _DEFAULT_FOLDER)
    return folder if os.path.isdir(folder) else _DEFAULT_FOLDER

def _count_reports(folder: str) -> int:
    if not os.path.isdir(folder):
        return 0
    return sum(1 for f in os.listdir(folder)
               if f.startswith("informe") and f.endswith((".xls", ".xlsx")))

import streamlit as st
import plotly.graph_objects as go
import plotly.express as px
import pandas as pd
import numpy as np

try:
    import xlrd
except ImportError:
    st.error("Falta xlrd. Ejecuta: pip install xlrd==1.2.0")
    st.stop()

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Page Config ──────────────────────────────────────────────────────────────

st.set_page_config(
    page_title="Margo · Nelí — Producción",
    page_icon="🍽️",
    layout="wide",
    initial_sidebar_state="expanded",
)

# ── Design System ─────────────────────────────────────────────────────────────


LOGO_B64 = "iVBORw0KGgoAAAANSUhEUgAAATEAAABhCAYAAAC+nJYrAAARPUlEQVR42u1dX0hk5xU/N7vZbLZZtakuxgiVYkFEggULxZdaCRRECDU+TV6W6UtZFh+KlM2SUtnCtE/TF/HJSqAiLVhjQegUSpHWLEsoEqiVsATZLFkJuplVE2T2T3L6sGe212+/7/6b79+dPT+QGe/cuXO/73733PPnd84BYDAYDAaDwWAwGAwGg8FgMBgMBoPBYDAYDAaDwWAwGAwGg8FgMBgMBuNZxt7dgzM8CwwGI3dAxFFErOFjjPGMMBiMvAivAUScw5OY5ZlhMBh5EGCTKMcozw6DwfBZeF1BNYo8QwwGw1fhVYoQXjuI2MOzpAn7947OIWI3IgY8G+BN1AoRe8mH0t3EY+xExJ7Q31CTCLChCAFW4BWud7IvhqIklTyO4eD+w1OIOIOIK4g4ntdrQeMYQcQNycK/gYgDOR1TDyKOIWIREdcRcRejUUPEy02gGCyGxrSMiBN5vIa+C7BeyQLqzOE4ZoUxVPLEv0HEgJ7cJYzH2P69o3M5Gdeg5NqkwULO76+gfq3YyjE3yTJcy9kYhmWDyJkQG0HE4xQ396Ln4+kWtJBG0J0Dk3gUEQukaRXIupmiv5GD+w9PMbnVzEJ7N483SEJBjIg4lZPzL2e8udc9HY8u4eVl9I605n7Smmspx1IlmsVgUwmTsJS2eCHGoyY6RwJsOmocLuY25flfa/AGX/ZoLJdizvU2Ii6Qb6yPtLVBMqMHaftliUa64pHrpaRZQJeaQqDRQAq2ojLkaI1ETuatP24cPguxmAdJOWx2kGN8S7HvZZfjvLN3eBYRb0WMpYKI/SnnJjzWXQ9M41k0i4W8+DljVW9fzJecRLy2EiyOPo/9KFLhlVHz7PfJH0nYzKpliNFLh/dnIaGJeJuix4uUUrRMPLC0mMhjDtVTF95R2oOI4RyakRXJtllPBbDsXKcTjr0oG7tHghgR8UqDQQ7nD9WYh/06Ig7FOenpWg8RY383qQ/wzt7h2bwIMdWghgz9Xp+KeEfRlFw4xYl/I0OPwqnc42HkrqFgikKQ9VgWYMcKftdwg/Oz5FqImfRBJgzkLPjuz62X4bAaVUPE+SgqgrB53uO5k2kxkxEs6XHPzl/0H21mCbtL8vFmLY5BJsCOddAhXLo3FA+YOtG4V/NDIE6YFSCntAAjF02iaT3xGSFiIDENqp7Om8xHsRt3g/miniuCEb261pEN57BC290ynKoTODSPC4YjntXcBaYUmoRRn1QSPpjvzn0Kw8ea3xSqf4rl7ik3r6LZN9hvgZSLpigQinQrWwGvmk3/tBB5XtHtW7S9CNDwwphKsjB8ZrwTwXBTMoyZFP7Gft98YTo0RJFuYdmCqFh42KIDjtusgyDJZpyV4bMZKRUyukygpCqybzd9Ah/iccT+E2npCy4K4hnwsW1bLCtT0RyxHXUhxOgBGcaGQ1fDptcWkYLeUKXPFiSfDWj4zYmkjHyVsxz8oCTcSlsN0zcfg8SXNG2C9W/JX7Rhk/dn0ioQ70sPsgJq3uaOKhx4ExF+nKKh35xMeNPPeCLAZM7eWpxAUqSIlByOwwiVRpwfQ2t3Rjj3QZtZJIaFmFe+U0llmaIvAkzG0VqCaO7TDQOO8Jorf0cDfqRqVgGvuCdGfCiQp9Fd0G1SiEkETMHA/Fx2kUEiRrs9LQqw6EUZH0S8HnVhaKHUdF08svNXkoaLKf/tQ2HfbU8KNmJWWoLCWVr2YBybpgIGBvxFvSZJtQlyeasGr8tNnx7aigDgpnMhpjCHCpLFMqeLakFCKbEWRj6PNZ+cihGLe6JBn6ALBvimqRw5kXtm2JSsGTj+mCIfMTaAo9k6KoKfxQEyCfHnNHf3FbWfe4cPHv0pvCEIAgQAmbN0PMvN39XRsir56GcxX63JnvKuBFjrmdN/lnz0/uGDR39JehzVvg58H98Th6jx2OEo8n0D5x72oc4bKEMlpvL8GgCODM1VGOJ9+Rn4g7Cr4ZxTTUzhl+pT7NupQ2tQPNkwgdZT8sF/FENsHdBELt5wTK3p1njsFZMmkcn+iIoMjE4bpFNJxNub2vcCb63mlLmPiKtpSKyN+H9Cx7idNnWChFjRh+qoEVUeyhmvQVHHvDbIyEaDfiujZFdTc6ZweUxJglwblh4s/Z5WXd52ponJai3FRaSoDpGIiymcsFezVjhVaHALnpQoyszxiqjecdlRlsaWxmMXTVdAMeXUl5FbQy3cjEfJfU1Nk1CjKj6x86dN1opSdC5KfHEUJtyWJ7mlS5od69ZMSklKy5KpNabb7JAEVwY0BiPEooFl+r0BGw9SX2vPSea85CrPr5jFhJCZHkl8KKSCbzeSg6lQ79GTSqGdJnhIlqo+zOtO6qU1tmC6wYvkhipqmpOrsnVGvzdmg3StoCD1gH+9CgZcnERnI9noWQolRpAFBzXkWXZbTIItmzBpIx4ORQccwYuGxjNkQYjNGbo/piMUgEsWG5uUPBBiHzinOckkfMpFczWNKRrh85nKUDF1x1WEUrcvLGEhv2UHHLFhDVqYSIzeNRXB0qmZq6rvhvsxSoo9TlmuEFz0yJTMHA1+rhHiHgD8RNj806Tfb3vh+a8AQJZuVJDlj9G2XygOl8qf9eAhfg0AMpZ+uw0tDE7ykep4j+ZEB/4o2famhfC1aKIcNXi8PwDAC8K2ksZ5EvGJ6LdtgBM2CgBvCR+9BQBwob3tAf3fInz+TyPC+auvHyk++r2rBjOtZ07/C07yR//hRZK3xgqTnSl8SOWMT4GrLmgWEWMeslDLbdRygKdHM69qy/D5i5pRppC/InNlJUEtuG6DY5uKaATS77iF37AveX5DGY81G+ffIqGzobMWmYJmUXaUmoWWBOWcTSHWwLXpd9GuTRG1vqwjYKPYd9tWSai4jk22+FmSh9MqeEKp2NR8U08lYeY3EsJXOF1XLURzt2xogKrmDDbXRZayMq79N2nycFNkXwwnFCojjooMGNfSIwIM/S4E2IRuRnic6q1wVDdMG5Ac8rqDJ33Nctu6XlsCIOP81HRz5zQI//U4Uy/iQVtKWsI75E97EgCy1QnMZAtF7yoQSzQJHbSAZdUNENGZ+IqBp27VcDRGRkS9BPCEuxYkOQ7t2ylqOvR07yOGeMG2SZl1Lum8VSakVWKmikNIlU966DoGofMuRnTALqdtnEt5vceNtLjLmOpmzD+lyFVGXyoxalEHZTdc/YlksnicJDhRte0LI8FWI//IGvX9W6On/zK9Xwtt+zB03lW6gdZl5YUi0GlBiG1pyPmseNgRu6rI2U2VtSIJIljtUB9jWg5o/J0FXwp2qiIcMyaTVCWlgrXmftHNjxZKHneiX7hiQYiVonyDIRNsLUmbPbDPYdpuYH4nM6aHyRrr9Dp4sCIilrJqgaShTkQcu+xTjqROiT0nCW8bNYdkxRlNcKlU6rRDbFtYH0sywRXSvKvocUdoErC1lPNaSXiDr8Q9ZA7uPzxlq5VghLaE9NlgCoJvMWZsJZ9aiU1baqtuLAdQZsbq7qAdEbFyjU4LD7kNMs3KMq3XGz9JejNXxI246B5pd9NJhJdHrQJF7NCDv0hUkjF6X6I5iEJNZ5VfHSklxyaeEgkmccCCj69b82/cjrioJlAlv9ka+dTmbEWiGjzvEoCXXez7FTfoDuXx9mhqJD3heqxEb6kYWJNLJhPOg6S5fgDwn9Cmt4IgWDJhdgHA24qP3w6C4LcGSkOLKRnfDYLgY11hZQCQlZ7+URAE68L8vgaP03ZeAoA2oWTxl3AyleeYth8AwO3d/aNPX73QWlOZLx3fal0CgDeFj94LgmBCtxDL+FVtc264jHjH7v7RgWquVVVVAeDbEF9e/LTBdKq04/wlAPxKw+F+s//54UwozcqbtkrGynhEOBoXLPr5hg0n3ZYdlEsatWG6ZXxKX4ImhKTbdm7M6JAGOp/hei6b5Jvp8lWVDVM4qpZZ5jdMhH4VzvyKw2socvF2LAixrVDVjo2Ihd/tRb9B8/6lY3zcxbzqTbQueXR9jGghs/RXInP6CvmWO30ewECS5h8Gmb2mk5aXdQsxRW/NZcfXscdyjfoTUUYKcFRsBGxcgqKP0r4J9Yhjs47dZyFWtM2e3rt7cMbWhZVoTKOaFnI4GrfmybUcIi2h01Jlk6Ek1U+cVjOwUy1iMrQuhn03JZtRiE2boh94Mr6rJqJEpMEWEbHotA2V3bm8FZcAHpFj6E2UTrMGfk0ohHgxT6ZkMy3OQtryzzka25jr1m1NNJeVJHSViBLj3pBddXECxebMYu4xm5IMXYtvoxk0AQ/msZSE1xfR+9No5oSLWnE01m5FnbRjXjUMncGLfltNZpt4HifTUHGIhlCg4ErdV/YhIhby6LpQUIRGIojV47xqGAyPqQXPkqlEGpeqhPpqHrhhDAbj//7TuWfNLK+bjSlIoX28WhgMho9CfDeBAHuHZ4rBYOTJL9b0aVYMBqP5aDsLQvXeGSdNMRgMBgMab5bMYDAYDAaDwWAwGAwGg8FgMBgMBoPBYDAYDAaDwWAwGAwGg8FgMMDPxHhOdmcwoAlbuj0j43zHdGn753g5MRjZNAzZ+yjc2Ts8Wy8UGQQBxh03qeALCwiVsNBdKTeFEH4ZAKCro+V1XjUMhn9NpBepEOJUvaOR5DvjiHhT+F5JJQTos4KiOkZFzMUkc22R6rhtyvpEUEL6DJ1LuNvUeujzW+L2iHl4V2iIMierikz7hVFsxmZDDEZeBFi4OusH9BfGdYXmVUeVhE0tXNk1LJRCvUGrkt9flDVfUVSTHRH2mZMI3y1Zc+fQ+3nFPIQFclkQZkMyUzKEK7ySGAw3Amxc1oaO+kjWNZt3RbMtdMO/IxxvSSYoQo11K6LZVtcCRU1M6LA0rjj/WZmgRcSdevls4byl5bND7eZuKhpez0vmoKyrKTWDwcguxK5FNAO+JLuxQ1rYzahu6YrtH0SYsj2Kzu63EpjB4wrtTty+rji3qqoVX72iLSgc+6YaNgOwY5/BSILXAQB294+2xA+CIJijtyc0pK6Olh/Q2+sKp/p/Qe18b5PsX6PXE0Ls7sEXe/R2JeL8zwIAHD549FfZ/b+7f/R38VQUx/lm/bfIh7ZOf6sA8AoAgKo5zN2DL45MteFjIcZgxKMLAODVC621iH32hP/rVVzX2l54/ivJ/h+L0cTWM6dfobc7kv0/o9ePFL//74hza4saV1dHi9gf9CknvSCcvg8ALQDwQxKqb9D2vz1/7sX7wlerAADtbee7FPPAQozBsICPElAtRGf8Nr2qeksOAgB8o/XcwyeC4vPDPUHrCmOMXk+Ycu1t579Db48izv8AAEAiRFQa111RwLa3nX+Z3n6yu3/0IgC8FjxGD70Gdw++mJD8xssxgpTBYIB5n1g9Argq+awS4cdCmQ+JaBGIiDVI5yt7qtVb6NwuJfCJdSoirt0yX5nkOHUaxjA87f+ryiK09a7wJnuQsibGYMQgCIL6zflGmIOFiGUA+DH9e0Hy1d/R63zoOyOtZ04/on9l3LJP6v07Q477SujzL4X9HwgaD6h8YqIWF9ouOt1fUhynTK/v1zXQg/sPT3V1tKzCY3/ZLZUZ3N52/ucmnfsMBiO5NqZCRea4rkf6JLim+J1Bxf51LWhU2L/esPdiAk2sT6YlSfhdi4iIivHMKc5vQ/HbU0In9G5eTQyGGyHWjYi9iDhJplOF3o+rzEmAJ9yvMRIkSyQ4BqNSgUj7mkLE6/S9Yfr9QVWvyygmPJ1Dv/h7dNw+RdL2UMRcjFAGwCad31DM3F2k+SqbilAyGIwI7N87OkfM9IuSG3SeyZwMBsN3LWxESM0p0LZbKkc8g8Fg+CbIeqN8QWwmuUPAU8BgJMOdvcOzXR0tvQDQDo8jeDtBEGzzzDAYDAaDwWAwnk38DySSbTlImmm6AAAAAElFTkSuQmCC"
GOLD     = "#C9A97A"
GOLD_2   = "#E0C090"
GOLD_DIM = "rgba(201,169,122,0.12)"
BG       = "#09080C"
BG_1     = "#0E0D12"
BG_2     = "#141219"
BG_3     = "#1B1920"
BG_4     = "#232128"
BORDER   = "rgba(255,255,255,0.06)"
TEXT     = "rgba(255,255,255,0.95)"
MUTED    = "rgba(255,255,255,0.62)"
SUBTLE   = "rgba(255,255,255,0.32)"

CAT_HEX = {
    "Fondo":    "#C9A97A",
    "Entrada":  "#7EB8F7",
    "Ensalada": "#5CE8D4",
    "Postre":   "#F7A8D0",
    "Niños":    "#C4B5FD",
    "Barra":    "#A78BFA",
}

PLOTLY_BASE = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(20,18,25,0.6)",
    font=dict(family="'Segoe UI','Segoe UI Light',system-ui,sans-serif", color=TEXT, size=12),
    margin=dict(l=16, r=16, t=48, b=16),
    xaxis=dict(gridcolor="rgba(255,255,255,0.04)", linecolor="rgba(255,255,255,0.06)",
               zerolinecolor="rgba(255,255,255,0.04)", tickfont=dict(size=11)),
    yaxis=dict(gridcolor="rgba(255,255,255,0.04)", linecolor="rgba(255,255,255,0.06)",
               zerolinecolor="rgba(255,255,255,0.04)", tickfont=dict(size=11)),
    hoverlabel=dict(bgcolor=BG_3, bordercolor="rgba(201,169,122,0.4)",
                    font=dict(family="Segoe UI", color=TEXT, size=12)),
    title=dict(font=dict(family="'Palatino Linotype','Palatino','Book Antiqua',Georgia,serif",
                         size=20, color=TEXT), x=0, xref="paper", pad=dict(b=12)),
    legend=dict(bgcolor="rgba(0,0,0,0)", bordercolor=BORDER,
                borderwidth=1, font=dict(size=11)),
    colorway=[GOLD, "#7EB8F7", "#5CE8D4", "#F7A8D0", "#C4B5FD", "#E0C090", "#A5D6A7"],
)

# ── Utility helpers (module-level — no redefinir en cada rerun) ───────────────

def _fmt_clp(v):
    if v >= 1e9: return f"${v/1e9:.2f}MM"
    if v >= 1e6: return f"${v/1e6:.1f}M"
    return f"${v/1e3:.0f}K"

def _delta_arrow(v):
    return ("▲" if v >= 0 else "▼"), ("#5CE8D4" if v >= 0 else "#F7A8D0")

def _pct_badge(v):
    c = '#5CE8D4' if v >= 0 else '#F7A8D0'
    s = '+' if v >= 0 else ''
    return f'<span style="color:{c};font-weight:600">{s}{v:.1f}%</span>'

def _fmt_q(v):
    if v >= 1000: return f"{v/1000:.1f}k"
    return f"{v:.0f}"

def cov_color(pct, is_current=False):
    border = f'border:2px solid {GOLD};' if is_current else 'border:1px solid rgba(255,255,255,.06);'
    if pct >= 90:   bg, color = 'rgba(100,230,170,.12)', 'rgba(100,230,170,.9)'
    elif pct >= 75: bg, color = 'rgba(245,158,11,.10)',  '#F59E0B'
    else:           bg, color = 'rgba(215,75,65,.10)',   'rgba(215,75,65,.9)'
    return (f'background:{bg};color:{color};{border}'
            f'border-radius:6px;padding:5px 8px;text-align:center;'
            f'font-family:var(--font-mono);font-size:13px;font-weight:600')

def _diag_card(title, body, border_color):
    st.markdown(
        f'<div style="border-left:3px solid {border_color};padding:9px 13px;'
        f'margin-bottom:9px;background:rgba(255,255,255,0.02);border-radius:0 6px 6px 0">'
        f'<div style="font-size:10px;letter-spacing:.09em;color:rgba(255,255,255,0.32);'
        f'margin-bottom:3px">{title}</div>'
        f'<div style="font-size:13px;color:rgba(255,255,255,0.92)">{body}</div>'
        f'</div>', unsafe_allow_html=True)

def _csum(df, common, anio, cat, campo):
    m = df[(df['año']==anio) & (df['mes'].isin(common)) & (df['categoria']==cat)]
    return float(m[campo].sum())

def _cat_totals(vc, cats, anio, meses):
    sub = vc[(vc['año'] == anio) & (vc['mes'].isin(meses))]
    agg = sub.groupby('categoria')['bruto'].sum()
    return [agg.get(c, 0) for c in cats]

def _mv_cat(src_cat, campo_cat, cat, anio, meses):
    sub = src_cat[(src_cat['año']==anio) & (src_cat['mes'].isin(meses))
                  & (src_cat['categoria']==cat)]
    return {int(r['mes']): float(r[campo_cat]) for _, r in sub.iterrows()}

# ── CSS Injection ─────────────────────────────────────────────────────────────

st.markdown("""
<style>
/* ═══════════════════════════════════════════════════════
   MARGO · NELÍ — Design System v4
   Premium Motion · Management Edition
   ═══════════════════════════════════════════════════════ */

@import url('https://fonts.googleapis.com/css2?family=Playfair+Display:ital,wght@0,300;0,400;0,600;1,300;1,400;1,600&family=Inter:wght@300;400;500;600&family=JetBrains+Mono:wght@300;400;500&display=swap');

:root {
  --font-serif: 'Playfair Display','Palatino Linotype','Book Antiqua',Georgia,serif;
  --font-sans:  'Inter','Segoe UI',system-ui,-apple-system,sans-serif;
  --font-mono:  'JetBrains Mono','Consolas','Cascadia Code',monospace;

  --bg:  #09080C; --s1: #0E0D12; --s2: #141219; --s3: #1B1920; --s4: #232128;

  --gold:    #C9A97A;
  --gold-hi: #E0C090;
  --g80: rgba(201,169,122,.8);
  --g50: rgba(201,169,122,.5);
  --g25: rgba(201,169,122,.25);
  --g12: rgba(201,169,122,.12);
  --g06: rgba(201,169,122,.06);
  --g03: rgba(201,169,122,.03);

  --glow-xs: 0 0 10px rgba(201,169,122,.2);
  --glow-sm: 0 0 20px rgba(201,169,122,.28), 0 0 40px rgba(201,169,122,.1);
  --glow-md: 0 0 32px rgba(201,169,122,.36), 0 0 70px rgba(201,169,122,.12);

  --t1: rgba(255,255,255,.95);
  --t2: rgba(255,255,255,.62);
  --t3: rgba(255,255,255,.32);
  --t4: rgba(255,255,255,.16);

  --b1: rgba(255,255,255,.06);
  --b2: rgba(255,255,255,.11);
  --b3: rgba(255,255,255,.20);

  --danger:    rgba(215,75,65,.9);
  --danger-bg: rgba(215,75,65,.08);

  --ease:   cubic-bezier(.4,0,.2,1);
  --spring: cubic-bezier(.34,1.56,.64,1);
  --out:    cubic-bezier(0,0,.2,1);
}

/* ── Animations ─────────────────────────────────────── */
@keyframes appear     { from{opacity:0;transform:translateY(14px)} to{opacity:1;transform:none} }
@keyframes float      { 0%,100%{transform:translateY(0)} 50%{transform:translateY(-5px)} }
@keyframes shimmer    { from{transform:translateX(-100%)} to{transform:translateX(100%)} }
@keyframes pulse-glow { 0%,100%{box-shadow:0 0 6px rgba(201,169,122,.2)} 50%{box-shadow:0 0 18px rgba(201,169,122,.55),0 0 40px rgba(201,169,122,.18)} }
@keyframes bar-grow   { from{transform:scaleX(0)} to{transform:scaleX(1)} }
@keyframes pulse      { 0%,100%{opacity:1;transform:scale(1)} 50%{opacity:.6;transform:scale(.85)} }

/* ── Base ─────────────────────────────────────────────── */
.stApp { background: var(--bg) !important; }
html,body { font-family: var(--font-sans) !important; font-size: 14px; font-weight: 300; line-height: 1.65; -webkit-font-smoothing: antialiased; font-feature-settings: "kern" 1; }
p,li,span,div { font-family: var(--font-sans) !important; }
* { box-sizing: border-box; }

/* Noise texture overlay */
.stApp::after {
  content:''; position:fixed; inset:0; pointer-events:none; z-index:9998; opacity:.028;
  background-image:url("data:image/svg+xml,%3Csvg viewBox='0 0 200 200' xmlns='http://www.w3.org/2000/svg'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='0.85' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='100%25' height='100%25' filter='url(%23n)'/%3E%3C/svg%3E");
  background-size:180px 180px;
}

/* Subtle radial glow top-center */
.stApp::before {
  content:''; position:fixed; top:0; left:0; right:0; height:500px; pointer-events:none; z-index:0;
  background: radial-gradient(ellipse 80% 60% at 50% -10%, rgba(201,169,122,.05) 0%, transparent 70%);
}

/* ── Layout ───────────────────────────────────────────── */
.main .block-container { padding: 1.5rem 2rem 5rem !important; max-width: 1440px !important; animation: appear .5s var(--ease) both; }
section[data-testid="stSidebar"] {
  background: linear-gradient(160deg, var(--s1) 0%, rgba(14,13,18,.97) 100%) !important;
  border-right: 1px solid var(--b1) !important;
}
section[data-testid="stSidebar"] > div { padding: 1.5rem 1rem !important; }

/* ── Headings ─────────────────────────────────────────── */
h1,h2,h3 { font-family: var(--font-serif) !important; font-weight: 300 !important; color: var(--t1) !important; letter-spacing: -.4px !important; }

/* ── Sidebar labels ───────────────────────────────────── */
section[data-testid="stSidebar"] label { font-size: 10px !important; font-weight: 700 !important; letter-spacing: .18em !important; text-transform: uppercase !important; color: var(--t3) !important; }
section[data-testid="stSidebar"] p { color: var(--t2) !important; font-size: 12px !important; }

/* ── KPI / Metric cards ───────────────────────────────── */
[data-testid="metric-container"] {
  background: linear-gradient(135deg, var(--s3) 0%, var(--s2) 100%) !important;
  border: 1px solid var(--b1) !important;
  border-radius: 12px !important;
  padding: 26px 20px !important;
  position: relative; overflow: hidden;
  transition: transform .22s var(--spring), box-shadow .2s, border-color .2s;
  animation: appear .4s var(--ease) both;
}
/* Permanent subtle top accent line (MargoStats pattern) */
[data-testid="metric-container"]::before {
  content: ''; position: absolute; top: 0; left: 0; right: 0; height: 1px;
  background: linear-gradient(90deg, transparent, rgba(201,169,122,.16), transparent);
}
[data-testid="metric-container"]:hover {
  transform: translateY(-4px) !important;
  border-color: rgba(201,169,122,.18) !important;
  box-shadow: 0 18px 52px rgba(0,0,0,.65), var(--glow-sm) !important;
}
[data-testid="stMetricLabel"] { font-family: var(--font-sans) !important; font-size: 8.5px !important; font-weight: 600 !important; letter-spacing: .24em !important; text-transform: uppercase !important; color: var(--t3) !important; margin-bottom: 4px !important; }
[data-testid="stMetricValue"] { font-family: var(--font-mono) !important; font-size: 42px !important; font-weight: 300 !important; letter-spacing: -2px !important; line-height: 1 !important;
  background: linear-gradient(135deg, var(--gold-hi) 0%, var(--gold) 55%, rgba(201,169,122,.8) 100%) !important;
  -webkit-background-clip: text !important; -webkit-text-fill-color: transparent !important; background-clip: text !important;
}
[data-testid="stMetricDelta"] > div { font-family: var(--font-sans) !important; font-size: 11px !important; color: var(--t2) !important; letter-spacing: .03em !important; }

/* ── Tabs ─────────────────────────────────────────────── */
.stTabs [data-baseweb="tab-list"] { background: transparent !important; border-bottom: 1px solid var(--b1) !important; gap: 0 !important; }
.stTabs [data-baseweb="tab"] {
  font-family: var(--font-sans) !important; font-size: 11px !important; font-weight: 600 !important;
  letter-spacing: .12em !important; text-transform: uppercase !important;
  color: var(--t3) !important; background: transparent !important;
  border-bottom: 2px solid transparent !important; padding: 12px 22px !important;
  transition: color .2s, border-color .2s !important;
}
.stTabs [data-baseweb="tab"]:hover { color: var(--t2) !important; }
.stTabs [aria-selected="true"] { color: var(--gold) !important; border-bottom-color: var(--gold) !important; text-shadow: var(--glow-xs) !important; }
.stTabs [data-baseweb="tab-panel"] { padding-top: 1.5rem !important; }

/* ── Buttons ──────────────────────────────────────────── */
.stButton>button {
  background: var(--g06) !important; color: var(--gold) !important;
  border: 1px solid var(--b3) !important; border-radius: 8px !important;
  font-family: var(--font-sans) !important; font-weight: 500 !important;
  font-size: 12px !important; letter-spacing: .08em !important;
  transition: all .2s var(--ease) !important; position: relative; overflow: hidden;
}
.stButton>button:hover { background: var(--g12) !important; border-color: var(--gold) !important; box-shadow: var(--glow-xs) !important; }
.stDownloadButton>button { background: var(--gold) !important; color: #09080C !important; border: none !important; font-weight: 700 !important; }
.stDownloadButton>button:hover { opacity: .9 !important; box-shadow: var(--glow-sm) !important; }

/* ── Inputs ───────────────────────────────────────────── */
.stSlider [data-baseweb="slider"] [role="slider"] { background: var(--gold) !important; border-color: var(--gold) !important; box-shadow: var(--glow-xs) !important; }
.stSlider [data-baseweb="slider"] div[data-testid="stThumbValue"] { color: var(--gold) !important; font-family: var(--font-mono) !important; }
/* Slider track fill color */
.stSlider [data-baseweb="slider"] [data-testid="stSlider"] div[role="progressbar"],
.stSlider div[data-baseweb="slider"] > div > div:nth-child(2) { background: var(--gold) !important; }
[data-baseweb="slider"] [data-testid="stSliderTickBarMin"],
[data-baseweb="slider"] [data-testid="stSliderTickBarMax"] { color: var(--t3) !important; font-family: var(--font-mono) !important; font-size:10px !important; }
/* Override red accent color from Streamlit theme */
:root { --primary-color: #C9A97A !important; }
[data-testid="stSlider"] > div > div > div > div { background: rgba(201,169,122,.18) !important; }
[data-testid="stSlider"] > div > div > div > div > div { background: var(--gold) !important; }
.stDateInput input { background: var(--s2) !important; border-color: var(--b2) !important; color: var(--t1) !important; border-radius: 8px !important; font-family: var(--font-mono) !important; }
[data-baseweb="select"] > div { background: var(--s2) !important; border-color: var(--b2) !important; }
[data-baseweb="menu"] { background: var(--s2) !important; border-color: var(--b2) !important; }
[data-baseweb="option"] { background: var(--s2) !important; color: var(--t1) !important; }
[data-baseweb="option"]:hover { background: var(--s3) !important; }

/* ── Divider ──────────────────────────────────────────── */
hr { border-color: var(--b1) !important; margin: 1rem 0 !important; }

/* ── Scrollbar ────────────────────────────────────────── */
::-webkit-scrollbar { width: 4px; height: 4px; }
::-webkit-scrollbar-track { background: var(--bg); }
::-webkit-scrollbar-thumb { background: var(--s4); border-radius: 2px; }
::-webkit-scrollbar-thumb:hover { background: var(--gold); }

/* ────────────────────────────────────────────────────────
   CUSTOM COMPONENTS
   ──────────────────────────────────────────────────────── */

/* ── Sidebar logo ─────────────────────────────────────── */
.sb-logo-wrap {
  padding: 28px 20px 14px;
  border-bottom: 1px solid var(--b1);
  margin-bottom: 8px;
  background: transparent;
}
.sb-logo-img {
  width: 115px; display: block;
  filter: drop-shadow(0 0 12px rgba(201,169,122,.3));
  margin-bottom: 8px;
}
.sb-logo-tagline {
  font-family: var(--font-sans); font-size: 8px; font-weight: 500;
  letter-spacing: .22em; text-transform: uppercase;
  color: rgba(201,169,122,.55);
}

/* ── Brand header ─────────────────────────────────────── */
.brand-header {
  background: linear-gradient(170deg, var(--s1) 0%, rgba(14,13,18,.8) 100%);
  border: 1px solid var(--b1); border-radius: 0;
  padding: 44px 52px 36px; margin-bottom: 24px; position: relative; overflow: hidden;
  animation: appear .5s var(--ease) both;
}
.brand-header::before {
  content:''; position:absolute; inset:0; pointer-events:none;
  background: radial-gradient(ellipse 60% 140% at 80% -10%, rgba(201,169,122,.09) 0%, transparent 60%);
}
.brand-header::after {
  content: 'MARGO'; position: absolute; right: -20px; bottom: -30px;
  font-family: var(--font-serif); font-size: 160px; font-weight: 600;
  color: rgba(201,169,122,.025); line-height: 1; letter-spacing: -6px;
  user-select: none; pointer-events: none;
}
.bh-inner { position: relative; z-index: 1; display: flex; justify-content: space-between; align-items: flex-end; gap: 24px; }
.bh-brand {
  font-family: var(--font-sans); font-size: 8.5px; font-weight: 600; letter-spacing: .24em; text-transform: uppercase;
  color: var(--gold); margin-bottom: 12px; display: flex; align-items: center; gap: 10px;
  text-shadow: 0 0 14px rgba(201,169,122,.4);
}
.bh-brand::before { content:''; display:inline-block; width:20px; height:1px; background:linear-gradient(90deg,transparent,var(--gold)); }
.bh-brand-line { display:none; }
.bh-title {
  font-family: var(--font-serif); font-size: clamp(28px,3.5vw,48px); font-weight: 300;
  line-height: 1; letter-spacing: .02em; margin: 0; font-style: italic;
  color: var(--t1);
}
.bh-title strong {
  font-style: normal; font-weight: 400;
  background: linear-gradient(135deg, var(--gold-hi) 0%, var(--gold) 55%, rgba(201,169,122,.7) 100%);
  -webkit-background-clip: text; -webkit-text-fill-color: transparent; background-clip: text;
}
.bh-title em { font-style: italic; color: var(--gold-hi); -webkit-text-fill-color: var(--gold-hi); }
.bh-period { margin-top: 14px; font-size: 11px; color: var(--t3); letter-spacing: .06em; display: flex; align-items: center; gap: 8px; flex-wrap: wrap; }
.bh-sep { color: var(--s4); }
.bh-right { text-align: right; flex-shrink: 0; }
.bh-date { font-family: var(--font-serif); font-style:italic; font-size: 13px; color: var(--t2); }
.bh-meta { font-family: var(--font-mono); font-size: 10px; color: var(--t3); margin-top: 4px; letter-spacing: .05em; }

/* ── Promo chip ───────────────────────────────────────── */
.promo-chip { display:inline-flex; align-items:center; gap:5px; background:var(--g06); border:1px solid var(--g25); color:var(--gold); font-size:9px; font-weight:700; letter-spacing:.15em; text-transform:uppercase; padding:3px 10px; border-radius:20px; animation:pulse-glow 2.5s ease-in-out infinite; }

/* ── Day cards ────────────────────────────────────────── */
.day-card {
  background: linear-gradient(160deg, var(--s3) 0%, var(--s2) 100%);
  border: 1px solid var(--b1); border-radius: 16px;
  padding: 16px 14px 14px; height: 100%;
  display: flex; flex-direction: column; align-items: center;
  transition: transform .25s var(--ease), box-shadow .25s var(--ease), border-color .25s;
  animation: appear .4s var(--ease) both;
  position: relative; overflow: hidden;
}
.day-card:hover { transform: translateY(-5px); box-shadow: 0 24px 60px rgba(0,0,0,.65), var(--glow-xs); border-color: rgba(255,255,255,.11); }
.day-card.promo { border-color: var(--g50); background: linear-gradient(160deg, #1a1510 0%, #120e08 100%); box-shadow: 0 0 0 1px rgba(201,169,122,.12) inset; }
.day-card.promo::before { content:''; position:absolute; top:0; left:0; right:0; height:2px; background:linear-gradient(90deg,transparent,var(--gold),transparent); }
/* arc animation */
@keyframes arc-draw { from{stroke-dashoffset:var(--arc-fill,189)} to{stroke-dashoffset:0} }
.dc-arc-fill { animation: arc-draw .9s cubic-bezier(.4,0,.2,1) both; }
/* card sub-elements */
.dc-hdr { width:100%; display:flex; justify-content:space-between; align-items:baseline; margin-bottom:4px; }
.dc-dow { font-size:9px; font-weight:700; letter-spacing:.22em; text-transform:uppercase; color:var(--t3); }
.dc-cal { font-family:var(--font-mono); font-size:12px; color:var(--t2); }
.dc-mon { font-size:9px; color:var(--t3); margin-left:2px; }
.dc-arc-wrap { position:relative; display:flex; align-items:center; justify-content:center; width:136px; height:80px; margin:2px 0; }
.dc-svg { position:absolute; top:0; left:0; overflow:visible; }
.dc-hero { position:absolute; bottom:2px; display:flex; flex-direction:column; align-items:center; gap:2px; }
.dc-big { font-family:var(--font-mono); font-size:36px; font-weight:300; letter-spacing:-2px; line-height:1; }
.dc-lbl { font-size:9px; font-weight:600; letter-spacing:.18em; text-transform:uppercase; color:var(--t3); }
.dc-total { margin-top:6px; font-family:var(--font-mono); font-size:13px; color:var(--t2); }
.dc-total-lbl { font-size:10px; color:var(--t3); margin-left:2px; }
.dc-chips { display:flex; flex-wrap:wrap; gap:4px; margin-top:8px; justify-content:center; width:100%; }
.dc-feriado-chip { display:inline-flex;align-items:center;gap:3px;background:rgba(245,158,11,.1);border:1px solid rgba(245,158,11,.3);color:#F59E0B;font-size:8px;font-weight:700;letter-spacing:.1em;text-transform:uppercase;padding:2px 7px;border-radius:20px; }
.dc-warn-chip { display:inline-flex;align-items:center;gap:3px;background:rgba(215,75,65,.08);border:1px solid rgba(215,75,65,.25);color:rgba(215,75,65,.9);font-size:8px;font-weight:700;letter-spacing:.1em;padding:2px 7px;border-radius:20px; }

/* ── Model state card ─────────────────────────────────── */
.kpi-model { background:linear-gradient(135deg,var(--s3),var(--s2)); border:1px solid var(--b1); border-radius:12px; padding:18px 20px; margin-top:12px; transition:border-color .25s; }
.kpi-model:hover { border-color: var(--g25); }
.kpi-model .km-label { font-size:9px; font-weight:700; letter-spacing:.2em; text-transform:uppercase; color:var(--t3); margin-bottom:6px; }
.kpi-model .km-val { font-family:var(--font-mono); font-size:13px; color:var(--t1); }
.kpi-model .km-sub { font-size:11px; color:var(--t3); margin-top:3px; }

/* ── Production table ─────────────────────────────────── */
.prod-wrap { background:linear-gradient(135deg,var(--s3),var(--s2)); border:1px solid var(--b1); border-radius:14px; overflow:hidden; }
.prod-table { width:100%; border-collapse:collapse; font-family:var(--font-sans); }
.prod-table thead tr { background:var(--s4); }
.prod-table th { padding:11px 16px; font-size:9px; font-weight:700; letter-spacing:.18em; text-transform:uppercase; color:var(--t3); text-align:left; border-bottom:1px solid var(--b1); white-space:nowrap; }
.prod-table th.r { text-align:right; }
.prod-table th.bh { width:160px; }
.prod-table td { padding:10px 16px; font-size:13px; border-bottom:1px solid rgba(255,255,255,.025); vertical-align:middle; color:var(--t2); }
.prod-table tr:hover td { background:var(--g03); }
.prod-table tr:last-child td { border-bottom:none; }
.td-name { font-weight:400; color:var(--t1); }
.td-name.r1 { font-weight:600; }
.td-name.r2 { font-weight:500; }
.td-name.dim { color:var(--t3); }
.td-num { text-align:right; font-family:var(--font-mono); font-size:12px; white-space:nowrap; }
.td-mu { color:var(--t3); }
.td-buf { color:rgba(100,230,170,.85); }
.td-total { font-weight:700; font-size:14px; color:var(--gold); }
.td-bar { width:160px; padding-right:20px !important; }
.bar-track { background:rgba(255,255,255,.04); border-radius:8px; height:8px; overflow:hidden; }
.bar-fill { height:100%; border-radius:8px; transform-origin:left; animation:bar-grow .6s var(--out) both; opacity:.88; }

/* ── Section headers ──────────────────────────────────── */
.sec-hd { display:flex; align-items:baseline; gap:14px; padding-bottom:14px; margin-bottom:20px; border-bottom:1px solid var(--b1); }
.sec-n { font-family:var(--font-mono); font-size:11px; color:var(--t3); min-width:26px; }
.sec-t { font-family:var(--font-serif); font-size:24px; font-weight:400; color:var(--t1); }
.sec-r { flex:1; height:1px; background:linear-gradient(90deg,var(--b2),transparent); margin-left:4px; }
.sec-c { font-size:11px; color:var(--t3); white-space:nowrap; }

/* ── Alert promo ──────────────────────────────────────── */
.alert-promo { background:var(--g06); border:1px solid var(--g25); border-left:3px solid var(--gold); border-radius:0 10px 10px 0; padding:12px 18px; font-size:13px; color:var(--t2); margin-bottom:20px; }
.alert-promo strong { color:var(--gold); text-shadow:var(--glow-xs); }

/* ── Totals strip ─────────────────────────────────────── */
.totals-strip { background:linear-gradient(135deg,var(--s3),var(--s2)); border:1px solid var(--b1); border-radius:12px; padding:20px 28px; display:flex; align-items:center; gap:0; margin-top:20px; }
.ts-item { flex:1; padding:0 20px; }
.ts-item:first-child { padding-left:0; }
.ts-sep { width:1px; height:44px; background:var(--b1); flex-shrink:0; }
.ts-item dt { font-size:8.5px; font-weight:700; letter-spacing:.22em; text-transform:uppercase; color:var(--t3); margin-bottom:5px; }
.ts-item dd { font-family:var(--font-mono); font-size:34px; letter-spacing:-2px; line-height:1; text-shadow:var(--glow-sm); }
.ts-sub { font-size:10px; color:var(--t3); margin-top:4px; letter-spacing:.02em; }

/* ── Confidence badges ────────────────────────────────── */
.conf-badge { display:inline-flex; align-items:center; gap:5px; padding:3px 10px; border-radius:20px; font-size:10px; font-weight:600; letter-spacing:.05em; }
.cb-warn { background:rgba(245,158,11,.1); color:#F59E0B; border:1px solid rgba(245,158,11,.22); }
.cb-mod  { background:rgba(126,184,247,.08); color:#7EB8F7; border:1px solid rgba(126,184,247,.18); }
.cb-ok   { background:rgba(100,230,170,.08); color:rgba(100,230,170,.85); border:1px solid rgba(100,230,170,.18); }
.cb-none { background:var(--g06); color:var(--t3); border:1px solid var(--b2); }

/* (chat CSS removed) */

/* ── Dataframes ───────────────────────────────────────── */
[data-testid="stDataFrame"] { border-radius: 12px; overflow: hidden; border: 1px solid var(--b1) !important; }
.dvn-scroller { background: var(--s2) !important; }

/* ── Alerts ───────────────────────────────────────────── */
.stAlert { border-radius: 10px !important; border: 1px solid !important; background: var(--s2) !important; }

/* ── Sidebar footer brand ─────────────────────────────── */
.sb-footer { margin-top:20px; padding:16px 0 4px; border-top:1px solid var(--b1); }
.sb-footer-name { font-family:var(--font-serif); font-style:italic; font-size:13px; color:rgba(255,255,255,.3); letter-spacing:.02em; margin-bottom:3px; }
.sb-footer-sub { font-family:var(--font-sans); font-size:8px; color:rgba(255,255,255,.2); letter-spacing:.1em; text-transform:uppercase; }

/* ── Headings override (Playfair) ─────────────────────── */
h1,h2,h3,h4,h5,h6 { font-family: var(--font-serif) !important; font-weight: 300 !important; color: var(--t1) !important; }
h5 { font-size: 10px !important; font-weight: 600 !important; letter-spacing: .18em !important; text-transform: uppercase !important; color: var(--t3) !important; font-family: var(--font-sans) !important; }

/* ── Page section title pattern (MargoStats .st) ─────── */
.section-title {
  font-family: var(--font-sans); font-size: 8.5px; font-weight: 600;
  letter-spacing: .24em; text-transform: uppercase;
  color: rgba(255,255,255,.38); margin-bottom: 14px; padding-bottom: 10px;
  border-bottom: 1px solid var(--b1); display: flex; align-items: center; gap: 10px;
}
.section-title::before { content:''; width:14px; height:1px; background:linear-gradient(90deg,transparent,rgba(201,169,122,.35)); flex-shrink:0; }

/* ── Sidebar container overrides ──────────────────────── */
/* Remove any Streamlit-added backgrounds inside sidebar markdown */
section[data-testid="stSidebar"] [data-testid="stMarkdownContainer"],
section[data-testid="stSidebar"] .stMarkdown,
section[data-testid="stSidebar"] .element-container { background: transparent !important; }

/* ── Hide Streamlit chrome ────────────────────────────── */
#MainMenu,footer,[data-testid="stToolbar"],[data-testid="stDecoration"] { visibility:hidden !important; }

/* Fix: sidebar collapse button renders icon name as text */
[data-testid="stSidebarCollapseButton"] { display:none !important; }
button[data-testid="baseButton-header"] { display:none !important; }
section[data-testid="stSidebar"] > div:first-child > div:first-child > button { display:none !important; }

/* ── Tab COMPRAS — ingredient table ──────────────────── */
.ci-name { font-family:var(--font-sans); font-size:13px; color:rgba(255,255,255,.88);
           padding:9px 14px; width:38%; }
.ci-qty  { font-family:var(--font-mono); font-size:13px; font-weight:600;
           padding:9px 10px; white-space:nowrap; width:14%; text-align:right; }
.ci-bar  { padding:9px 10px; width:36%; }
.ci-bar-bg { height:5px; background:rgba(255,255,255,.07); border-radius:3px; overflow:hidden; }
.ci-bar-fill { height:5px; border-radius:3px; transition:width .6s ease; }
.ci-meta { padding:9px 6px; font-size:10px; white-space:nowrap; color:rgba(255,255,255,.35);
           text-transform:uppercase; letter-spacing:.06em; }
.ci-platos { padding:9px 10px; font-size:10px; color:rgba(255,255,255,.28); text-align:right; }
table.ci-table tr:not(:last-child) td { border-bottom:1px solid rgba(255,255,255,.04); }
</style>
""", unsafe_allow_html=True)

# ── Model constants ───────────────────────────────────────────────────────────

FOLDER          = _get_folder()
PROMO_FROM      = datetime(2026, 6, 7)
K_SAFETY        = 0.7
BUFFER_FALLBACK = 0.20
DEFAULT_HORIZON = 3

EXCLUDED = {
    datetime(2026, 5, 18),
    datetime(2026, 5, 21), datetime(2026, 5, 22),
    datetime(2026, 5, 23), datetime(2026, 5, 24),
}

# Feriados chilenos — fuente: Ley 20.215 + decretos
FERIADOS = {
    datetime(2026,  1,  1): "Año Nuevo",
    datetime(2026,  4,  3): "Viernes Santo",
    datetime(2026,  4,  4): "Sábado Santo",
    datetime(2026,  5,  1): "Día del Trabajo",
    datetime(2026,  5, 21): "Glorias Navales",
    datetime(2026,  6, 29): "San Pedro y San Pablo",
    datetime(2026,  7, 16): "Virgen del Carmen",
    datetime(2026,  8, 15): "Asunción de la Virgen",
    datetime(2026,  9, 18): "Fiestas Patrias",
    datetime(2026,  9, 19): "Glorias del Ejército",
    datetime(2026, 10, 12): "Encuentro de Dos Mundos",
    datetime(2026, 10, 31): "Iglesias Evangélicas",
    datetime(2026, 11,  1): "Todos los Santos",
    datetime(2026, 12,  8): "Inmaculada Concepción",
    datetime(2026, 12, 25): "Navidad",
    datetime(2027,  1,  1): "Año Nuevo",
}

PREFIXES = {
    # Cocina
    'VECARFON': 'Fondo',   'VECARFJT': 'Fondo',   'VECARFUB': 'Fondo',
    'VECARENT': 'Entrada',  'VECARAJT': 'Entrada',  'VECARAUB': 'Entrada',
    'VECARENS': 'Ensalada', 'VECAREJT': 'Ensalada', 'VECAREUB': 'Ensalada',
    'VECARPOS': 'Postre',   'VECARPUB': 'Postre',
    'VECARNIN': 'Niños',
    # Barra
    'VELICACO': 'Barra',   'VELICASA': 'Barra',
    'VELICATR': 'Barra',   'VELICDES': 'Barra',
    'VCVINCOP': 'Barra',   'VCVINCAR': 'Barra',
    'VCVINESP': 'Barra',   'VCVINSAU': 'Barra',
}

# Items a excluir: nombres que terminen en " PY" o " JT" (variantes de servicio)
_EXCLUDE_SUFFIXES = (' PY', ' JT')

CAT_ORDER = ['Fondo', 'Entrada', 'Ensalada', 'Postre', 'Niños', 'Barra']
DAYS_ES   = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo']
MONTHS_ES = ['','enero','febrero','marzo','abril','mayo','junio',
             'julio','agosto','septiembre','octubre','noviembre','diciembre']

# ── Core model functions ──────────────────────────────────────────────────────

def _fecha_es(d):
    return f"{d.day} de {MONTHS_ES[d.month]} de {d.year}"

def _confidence_badge(n):
    if n == 0: return '<span class="conf-badge cb-none">Sin datos</span>'
    if n == 1: return '<span class="conf-badge cb-warn">⚠ 1 obs · muy baja</span>'
    if n < 4:  return f'<span class="conf-badge cb-mod">△ {n} obs · moderada</span>'
    return         f'<span class="conf-badge cb-ok">✓ {n} obs · buena</span>'

def day_type(date):
    if date.weekday() == 6 and date >= PROMO_FROM:
        return 'Dom-Promo'
    return DAYS_ES[date.weekday()]

def parse_report(path):
    result = {}
    def handle_row(code, name, raw):
        code = str(code or '').strip()
        name = str(name or '').strip()
        if not code or not name:
            return
        # Excluir variantes PY y JT (son duplicados de servicio)
        name_up = name.upper()
        if any(name_up.endswith(s.upper()) for s in _EXCLUDE_SUFFIXES):
            return
        for prefix, cat in PREFIXES.items():
            if code.startswith(prefix):
                try:
                    qty = float(raw or 0)
                    if qty > 0:
                        result[(code, name, cat)] = qty
                except (ValueError, TypeError):
                    pass
                break
    if path.endswith('.xlsx'):
        if not HAS_OPENPYXL:
            return result
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for row in wb.active.iter_rows(min_row=22, values_only=True):
            if len(row) > 9:
                handle_row(row[0], row[1], row[9])
        wb.close()
    else:
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        if ws.ncols > 9:
            for r in range(22, ws.nrows):
                handle_row(ws.cell_value(r, 0), ws.cell_value(r, 1), ws.cell_value(r, 9))
    return result

@st.cache_data(show_spinner=False)
def load_history_cached(folder: str):
    history = {}
    for fname in sorted(os.listdir(folder)):
        if not fname.startswith('informe'):
            continue
        if not (fname.endswith('.xls') or fname.endswith('.xlsx')):
            continue
        raw = fname.replace('informe ', '').replace('.xlsx','').replace('.xls','').strip()
        try:
            d, m, y = raw.split('-')
            date = datetime(int(y), int(m), int(d))
        except ValueError:
            continue
        if date in EXCLUDED:
            continue
        data = parse_report(os.path.join(folder, fname))
        if data:
            history[date] = data
    return history

def build_model(history):
    accum = defaultdict(lambda: defaultdict(list))
    for date, data in history.items():
        dt = day_type(date)
        for (code, name, cat), qty in data.items():
            accum[dt][(name, cat)].append(qty)
    model = {}
    for dt, dishes in accum.items():
        model[dt] = {}
        for dish, vals in dishes.items():
            n    = len(vals)
            mean = sum(vals) / n
            std  = (sum((v - mean)**2 for v in vals) / n)**0.5 if n > 1 else None
            model[dt][dish] = {'mean': mean, 'std': std, 'n': n}
    return model

def build_name_to_sku(history):
    """Build {dish_name: sku_code} lookup from history for recipe matching."""
    lookup = {}
    for data in history.values():
        for (code, name, cat) in data:
            if name not in lookup:
                lookup[name] = code
    return lookup

def dish_fc(stats, k=K_SAFETY):
    mean, std = stats['mean'], stats['std']
    return mean + k * std if std is not None else mean * (1 + BUFFER_FALLBACK)

def make_forecast(model, start, horizon, k=K_SAFETY):
    day_details, totals, mu_acc, sg_acc = [], defaultdict(float), defaultdict(float), defaultdict(float)
    for i in range(horizon):
        d  = start + timedelta(days=i)
        dt = day_type(d)
        dm = model.get(dt) or model.get('Domingo', {})
        day_details.append((d, dt, dm, dt not in model))
        for dish, stats in dm.items():
            mean   = stats['mean']
            buffer = k * stats['std'] if stats['std'] is not None else mean * BUFFER_FALLBACK
            mu_acc[dish] += mean
            sg_acc[dish] += buffer
            totals[dish] += mean + buffer
    dish_stats = {d: {'mu': mu_acc[d], 'sigma': sg_acc[d]} for d in totals}
    return day_details, dict(totals), dish_stats

@st.cache_data(show_spinner=False)
def build_df(_history):
    rows = []
    for date, data in _history.items():
        dt = day_type(date)
        for (code, name, cat), qty in data.items():
            rows.append({
                'date': date, 'day_type': dt,
                'weekday': DAYS_ES[date.weekday()],
                'is_promo': date.weekday() == 6 and date >= PROMO_FROM,
                'dish': name, 'category': cat, 'quantity': qty,
            })
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df['date'] = pd.to_datetime(df['date'])
    return df

# ── HTML builders ─────────────────────────────────────────────────────────────

def compute_dish_coverage(history, day_fc, model, k):
    """
    Por cada (plato, cat) del pronóstico actual, calcula qué % de días
    históricos del mismo tipo de día habrías cubierto con el factor k.
    """
    accum = {}  # (name, cat) → {'covered': int, 'total': int}

    for _, dt, _, _ in day_fc:
        dm        = model.get(dt, {})
        hist_days = [(d, data) for d, data in history.items() if day_type(d) == dt]
        if not hist_days:
            continue
        for (name, cat), stats in dm.items():
            mean = stats['mean']
            std  = stats['std']
            fc   = mean + k * std if std is not None else mean * (1 + BUFFER_FALLBACK)
            key  = (name, cat)
            if key not in accum:
                accum[key] = {'covered': 0, 'total': 0}
            for _, actual_data in hist_days:
                actual = next(
                    (qty for (_, n, c), qty in actual_data.items() if n == name and c == cat),
                    None,
                )
                if actual is None:
                    continue
                accum[key]['total']   += 1
                if actual <= fc:
                    accum[key]['covered'] += 1

    return {
        key: round(v['covered'] / v['total'] * 100)
        for key, v in accum.items()
        if v['total'] > 0
    }


def _cov_badge(pct):
    """Colored inline badge showing historical coverage %."""
    if pct is None:
        return ''
    if pct >= 90:
        bg, fg = 'rgba(100,230,170,.12)', 'rgba(100,230,170,.9)'
        dot = '●'
    elif pct >= 75:
        bg, fg = 'rgba(245,158,11,.12)', '#F59E0B'
        dot = '●'
    else:
        bg, fg = 'rgba(215,75,65,.12)', 'rgba(215,75,65,.9)'
        dot = '●'
    return (
        f'<span style="display:inline-flex;align-items:center;gap:3px;'
        f'background:{bg};border:1px solid {fg}44;color:{fg};'
        f'font-size:9px;font-weight:700;padding:1px 6px;border-radius:10px;'
        f'margin-left:7px;vertical-align:middle;letter-spacing:.04em">'
        f'{dot} {pct}%</span>'
    )


def prod_table_html(items, dish_stats, color, dish_coverage=None):
    if not items:
        return '<p style="color:rgba(255,255,255,.32);padding:20px">Sin datos para esta categoría.</p>'
    max_v = max(items.values())
    rows  = ''
    for i, ((name, cat), qty) in enumerate(sorted(items.items(), key=lambda x: -x[1])):
        st_d  = dish_stats.get((name, cat), {'mu': qty, 'sigma': 0})
        mu    = round(st_d['mu'])
        sigma = round(st_d['sigma'])
        pct   = round(qty / max_v * 100)
        rank  = {0: 'r1', 1: 'r2', 2: 'r3'}.get(i, 'dim' if i >= 10 else '')
        total_style = f'color:{color}' if i < 3 else ''
        cov_pct     = dish_coverage.get((name, cat)) if dish_coverage is not None else None
        badge       = _cov_badge(cov_pct)
        rows += f"""<tr>
          <td class="td-name {rank}">{_html.escape(name)}{badge}</td>
          <td class="td-num td-mu">{mu}</td>
          <td class="td-num td-buf">+{sigma}</td>
          <td class="td-num td-total" style="{total_style}">{round(qty)}</td>
          <td class="td-bar">
            <div class="bar-track">
              <div class="bar-fill" style="width:{pct}%;background:{color}"></div>
            </div>
          </td>
        </tr>"""
    cov_legend = (
        '<tr><td colspan="5" style="padding:10px 16px 8px;border-top:1px solid rgba(255,255,255,.04)">'
        '<span style="font-size:9px;color:rgba(255,255,255,.25);letter-spacing:.06em">'
        'Badge = cobertura histórica con el factor actual &nbsp;·&nbsp; '
        '<span style="color:rgba(100,230,170,.7)">●</span> ≥90% seguro &nbsp;'
        '<span style="color:#F59E0B">●</span> 75–89% aceptable &nbsp;'
        '<span style="color:rgba(215,75,65,.7)">●</span> &lt;75% riesgo'
        '</span></td></tr>'
    ) if dish_coverage else ''
    return f"""<div class="prod-wrap"><table class="prod-table">
      <thead><tr>
        <th>Plato</th><th class="r">Prom μ</th>
        <th class="r">Buffer σ</th>
        <th class="r" style="color:{color}">Total</th>
        <th class="bh"></th>
      </tr></thead>
      <tbody>{rows}{cov_legend}</tbody>
    </table></div>"""


def day_card_html(d, dt, ff, pf, max_ff, is_fallback=False, idx=0):
    is_p       = dt == 'Dom-Promo'
    feriado    = FERIADOS.get(datetime(d.year, d.month, d.day))
    cls        = 'day-card promo' if is_p else 'day-card'
    bar_c      = GOLD if is_p else '#7EB8F7'
    nc         = GOLD if is_p else TEXT
    card_delay = f'{idx * 0.09:.2f}s'
    arc_delay  = f'{idx * 0.09 + 0.2:.2f}s'
    # SVG semicircle arc: center (68,72), radius 60, length ≈ 188.5
    ARC_LEN  = 188.5
    fill_len = max(0.0, min(ARC_LEN * (ff / max_ff), ARC_LEN)) if max_ff else 0
    badges = []
    if is_p:
        badges.append('<span class="promo-chip">◆ AMEX Promo</span>')
    if feriado:
        badges.append(f'<span class="dc-feriado-chip">🗓 {feriado}</span>')
    if is_fallback:
        badges.append('<span class="dc-warn-chip">⚠ Sin historial</span>')
    chips = f'<div class="dc-chips">{"".join(badges)}</div>' if badges else ''
    return f"""<div class="{cls}" style="animation-delay:{card_delay}">
  <div class="dc-hdr">
    <span class="dc-dow">{DAYS_ES[d.weekday()]}</span>
    <span class="dc-cal">{d.strftime('%d')}<span class="dc-mon"> {MONTHS_ES[d.month][:3]}</span></span>
  </div>
  <div class="dc-arc-wrap">
    <svg width="136" height="80" viewBox="0 0 136 80" class="dc-svg">
      <path d="M 8,72 A 60,60 0 0,1 128,72"
            fill="none" stroke="rgba(255,255,255,.05)" stroke-width="5" stroke-linecap="round"/>
      <path d="M 8,72 A 60,60 0 0,1 128,72"
            fill="none" stroke="{bar_c}" stroke-width="5" stroke-linecap="round"
            stroke-dasharray="{fill_len:.1f} {ARC_LEN:.1f}"
            class="dc-arc-fill"
            style="--arc-fill:{fill_len:.1f};animation-delay:{arc_delay}"/>
    </svg>
    <div class="dc-hero">
      <span class="dc-big" style="color:{nc}">{round(ff):,}</span>
      <span class="dc-lbl">fondos</span>
    </div>
  </div>
  <div class="dc-total">{round(pf):,}<span class="dc-total-lbl"> platos</span></div>
  {chips}
</div>"""

# ── Chart builders ────────────────────────────────────────────────────────────

def chart_forecast_bars(totals, dish_stats, cat='Fondo', top_n=20):
    items = sorted(
        [(n, c, q, dish_stats.get((n, c), {'mu': q, 'sigma': 0}))
         for (n, c), q in totals.items() if c == cat and q >= 1],
        key=lambda x: -x[2]
    )[:top_n]
    if not items:
        return go.Figure()

    names  = [x[0] for x in items]
    mu_v   = [round(x[3]['mu'])    for x in items]
    sig_v  = [round(x[3]['sigma']) for x in items]
    total  = [round(x[2])          for x in items]
    color  = CAT_HEX.get(cat, GOLD)

    fig = go.Figure()
    fig.add_trace(go.Bar(
        name='Promedio (μ)', y=names, x=mu_v, orientation='h',
        marker=dict(color=f'rgba({int(color[1:3],16)},{int(color[3:5],16)},{int(color[5:7],16)},.45)',
                    cornerradius=4),
        hovertemplate='<b>%{y}</b><br>Promedio: %{x}<extra></extra>',
    ))
    fig.add_trace(go.Bar(
        name='Buffer (σ)', y=names, x=sig_v, orientation='h',
        marker=dict(color='rgba(110,231,183,.65)', cornerradius=4),
        hovertemplate='<b>%{y}</b><br>Buffer: +%{x}<extra></extra>',
    ))
    # Annotation: total
    for name, tot in zip(names, total):
        fig.add_annotation(
            x=tot, y=name, text=f'<b>{tot}</b>', showarrow=False,
            xanchor='left', xshift=8, font=dict(size=11, color=color),
        )

    layout = dict(PLOTLY_BASE)
    layout.update(dict(
        barmode='stack', height=max(380, len(names) * 32),
        title=dict(text=f'Producción Forecast — {cat}s', **PLOTLY_BASE['title']),
        yaxis={**PLOTLY_BASE['yaxis'], 'autorange': 'reversed'},
        xaxis={**PLOTLY_BASE['xaxis'], 'title': 'Unidades a producir'},
        showlegend=True,
    ))
    fig.update_layout(**layout)
    return fig


def chart_heatmap(df, top_n=14):
    if df.empty:
        return go.Figure()
    fondo_df = df[df['category'] == 'Fondo']
    top_dishes = (fondo_df.groupby('dish')['quantity']
                  .mean().sort_values(ascending=False)
                  .head(top_n).index.tolist())
    pivot = (fondo_df[fondo_df['dish'].isin(top_dishes)]
             .groupby(['weekday','dish'])['quantity']
             .mean().unstack(fill_value=0))
    order = [d for d in (DAYS_ES + ['Dom-Promo']) if d in pivot.index]
    pivot = pivot.reindex(order)[top_dishes]

    fig = go.Figure(go.Heatmap(
        z=pivot.values,
        x=[n[:20] for n in pivot.columns],
        y=pivot.index,
        colorscale=[[0,'#0B1428'],[0.3,'#3A2A00'],[0.65,'#8B6B1A'],[1,'#C9A97A']],
        hovertemplate='<b>%{x}</b><br>%{y}: <b>%{z:.1f}</b> uds<extra></extra>',
        showscale=True,
        colorbar=dict(bgcolor='rgba(0,0,0,0)', bordercolor=BORDER,
                      tickfont=dict(color=MUTED, size=10), thickness=12, len=0.8),
        text=np.round(pivot.values, 1),
        texttemplate='<b>%{text}</b>',
        textfont=dict(size=10, color='rgba(255,255,255,0.7)'),
    ))
    layout = dict(PLOTLY_BASE)
    layout.update(dict(
        title=dict(text='Patrón de Demanda por Día de Semana', **PLOTLY_BASE['title']),
        height=340,
        xaxis={**PLOTLY_BASE['xaxis'], 'tickangle': -40, 'tickfont': dict(size=10)},
        yaxis=dict(**PLOTLY_BASE['yaxis']),
    ))
    fig.update_layout(**layout)
    return fig


def chart_trend(df, dishes):
    if df.empty or not dishes:
        return go.Figure()
    colors = [GOLD,'#4F96FF','#2DD4BF','#F472B6','#A78BFA','#E8C878','#93C5FD','#FCA5A5']
    fig    = go.Figure()
    for i, dish in enumerate(dishes):
        data = (df[df['dish'] == dish]
                .groupby('date')['quantity'].sum()
                .reset_index().sort_values('date'))
        if data.empty:
            continue
        data['roll'] = data['quantity'].rolling(7, min_periods=1).mean()
        c = colors[i % len(colors)]
        fig.add_trace(go.Scatter(
            x=data['date'], y=data['quantity'], name=f'{dish} (diario)',
            mode='markers', marker=dict(size=5, color=c, opacity=0.35),
            showlegend=False, hovertemplate=f'<b>{dish}</b><br>%{{x|%d/%m}}: %{{y}}<extra></extra>',
        ))
        fig.add_trace(go.Scatter(
            x=data['date'], y=data['roll'], name=dish,
            mode='lines', line=dict(width=2.5, color=c),
            hovertemplate=f'<b>{dish}</b><br>%{{x|%d/%m}}: %{{y:.1f}} (media 7d)<extra></extra>',
        ))
    layout = dict(PLOTLY_BASE)
    layout.update(dict(
        title=dict(text='Tendencia Histórica — Media 7 días', **PLOTLY_BASE['title']),
        height=420, hovermode='x unified',
        xaxis=dict(**PLOTLY_BASE['xaxis'], title=None, tickformat='%d/%m'),
        yaxis=dict(**PLOTLY_BASE['yaxis'], title='Unidades vendidas'),
    ))
    fig.update_layout(**layout)
    return fig


def compute_buffer_coverage(history, model):
    """
    Para cada tipo de día, calcula qué % de días históricos habrías cubierto
    con cada factor k (0, 0.5, 0.7, 1.0, 1.5, 2.0).
    'Cubierto' = total fondos producidos ese día ≤ pronóstico con ese factor.
    """
    K_VALUES = [0.0, 0.5, 0.7, 1.0, 1.5, 2.0]
    results   = {}

    for dt, dm in model.items():
        days = [(d, data) for d, data in history.items() if day_type(d) == dt]
        n    = len(days)
        if n < 2:
            continue

        # Pre-calcular fondos base del modelo para este tipo de día
        base_fondos = {
            name: s
            for (name, cat), s in dm.items()
            if cat == 'Fondo'
        }
        if not base_fondos:
            continue

        covered = {k: 0 for k in K_VALUES}

        for _, actual_data in days:
            actual_total = sum(
                qty for (_, name, cat), qty in actual_data.items()
                if cat == 'Fondo'
            )
            for k in K_VALUES:
                forecast = sum(
                    s['mean'] + k * (s['std'] if s['std'] is not None else s['mean'] * BUFFER_FALLBACK)
                    for s in base_fondos.values()
                )
                if actual_total <= forecast:
                    covered[k] += 1

        # Calcular también la desviación promedio: cuánto % por encima del promedio
        # sucede en días no cubiertos con k=0
        excesos = []
        for _, actual_data in days:
            actual_f = sum(qty for (_, _, cat), qty in actual_data.items() if cat == 'Fondo')
            base_f   = sum(s['mean'] for s in base_fondos.values())
            if base_f > 0:
                excesos.append((actual_f - base_f) / base_f * 100)

        results[dt] = {
            'n':        n,
            'coverage': {k: round(covered[k] / n * 100) for k in K_VALUES},
            'avg_excess_pct': round(sum(e for e in excesos if e > 0) / max(1, sum(1 for e in excesos if e > 0)), 1) if excesos else 0,
        }

    return results


def chart_variability(model, cat='Fondo', top_n=18):
    rows = []
    for dt, dishes in model.items():
        for (name, c), stats in dishes.items():
            if c == cat and stats['std'] is not None:
                rows.append({'dish': name, 'day_type': dt,
                             'mean': stats['mean'], 'std': stats['std'],
                             'cv': stats['std'] / stats['mean'] if stats['mean'] > 0 else 0})
    if not rows:
        return go.Figure()
    df_v = pd.DataFrame(rows)
    top  = df_v.groupby('dish')['mean'].mean().sort_values(ascending=False).head(top_n).index
    df_v = df_v[df_v['dish'].isin(top)]
    agg  = df_v.groupby('dish').agg(mean_avg=('mean','mean'), std_avg=('std','mean'), cv=('cv','mean')).reset_index()
    agg  = agg.sort_values('mean_avg', ascending=False)

    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=agg['dish'], x=agg['mean_avg'], name='Media', orientation='h',
        error_x=dict(type='data', array=agg['std_avg'], color='rgba(110,231,183,.6)',
                     thickness=1.5, width=6),
        marker=dict(color=[f'rgba(201,169,122,{max(.3, 1-row.cv*.8):.2f})' for _, row in agg.iterrows()],
                    cornerradius=4),
        hovertemplate='<b>%{y}</b><br>Media: %{x:.1f} ± %{error_x.array:.1f}<extra></extra>',
    ))
    layout = dict(PLOTLY_BASE)
    layout.update(dict(
        title=dict(text='Variabilidad por Plato — Barras = ±1σ', **PLOTLY_BASE['title']),
        height=max(380, len(agg) * 30),
        yaxis=dict(**PLOTLY_BASE['yaxis'], autorange='reversed'),
        xaxis=dict(**PLOTLY_BASE['xaxis'], title='Unidades promedio'),
        showlegend=False,
    ))
    fig.update_layout(**layout)
    return fig


def chart_promo_impact(model, top_n=15):
    normal_d = model.get('Domingo',   {})
    promo_d  = model.get('Dom-Promo', {})
    if not promo_d:
        return go.Figure()

    dishes = sorted(
        [(n, s['mean']) for (n, c), s in promo_d.items() if c == 'Fondo'],
        key=lambda x: -x[1]
    )[:top_n]

    names     = [d[0] for d in dishes]
    promo_v   = [d[1] for d in dishes]
    normal_v = []
    for name in names:
        val = 0
        for (n, c), stats in normal_d.items():
            if n == name and c == 'Fondo':
                val = stats['mean']
                break
        normal_v.append(val)

    factors = [p/n if n > 0 else 0 for p, n in zip(promo_v, normal_v)]

    fig = go.Figure()
    fig.add_trace(go.Bar(
        name='Domingo Normal', y=names, x=normal_v, orientation='h',
        marker=dict(color='rgba(74,86,104,.7)', cornerradius=4),
        hovertemplate='<b>%{y}</b><br>Normal: %{x:.1f}<extra></extra>',
    ))
    fig.add_trace(go.Bar(
        name='Dom-Promo AMEX', y=names, x=promo_v, orientation='h',
        marker=dict(color='rgba(201,169,122,.85)', cornerradius=4),
        hovertemplate='<b>%{y}</b><br>Promo: %{x:.1f}<extra></extra>',
    ))
    for name, pv, nv, fct in zip(names, promo_v, normal_v, factors):
        if nv > 0 and fct > 0:
            fig.add_annotation(
                x=max(pv, nv), y=name, text=f'  {fct:.1f}×',
                showarrow=False, xanchor='left',
                font=dict(size=10, color=GOLD if fct > 1.5 else MUTED),
            )

    layout = dict(PLOTLY_BASE)
    layout.update(dict(
        barmode='overlay',
        title=dict(text='Impacto Promoción AMEX — Fondos por Plato', **PLOTLY_BASE['title']),
        height=max(400, len(names) * 32),
        yaxis={**PLOTLY_BASE['yaxis'], 'autorange': 'reversed'},
        xaxis={**PLOTLY_BASE['xaxis'], 'title': 'Unidades promedio por domingo'},
    ))
    fig.update_layout(**layout)
    return fig


def chart_weekly_total(df):
    if df.empty:
        return go.Figure()
    weekly = (df.groupby([pd.Grouper(key='date', freq='W'), 'is_promo'])['quantity']
              .sum().reset_index())
    weekly.columns = ['week', 'is_promo', 'total']

    fig = go.Figure()
    for is_p, clr, name in [(False,'rgba(255,255,255,.32)','Semana normal'),(True,GOLD,'Semana con Dom-Promo')]:
        d = weekly[weekly['is_promo'] == is_p]
        if not d.empty:
            fig.add_trace(go.Scatter(
                x=d['week'], y=d['total'], name=name,
                mode='lines+markers',
                line=dict(color=clr, width=2.5),
                marker=dict(size=7, color=clr),
                hovertemplate='%{x|%d/%m}: <b>%{y:,}</b> uds<extra></extra>',
            ))
    layout = dict(PLOTLY_BASE)
    layout.update(dict(
        title=dict(text='Volumen Semanal Total — Todos los Platos', **PLOTLY_BASE['title']),
        height=340,
        xaxis=dict(**PLOTLY_BASE['xaxis'], tickformat='%d/%m'),
        yaxis=dict(**PLOTLY_BASE['yaxis'], title='Unidades'),
        hovermode='x unified',
    ))
    fig.update_layout(**layout)
    return fig


def make_excel_download(day_details, totals, dish_stats, history_len, k_factor=K_SAFETY):
    if not HAS_OPENPYXL:
        return None
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Producción'

    NAVY = PatternFill('solid', fgColor='1F3460')
    GOLD_F = PatternFill('solid', fgColor='C9A84C')
    GRAY = PatternFill('solid', fgColor='F2F4F8')
    thin = Side(style='thin', color='D0D5E0')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def c(r, col, val, bold=False, fill=None, clr=None, align='left'):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font      = Font(bold=bold, color=clr or ('FFFFFF' if fill in (NAVY,GOLD_F) else '111827'))
        cell.fill      = fill or PatternFill()
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border    = brd
        return cell

    ws.merge_cells('A1:E1')
    ws['A1'] = 'PRONÓSTICO DE PRODUCCIÓN — Margo / Nelí'
    ws['A1'].font      = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill      = NAVY
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24

    period = '  +  '.join(
        f"{DAYS_ES[d.weekday()]} {d.strftime('%d/%m/%Y')} [{dt}]"
        for d, dt, _, _ in day_details
    )
    ws.merge_cells('A2:E2')
    ws['A2'] = f'Período: {period}'
    ws.merge_cells('A3:E3')
    ws['A3'] = f'Generado: {_fecha_es(datetime.now())}   Buffer: μ + {k_factor}σ   Modelo: {history_len} días'
    ws['A3'].font = Font(italic=True, color='6B7280')

    row = 5
    for cat in CAT_ORDER:
        items = {k: v for k, v in totals.items() if k[1] == cat and v >= 1}
        if not items:
            continue
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = cat.upper()
        ws[f'A{row}'].font      = Font(bold=True, color='FFFFFF')
        ws[f'A{row}'].fill      = GOLD_F
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1
        for lbl, col in [('Plato',1),('Promedio μ',2),('Buffer σ',3),('Total',4)]:
            c(row, col, lbl, bold=True, fill=GRAY, align='center')
        row += 1
        for (name, _), qty in sorted(items.items(), key=lambda x: -x[1]):
            st_d = dish_stats.get((name, _), {'mu': qty, 'sigma': 0})
            c(row, 1, name)
            c(row, 2, round(st_d['mu']),    align='center')
            c(row, 3, round(st_d['sigma']), align='center')
            c(row, 4, round(qty),           align='center', bold=True)
            row += 1
        row += 1

    ws.column_dimensions['A'].width = 40
    for col in ['B','C','D']:
        ws.column_dimensions[col].width = 14

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Load data ─────────────────────────────────────────────────────────────────

history = load_history_cached(FOLDER)
if not history:
    st.error("No se encontraron archivos de informe en la carpeta.")
    st.stop()

dates_sorted = sorted(history)
model        = build_model(history)
df           = build_df(history)
name_to_sku  = build_name_to_sku(history)
n_promo      = sum(1 for d in history if day_type(d) == 'Dom-Promo')

# ── Recetas (cargado una vez, compartido entre tabs COMPRAS y RECETAS) ────────
_RECETAS_PATH = os.path.join(_BASE_DIR, "recetas.json")
try:
    with open(_RECETAS_PATH, encoding='utf-8') as _rf:
        _recetas_raw = json.load(_rf)
    RECETAS = {k: v for k, v in _recetas_raw.items() if not k.startswith('_')}
except Exception:
    _recetas_raw = {}
    RECETAS = {}

# ── Ventas Financieras (Ventas/*.xls) ────────────────────────────────────────

_VENTAS_DIR   = os.path.join(_BASE_DIR, "Ventas")
_MESES_MAP_V  = {'Ene':1,'Feb':2,'Mar':3,'Abr':4,'May':5,'Jun':6,
                 'Jul':7,'Ago':8,'Sep':9,'Oct':10,'Nov':11,'Dic':12}
_MESES_NOM_V  = ['','Ene','Feb','Mar','Abr','May','Jun',
                  'Jul','Ago','Sep','Oct','Nov','Dic']
_GRUPOS_VTA   = {'Alimentos','Bebidas S/Alcohol','Bebidas C/Alcohol','Vinos'}

@st.cache_data(ttl=3600)
def _load_ventas_df():
    rows = []
    if not os.path.isdir(_VENTAS_DIR):
        return pd.DataFrame()
    for fname in sorted(os.listdir(_VENTAS_DIR)):
        if not fname.endswith('.xls'):
            continue
        mes_str = fname[:3]
        try:
            anio = int('20' + fname[3:5])
            mes  = _MESES_MAP_V.get(mes_str, 0)
            if mes == 0:
                continue
        except Exception:
            continue
        try:
            wb = xlrd.open_workbook(os.path.join(_VENTAS_DIR, fname))
            sh = wb.sheets()[0]
            for r in range(sh.nrows):
                try:
                    nombre = str(sh.cell_value(r, 7)).strip()
                    if nombre in _GRUPOS_VTA:
                        bruto = float(sh.cell_value(r, 21) or 0)
                        cant  = float(sh.cell_value(r, 16) or 0)
                        desc  = float(sh.cell_value(r, 28) or 0)
                        rows.append({'año': anio, 'mes': mes, 'categoria': nombre,
                                     'bruto': bruto, 'cantidad': cant, 'descuento': desc})
                except Exception:
                    pass
        except Exception:
            pass
    if not rows:
        return pd.DataFrame()
    return pd.DataFrame(rows)

_vdf = _load_ventas_df()

# ── Sidebar ────────────────────────────────────────────────────────────────────

with st.sidebar:
    st.markdown(f"""
    <div class="sb-logo-wrap">
      <img src="data:image/png;base64,{LOGO_B64}" class="sb-logo-img" alt="Margó Gourmet">
      <div class="sb-logo-tagline">Sistema de Producción</div>
    </div>
    """, unsafe_allow_html=True)

    st.markdown("##### Parámetros")

    start_date = st.date_input(
        "Fecha de inicio",
        value=datetime.now().date() + timedelta(days=1),
        min_value=datetime(2026, 1, 1).date(),
        key="start_date",
    )
    horizon = st.slider("Días de producción", 1, 7, DEFAULT_HORIZON,
                        help="Cuántos días hacia adelante producir")
    k_factor = st.slider(
        "Factor buffer (σ)", 0.5, 2.5, float(K_SAFETY), 0.1,
        help="Cuánto extra producir sobre el promedio. 1.0 = promedio + 1 desviación estándar (~84% de cobertura). Sube en días especiales, baja si hay mucho desperdicio.",
        format="%.1f",
    )
    # Cobertura teórica aproximada (distribución normal)
    _cov_approx = {0.5: 69, 0.7: 76, 1.0: 84, 1.5: 93, 2.0: 98, 2.5: 99}
    _k_near = min(_cov_approx, key=lambda x: abs(x - k_factor))
    _cov_pct = _cov_approx[_k_near]
    _cov_c = 'rgba(100,230,170,.7)' if _cov_pct >= 90 else ('#F59E0B' if _cov_pct >= 80 else 'rgba(201,169,122,.7)')
    st.markdown(
        f'<div style="font-size:10px;color:rgba(255,255,255,.28);margin-top:-6px;display:flex;justify-content:space-between">'
        f'<span>μ + {k_factor:.1f}σ por plato</span>'
        f'<span style="color:{_cov_c}">~{_cov_pct}% cobertura</span></div>',
        unsafe_allow_html=True,
    )

    st.divider()
    st.markdown("##### Estado del Modelo")

    d0 = dates_sorted[0].strftime('%d/%m/%y')
    d1 = dates_sorted[-1].strftime('%d/%m/%y')
    st.markdown(f"""
    <div class="kpi-model">
      <div class="km-label">Período datos</div>
      <div class="km-val">{d0} → {d1}</div>
      <div class="km-sub">{len(history)} días cargados</div>
    </div>
    <div class="kpi-model" style="margin-top:8px">
      <div class="km-label">Dom-Promo AMEX</div>
      <div class="km-val">{n_promo} observaciones</div>
      <div class="km-sub">{_confidence_badge(n_promo)}</div>
    </div>
    """, unsafe_allow_html=True)

    st.divider()

    # ── Carpeta de reportes ────────────────────────────────────────
    st.markdown("##### Carpeta de Reportes")
    current_folder = _get_folder()
    n_files = _count_reports(current_folder)

    # Mostrar carpeta actual
    folder_name = os.path.basename(current_folder) or current_folder
    st.markdown(
        f'<div style="font-family:var(--font-mono);font-size:10px;color:rgba(255,255,255,.5);'
        f'background:rgba(255,255,255,.04);border:1px solid rgba(255,255,255,.08);'
        f'border-radius:6px;padding:6px 10px;margin-bottom:6px;word-break:break-all">'
        f'{current_folder}</div>'
        f'<div style="font-size:11px;color:rgba(201,169,122,.7);margin-bottom:8px">'
        f'{n_files} reportes encontrados</div>',
        unsafe_allow_html=True,
    )

    # Expandir para cambiar carpeta
    with st.expander("📁  Cambiar carpeta"):
        nueva = st.text_input(
            "Ruta de la carpeta",
            value=current_folder,
            placeholder="Ej: C:\\Users\\Juan\\Reportes",
            label_visibility="collapsed",
            key="nueva_carpeta",
        )
        col_ok, col_cancel = st.columns(2)
        with col_ok:
            if st.button("Guardar", use_container_width=True, key="btn_guardar_carpeta"):
                nueva = nueva.strip().strip('"').strip("'")
                if not os.path.isdir(nueva):
                    st.error("La carpeta no existe")
                elif _count_reports(nueva) == 0:
                    st.warning("No se encontraron archivos 'informe*.xls' en esa carpeta")
                else:
                    cfg = _load_config()
                    cfg["reportes_folder"] = nueva
                    _save_config(cfg)
                    st.cache_data.clear()
                    st.success(f"Guardado · {_count_reports(nueva)} reportes")
                    st.rerun()

    if st.button("🔄  Recargar archivos", use_container_width=True):
        st.cache_data.clear()
        st.rerun()

    # ── Próximos feriados ──────────────────────────────────────
    st.divider()
    st.markdown("##### Próximos Feriados")
    hoy = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    proximos = [(d, n) for d, n in sorted(FERIADOS.items()) if d >= hoy][:5]
    if proximos:
        for fd, fn in proximos:
            dias = (fd - hoy).days
            tag = "**HOY**" if dias == 0 else f"en {dias}d"
            color = "#F59E0B" if dias <= 7 else "rgba(255,255,255,.45)"
            st.markdown(
                f'<div style="font-size:12px;color:{color};padding:3px 0">'
                f'<b>{fd.strftime("%d/%m")}</b> · {fn} <span style="opacity:.5;font-size:10px">({tag})</span>'
                f'</div>',
                unsafe_allow_html=True,
            )
    else:
        st.caption("Sin feriados próximos")

    # ── Semáforo de ventas ────────────────────────────────────────
    if not _vdf.empty:
        st.divider()
        st.markdown("##### Tendencia de Ventas")
        _sb_common = [1,2,3,4,5]
        _sb_vm = _vdf.groupby(['año','mes'])['bruto'].sum().reset_index()
        _sb25  = _sb_vm[_sb_vm['año']==2025].set_index('mes')['bruto']
        _sb26  = _sb_vm[_sb_vm['año']==2026].set_index('mes')['bruto']
        _sb_ok = sorted(set(_sb25.index) & set(_sb26.index))
        if _sb_ok:
            # Tendencia últimos 3 meses comparables
            _sb_ult3 = _sb_ok[-3:]
            _sb_yoy  = [((_sb26[m]-_sb25[m])/_sb25[m]*100) for m in _sb_ult3]
            _sb_tend = sum(_sb_yoy) / len(_sb_yoy)
            # Último mes
            _sb_last_m  = _sb_ok[-1]
            _sb_last_pct= (_sb26[_sb_last_m]-_sb25[_sb_last_m])/_sb25[_sb_last_m]*100
            # Semáforo
            if _sb_tend >= 0:
                _sb_clr, _sb_ico, _sb_lbl = '#5CE8D4', '●', 'Crecimiento'
            elif _sb_tend >= -5:
                _sb_clr, _sb_ico, _sb_lbl = '#F59E0B', '●', 'Estable / Leve baja'
            elif _sb_tend >= -12:
                _sb_clr, _sb_ico, _sb_lbl = '#F7A8D0', '●', 'En baja'
            else:
                _sb_clr, _sb_ico, _sb_lbl = '#EF4444', '●', 'Alerta'
            _MESES_SB = ['','Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
            st.markdown(
                f'<div style="background:rgba(255,255,255,0.03);border:1px solid rgba(255,255,255,.06);'
                f'border-radius:8px;padding:12px 14px">'
                f'<div style="font-size:18px;color:{_sb_clr};margin-bottom:4px">'
                f'{_sb_ico} {_sb_lbl}</div>'
                f'<div style="font-size:11px;color:rgba(255,255,255,.55)">Últimos 3 meses vs 2025</div>'
                f'<div style="font-size:22px;font-weight:600;color:{_sb_clr};margin:4px 0">'
                f'{"+" if _sb_tend>=0 else ""}{_sb_tend:.1f}%</div>'
                f'<div style="font-size:10px;color:rgba(255,255,255,.35);margin-top:6px">'
                + '  '.join([f'{_MESES_SB[m]}: {"+" if ((_sb26[m]-_sb25[m])/_sb25[m]*100)>=0 else ""}{((_sb26[m]-_sb25[m])/_sb25[m]*100):.0f}%'
                              for m in _sb_ult3])
                + f'</div></div>',
                unsafe_allow_html=True)

    st.markdown("""
    <div class="sb-footer">
      <div class="sb-footer-name">Margó Gourmet</div>
      <div class="sb-footer-sub">Isidora Goyenechea · Uso interno</div>
    </div>
    """, unsafe_allow_html=True)

# ── Compute forecast ───────────────────────────────────────────────────────────

start_dt   = datetime.combine(start_date, datetime.min.time())
day_fc, totals, dish_stats = make_forecast(model, start_dt, horizon, k_factor)
dish_coverage = compute_dish_coverage(history, day_fc, model, k_factor)

total_fondos = round(sum(v for (_, c), v in totals.items() if c == 'Fondo'))
total_all    = round(sum(totals.values()))
has_promo    = any(dt == 'Dom-Promo' for _, dt, _, _ in day_fc)

# ── Header ─────────────────────────────────────────────────────────────────────

period_parts = []
for d, dt, _, _ in day_fc:
    chip = f' <span class="promo-chip" style="font-size:8px">AMEX</span>' if dt == 'Dom-Promo' else ''
    period_parts.append(f'{DAYS_ES[d.weekday()][:3]} {d.strftime("%d/%m")}{chip}')
period_str = ' <span class="bh-sep">·</span> '.join(period_parts)

st.markdown(f"""
<div class="brand-header">
  <div class="bh-inner">
    <div>
      <div class="bh-brand">
        Margo · Nelí · Sistema de Producción
      </div>
      <h1 class="bh-title"><em>Pronóstico de</em> <strong>Producción</strong></h1>
      <div class="bh-period">{period_str}</div>
    </div>
    <div class="bh-right">
      <div class="bh-date">{_fecha_es(datetime.now())}</div>
      <div class="bh-meta">{len(history)} días · buffer μ+{k_factor}σ</div>
    </div>
  </div>
</div>
""", unsafe_allow_html=True)

# ── KPI Metrics ───────────────────────────────────────────────────────────────

col1, col2, col3, col4 = st.columns(4)
with col1:
    st.metric("Fondos a Producir", f"{total_fondos:,}",
              delta=f"{horizon} día{'s' if horizon>1 else ''} de horizonte")
with col2:
    st.metric("Total Platos", f"{total_all:,}")
with col3:
    st.metric("Días en Modelo", f"{len(history):,}",
              delta=f"μ + {k_factor}σ por plato")
with col4:
    conf_delta = "⚠ Confianza baja" if n_promo <= 1 else f"✓ {n_promo} obs" if n_promo >= 4 else f"△ {n_promo} obs"
    st.metric("Dom-Promo AMEX", f"{n_promo}",
              delta=conf_delta, delta_color="off")

st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

# ── Main Tabs ─────────────────────────────────────────────────────────────────

tab_prod, tab_analisis, tab_tend, tab_promo, tab_hist, tab_compras, tab_recetas, tab_ventas = st.tabs([
    "🍽  PRODUCCIÓN",
    "📊  ANÁLISIS",
    "📈  TENDENCIAS",
    "⭐  PROMO AMEX",
    "📋  HISTORIAL",
    "🛒  COMPRAS",
    "📖  RECETAS",
    "💰  VENTAS",
])

# ──────────────────────────────────────────────────────────────────────────────
# TAB 1: PRODUCCIÓN
# ──────────────────────────────────────────────────────────────────────────────

with tab_prod:

    # Day cards
    all_ff  = [sum(dish_fc(s, k_factor) for (_, c), s in dm.items() if c == 'Fondo')
               for _, _, dm, _ in day_fc]
    all_pf  = [sum(dish_fc(s, k_factor) for s in dm.values())
               for _, _, dm, _ in day_fc]
    max_ff  = max(all_ff) if all_ff else 1

    if has_promo:
        st.markdown("""<div class="alert-promo">
          <strong>◆ Dom-Promo AMEX activa en este período</strong> —
          El modelo aplica el factor histórico de domingos con promo.
          Confianza baja con pocas observaciones: mantén buffer generoso.
        </div>""", unsafe_allow_html=True)

    cols = st.columns(len(day_fc))
    for i, ((d, dt, dm, is_fb), ff, pf) in enumerate(zip(day_fc, all_ff, all_pf)):
        with cols[i]:
            st.markdown(day_card_html(d, dt, ff, pf, max_ff, is_fb, idx=i), unsafe_allow_html=True)

    st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

    # Category tabs inside Producción
    cat_tabs = st.tabs([f"  {c}  " for c in CAT_ORDER])
    for ci, (cat, ctab) in enumerate(zip(CAT_ORDER, cat_tabs)):
        with ctab:
            items = {k: v for k, v in totals.items() if k[1] == cat and v >= 1}
            color = CAT_HEX[cat]
            n_items = len(items)
            total_cat = round(sum(items.values()))

            r, g, b = int(color[1:3], 16), int(color[3:5], 16), int(color[5:7], 16)
            st.markdown(f"""<div class="sec-hd" style="border-left:2px solid rgba({r},{g},{b},.45);padding-left:14px">
              <span class="sec-n">0{ci+1}</span>
              <span class="sec-t" style="color:{color}">{cat}</span>
              <span class="sec-r"></span>
              <span class="sec-c">{n_items} platos · {total_cat:,} uds</span>
            </div>""", unsafe_allow_html=True)

            st.markdown(prod_table_html(items, dish_stats, color, dish_coverage), unsafe_allow_html=True)

    # Totals strip + download
    st.markdown(f"""<div class="totals-strip">
      <div class="ts-item">
        <dt>Total Fondos</dt>
        <dd style="color:{GOLD}">{total_fondos:,}</dd>
        <div class="ts-sub">platos de fondo · {horizon}d</div>
      </div>
      <div class="ts-sep"></div>
      <div class="ts-item">
        <dt>Total Platos</dt>
        <dd style="color:#7EB8F7">{total_all:,}</dd>
        <div class="ts-sub">todas las categorías</div>
      </div>
      <div class="ts-sep"></div>
      <div class="ts-item">
        <dt>Horizonte</dt>
        <dd style="color:var(--t2);font-size:28px;text-shadow:none">{horizon}d</dd>
        <div class="ts-sub">{horizon} día{'s' if horizon > 1 else ''} de producción</div>
      </div>
      <div class="ts-sep"></div>
      <div class="ts-item">
        <dt>Días Modelo</dt>
        <dd style="color:var(--t2);font-size:28px;text-shadow:none">{len(history)}</dd>
        <div class="ts-sub">observaciones históricas</div>
      </div>
    </div>""", unsafe_allow_html=True)

    st.markdown("<div style='height:16px'></div>", unsafe_allow_html=True)

    excel_data = make_excel_download(day_fc, totals, dish_stats, len(history), k_factor)
    if excel_data:
        fname = f"produccion_{start_date.strftime('%Y%m%d')}.xlsx"
        st.download_button(
            "⬇  Descargar Excel de Producción",
            data=excel_data,
            file_name=fname,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )

# ──────────────────────────────────────────────────────────────────────────────
# TAB 2: ANÁLISIS
# ──────────────────────────────────────────────────────────────────────────────

with tab_analisis:

    col_left, col_right = st.columns([3, 1])

    with col_left:
        selected_cat = st.selectbox(
            "Categoría",
            CAT_ORDER,
            label_visibility="collapsed",
            key="analysis_cat",
        )
        fig_bars = chart_forecast_bars(totals, dish_stats, selected_cat)
        st.plotly_chart(fig_bars, use_container_width=True, config={"displayModeBar": False})

    with col_right:
        # Category breakdown donut
        cat_totals = {}
        for cat in CAT_ORDER:
            v = sum(q for (_, c), q in totals.items() if c == cat and q >= 1)
            if v > 0:
                cat_totals[cat] = round(v)

        fig_donut = go.Figure(go.Pie(
            labels=list(cat_totals.keys()),
            values=list(cat_totals.values()),
            hole=0.62,
            marker=dict(colors=[CAT_HEX[c] for c in cat_totals],
                        line=dict(color=BG_2, width=2)),
            textinfo='none',
            hovertemplate='<b>%{label}</b><br>%{value:,} uds (%{percent})<extra></extra>',
        ))
        fig_donut.add_annotation(
            text=f'<b>{total_all:,}</b><br><span style="font-size:10px">platos</span>',
            x=0.5, y=0.5, showarrow=False,
            font=dict(size=20, color=GOLD, family='Consolas'),
        )
        layout_d = dict(PLOTLY_BASE)
        layout_d.update(dict(
            title=dict(text='Distribución', **PLOTLY_BASE['title']),
            showlegend=True, height=340,
            legend=dict(orientation='v', x=1, y=0.5, bgcolor='rgba(0,0,0,0)'),
            margin=dict(l=8, r=8, t=48, b=8),
        ))
        fig_donut.update_layout(**layout_d)
        st.plotly_chart(fig_donut, use_container_width=True, config={"displayModeBar": False})

        st.divider()

        # Mini stats per category
        for cat, val in sorted(cat_totals.items(), key=lambda x: -x[1]):
            pct = round(val / total_all * 100) if total_all else 0
            st.markdown(f"""
            <div style="display:flex;justify-content:space-between;align-items:center;
                        padding:6px 0;border-bottom:1px solid {BORDER}">
              <span style="font-size:12px;color:{CAT_HEX[cat]};font-weight:600">{cat}</span>
              <span style="font-family:'Consolas','Cascadia Code',monospace;font-size:12px;color:{TEXT}">{val:,} <span style="color:{SUBTLE}">({pct}%)</span></span>
            </div>
            """, unsafe_allow_html=True)

    st.divider()

    # Variability chart
    st.markdown(f"""<div class="sec-hd">
      <span class="sec-t" style="font-size:18px">Variabilidad por Plato</span>
      <span class="sec-r"></span>
      <span class="sec-c">Barras muestran ±1σ — platos más predecibles tienen barras más cortas</span>
    </div>""", unsafe_allow_html=True)

    fig_var = chart_variability(model, selected_cat)
    st.plotly_chart(fig_var, use_container_width=True, config={"displayModeBar": False})

    # ── Cobertura histórica del buffer ────────────────────────────────────────
    st.markdown(f"""<div class="sec-hd" style="margin-top:12px">
      <span class="sec-t" style="font-size:18px">Cobertura Histórica del Buffer</span>
      <span class="sec-r"></span>
      <span class="sec-c">% de días que habrías cubierto con cada factor (solo Fondos)</span>
    </div>""", unsafe_allow_html=True)

    coverage_data = compute_buffer_coverage(history, model)

    if coverage_data:
        K_COLS  = [0.0, 0.5, 0.7, 1.0, 1.5, 2.0]
        DT_ORDER = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo','Dom-Promo']

        # Header
        header_html = '<div style="display:grid;grid-template-columns:120px repeat(6,1fr);gap:6px;margin-bottom:6px">'
        header_html += '<div style="font-size:9px;letter-spacing:.15em;text-transform:uppercase;color:rgba(255,255,255,.32);padding:5px 0">Tipo de día</div>'
        for k in K_COLS:
            is_cur = abs(k - k_factor) < 0.01
            style  = f'text-align:center;font-size:9px;letter-spacing:.12em;text-transform:uppercase;color:{"#C9A97A" if is_cur else "rgba(255,255,255,.32)"};padding:5px 0;{"font-weight:700" if is_cur else ""}'
            label  = f'σ×{k:.1f}{"  ←" if is_cur else ""}'
            header_html += f'<div style="{style}">{label}</div>'
        header_html += '</div>'

        rows_html = '<div style="display:grid;grid-template-columns:120px repeat(6,1fr);gap:6px">'
        for dt_name in DT_ORDER:
            if dt_name not in coverage_data:
                continue
            info = coverage_data[dt_name]
            n    = info['n']
            avg_ex = info['avg_excess_pct']
            rows_html += f'<div style="font-size:12px;color:rgba(255,255,255,.7);padding:5px 0;display:flex;flex-direction:column;justify-content:center"><b>{dt_name}</b><span style="font-size:10px;color:rgba(255,255,255,.32)">{n} días · +{avg_ex:.0f}% exceso prom.</span></div>'
            for k in K_COLS:
                pct     = info['coverage'].get(k, 0)
                is_cur  = abs(k - k_factor) < 0.01
                rows_html += f'<div style="{cov_color(pct, is_cur)}">{pct}%</div>'
        rows_html += '</div>'

        legend = (
            f'<div style="display:flex;gap:16px;margin-top:12px;font-size:11px;color:rgba(255,255,255,.4)">'
            f'<span><span style="color:rgba(100,230,170,.9)">■</span> ≥90% seguro</span>'
            f'<span><span style="color:#F59E0B">■</span> 75–89% aceptable</span>'
            f'<span><span style="color:rgba(215,75,65,.9)">■</span> &lt;75% riesgo de quiebre</span>'
            f'<span style="margin-left:auto;color:{GOLD}">← factor actual ({k_factor:.1f})</span>'
            f'</div>'
        )

        st.markdown(
            f'<div style="background:linear-gradient(135deg,{BG_3},{BG_2});border:1px solid rgba(255,255,255,.06);border-radius:14px;padding:20px 22px">'
            f'{header_html}{rows_html}{legend}</div>',
            unsafe_allow_html=True,
        )

        # Summary callout
        k_nearest = min(K_COLS, key=lambda x: abs(x - k_factor))
        avg_cov = round(sum(coverage_data[dt]['coverage'].get(k_nearest, 0)
                            for dt in coverage_data) / len(coverage_data))
        color_sum = 'rgba(100,230,170,.85)' if avg_cov >= 90 else ('#F59E0B' if avg_cov >= 75 else 'rgba(215,75,65,.9)')
        st.markdown(
            f'<div style="margin-top:12px;background:rgba(201,169,122,.06);border:1px solid rgba(201,169,122,.2);'
            f'border-left:3px solid {GOLD};border-radius:0 10px 10px 0;padding:12px 18px;font-size:13px;color:rgba(255,255,255,.7)">'
            f'Con factor <b style="color:{GOLD}">{k_factor:.1f}</b>, históricamente habrías cubierto el '
            f'<b style="color:{color_sum}">{avg_cov}%</b> de los días (promedio entre tipos). '
            f'{"✓ Cobertura sólida." if avg_cov >= 90 else ("⚠ Considera subir el factor en días críticos." if avg_cov >= 75 else "✗ Alto riesgo — considera subir el factor.")}'
            f'</div>',
            unsafe_allow_html=True,
        )
    else:
        st.info("Se necesitan al menos 2 días por tipo para calcular la cobertura.")

    # ── Últimos 7 días: Real vs Pronóstico ───────────────────────────────────
    st.markdown(f"""<div class="sec-hd" style="margin-top:28px">
      <span class="sec-t" style="font-size:18px">Últimos 7 Días · Real vs Pronóstico</span>
      <span class="sec-r"></span>
      <span class="sec-c">Factor actual {k_factor:.1f}σ · solo Fondos</span>
    </div>""", unsafe_allow_html=True)

    # Calcular real vs forecast para los últimos 7 días con datos
    last_days = sorted(history.keys(), reverse=True)[:7]

    if last_days:
        rows_7 = []
        for d in sorted(last_days):
            dt_name  = day_type(d)
            dm       = model.get(dt_name, {})
            actual_f = sum(qty for (_, name, cat), qty in history[d].items() if cat == 'Fondo')
            fc_f     = sum(
                s['mean'] + k_factor * (s['std'] if s['std'] is not None else s['mean'] * BUFFER_FALLBACK)
                for (name, cat), s in dm.items() if cat == 'Fondo'
            )
            diff     = actual_f - fc_f
            covered  = actual_f <= fc_f
            rows_7.append({
                'date': d, 'dt': dt_name,
                'actual': round(actual_f), 'fc': round(fc_f),
                'diff': round(diff), 'covered': covered,
            })

        # Chart
        labels   = [f"{r['dt'][:3]}<br>{r['date'].strftime('%d/%m')}" for r in rows_7]
        actuals  = [r['actual'] for r in rows_7]
        fcs      = [r['fc']     for r in rows_7]
        bar_colors = [
            'rgba(100,230,170,.75)' if r['covered'] else 'rgba(215,75,65,.75)'
            for r in rows_7
        ]

        fig7 = go.Figure()
        fig7.add_trace(go.Bar(
            x=labels, y=actuals, name='Real',
            marker=dict(color=bar_colors, cornerradius=4),
            hovertemplate='<b>%{x}</b><br>Real: <b>%{y}</b><extra></extra>',
        ))
        fig7.add_trace(go.Scatter(
            x=labels, y=fcs, name=f'Pronóstico (k={k_factor:.1f})',
            mode='lines+markers',
            line=dict(color=GOLD, width=2, dash='dot'),
            marker=dict(color=GOLD, size=8, symbol='diamond'),
            hovertemplate='<b>%{x}</b><br>Pronóstico: <b>%{y}</b><extra></extra>',
        ))
        layout7 = dict(PLOTLY_BASE)
        layout7.update(dict(
            height=300,
            barmode='group',
            showlegend=True,
            margin=dict(l=16, r=16, t=24, b=16),
            legend=dict(orientation='h', x=0, y=1.15, bgcolor='rgba(0,0,0,0)'),
            xaxis={**PLOTLY_BASE['xaxis'], 'tickfont': dict(size=11)},
            yaxis={**PLOTLY_BASE['yaxis'], 'title': 'Fondos'},
        ))
        fig7.update_layout(**layout7)
        st.plotly_chart(fig7, use_container_width=True, config={"displayModeBar": False})

        # Tabla resumen
        html_rows = ''
        for r in reversed(rows_7):   # más reciente primero
            icon  = '✓' if r['covered'] else '✗'
            c_icon = 'rgba(100,230,170,.9)' if r['covered'] else 'rgba(215,75,65,.9)'
            c_diff = 'rgba(100,230,170,.75)' if r['diff'] <= 0 else 'rgba(215,75,65,.75)'
            diff_txt = f'+{r["diff"]}' if r['diff'] > 0 else str(r['diff'])
            tooltip = 'Cubierto — el pronóstico fue suficiente' if r['covered'] else f'Déficit de {abs(r["diff"])} fondos'
            html_rows += (
                f'<tr title="{tooltip}">'
                f'<td style="color:rgba(255,255,255,.8);font-size:12px">'
                f'  {r["date"].strftime("%a %d/%m").capitalize()}'
                f'  <span style="font-size:10px;color:rgba(255,255,255,.35);margin-left:4px">{r["dt"]}</span>'
                f'</td>'
                f'<td style="text-align:right;font-family:var(--font-mono);font-size:13px;color:rgba(255,255,255,.9)">{r["actual"]:,}</td>'
                f'<td style="text-align:right;font-family:var(--font-mono);font-size:13px;color:{GOLD}">{r["fc"]:,}</td>'
                f'<td style="text-align:right;font-family:var(--font-mono);font-size:12px;color:{c_diff}">{diff_txt}</td>'
                f'<td style="text-align:center;color:{c_icon};font-weight:700;font-size:13px">{icon}</td>'
                f'</tr>'
            )
        n_ok  = sum(1 for r in rows_7 if r['covered'])
        n_bad = len(rows_7) - n_ok
        c_sum = 'rgba(100,230,170,.9)' if n_ok >= len(rows_7)*0.7 else 'rgba(215,75,65,.9)'
        st.markdown(
            f'<div class="prod-wrap"><table class="prod-table">'
            f'<thead><tr>'
            f'<th>Día</th>'
            f'<th class="r">Real</th>'
            f'<th class="r" style="color:{GOLD}">Pronóstico</th>'
            f'<th class="r">Diferencia</th>'
            f'<th style="text-align:center">Estado</th>'
            f'</tr></thead>'
            f'<tbody>{html_rows}</tbody>'
            f'<tfoot><tr><td colspan="5" style="padding:10px 16px;border-top:1px solid rgba(255,255,255,.06);'
            f'font-size:11px;color:rgba(255,255,255,.35)">'
            f'<span style="color:{c_sum};font-weight:700">{n_ok}/{len(rows_7)} días cubiertos</span>'
            f' con factor {k_factor:.1f}σ — '
            f'{"el modelo cubre bien la demanda reciente" if n_ok==len(rows_7) else f"en {n_bad} día(s) la demanda superó el pronóstico"}'
            f'</td></tr></tfoot>'
            f'</table></div>',
            unsafe_allow_html=True,
        )
    else:
        st.info("Sin datos históricos suficientes.")

# ──────────────────────────────────────────────────────────────────────────────
# TAB 3: TENDENCIAS
# ──────────────────────────────────────────────────────────────────────────────

with tab_tend:

    # Heatmap
    fig_heat = chart_heatmap(df)
    st.plotly_chart(fig_heat, use_container_width=True, config={"displayModeBar": False})

    st.divider()

    # Trend lines
    col_ctrl, _ = st.columns([2, 3])
    with col_ctrl:
        all_fondos = (df[df['category'] == 'Fondo']
                      .groupby('dish')['quantity'].mean()
                      .sort_values(ascending=False)
                      .head(20).index.tolist())
        selected_trend = st.multiselect(
            "Platos a graficar",
            all_fondos,
            default=all_fondos[:5],
            key="trend_dishes",
        )

    if selected_trend:
        fig_trend = chart_trend(df, selected_trend)
        st.plotly_chart(fig_trend, use_container_width=True, config={"displayModeBar": False})

    st.divider()

    # Weekly volume
    fig_weekly = chart_weekly_total(df)
    st.plotly_chart(fig_weekly, use_container_width=True, config={"displayModeBar": False})

# ──────────────────────────────────────────────────────────────────────────────
# TAB 4: PROMO AMEX
# ──────────────────────────────────────────────────────────────────────────────

with tab_promo:

    if n_promo == 0:
        st.info("Aún no hay datos de domingos con promo AMEX registrados en el modelo.")
    else:
        # Summary banner
        sunday_normal_fondos = round(sum(
            s['mean'] for (n, c), s in model.get('Domingo', {}).items() if c == 'Fondo'
        ))
        sunday_promo_fondos  = round(sum(
            s['mean'] for (n, c), s in model.get('Dom-Promo', {}).items() if c == 'Fondo'
        ))
        factor = sunday_promo_fondos / sunday_normal_fondos if sunday_normal_fondos else 0

        col_a, col_b, col_c = st.columns(3)
        with col_a:
            st.metric("Dom Normal (fondos avg)", f"{sunday_normal_fondos:,}")
        with col_b:
            st.metric("Dom-Promo AMEX (fondos avg)", f"{sunday_promo_fondos:,}",
                      delta=f"+{sunday_promo_fondos - sunday_normal_fondos:,} uds")
        with col_c:
            st.metric("Factor multiplicador", f"{factor:.2f}×",
                      delta=f"{n_promo} domingo{'s' if n_promo>1 else ''} con promo")

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        if n_promo <= 2:
            st.markdown("""<div class="alert-promo">
              <strong>⚠ Confianza estadística baja</strong> —
              Con pocas observaciones el factor puede variar significativamente.
              Cada domingo nuevo mejora la precisión del modelo.
            </div>""", unsafe_allow_html=True)

        # Promo comparison chart
        fig_promo = chart_promo_impact(model)
        st.plotly_chart(fig_promo, use_container_width=True, config={"displayModeBar": False})

        st.divider()

        # Per-dish factor table
        st.markdown(f"""<div class="sec-hd">
          <span class="sec-t" style="font-size:18px">Factor por Plato</span>
          <span class="sec-r"></span>
          <span class="sec-c">Domingo Normal vs Dom-Promo AMEX</span>
        </div>""", unsafe_allow_html=True)

        promo_items, normal_items = {}, {}
        for (n, c), s in model.get('Dom-Promo', {}).items():
            if c == 'Fondo':
                promo_items[n] = s['mean']
        for (n, c), s in model.get('Domingo', {}).items():
            if c == 'Fondo':
                normal_items[n] = s['mean']

        all_names = sorted(promo_items.keys(), key=lambda x: -promo_items.get(x, 0))
        rows_data  = []
        for name in all_names:
            p = promo_items.get(name, 0)
            n = normal_items.get(name, 0)
            f = p / n if n > 0 else 0
            rows_data.append({
                'Plato':          name,
                'Dom Normal (μ)': round(n, 1),
                'Dom Promo (μ)':  round(p, 1),
                'Factor':         round(f, 2),
            })

        df_factor = pd.DataFrame(rows_data).sort_values('Dom Promo (μ)', ascending=False)
        st.dataframe(
            df_factor,
            use_container_width=True,
            hide_index=True,
            column_config={
                'Factor': st.column_config.ProgressColumn(
                    'Factor ×', min_value=0, max_value=5, format="%.2f×"
                ),
                'Dom Promo (μ)': st.column_config.NumberColumn(format="%.1f"),
                'Dom Normal (μ)': st.column_config.NumberColumn(format="%.1f"),
            }
        )

# ──────────────────────────────────────────────────────────────────────────────
# ──────────────────────────────────────────────────────────────────────────────
# TAB 5: HISTORIAL — Real vs Pronóstico por plato
# ──────────────────────────────────────────────────────────────────────────────

with tab_hist:

    # Selector de fecha (últimos 30 días con datos)
    avail_dates = sorted(history.keys(), reverse=True)[:30]
    if not avail_dates:
        st.info("Sin datos históricos.")
    else:
        # Encabezado
        st.markdown("""<div class="sec-hd">
          <span class="sec-t" style="font-size:22px">Historial · Real vs Pronóstico</span>
          <span class="sec-r"></span>
          <span class="sec-c">Selecciona un día para ver el detalle por plato</span>
        </div>""", unsafe_allow_html=True)

        # Selector de día + info de contexto en columnas
        col_sel, col_info = st.columns([2, 3])
        with col_sel:
            date_options = {
                f"{DAYS_ES[d.weekday()]} {d.strftime('%d/%m/%Y')} — {day_type(d)}": d
                for d in avail_dates
            }
            sel_label = st.selectbox(
                "Día a revisar",
                list(date_options.keys()),
                label_visibility="collapsed",
                key="hist_date",
            )
            sel_date = date_options[sel_label]
            sel_dt   = day_type(sel_date)

        # Calcular métricas del día seleccionado
        actual_data = history[sel_date]
        dm = model.get(sel_dt, {})

        # Construir tabla detalle por plato
        detail_rows = []
        for (name, cat), stats in dm.items():
            fc_qty = stats['mean'] + k_factor * (stats['std'] if stats['std'] is not None else stats['mean'] * BUFFER_FALLBACK)
            actual_qty = next(
                (qty for (_, n, c), qty in actual_data.items() if n == name and c == cat),
                0.0,
            )
            diff = actual_qty - fc_qty
            detail_rows.append({
                'name': name, 'cat': cat,
                'actual': actual_qty, 'fc': fc_qty,
                'diff': diff, 'covered': actual_qty <= fc_qty,
                'mu': stats['mean'], 'std': stats['std'] or 0,
                'n': stats.get('n', 0),
            })

        total_actual_f = sum(r['actual'] for r in detail_rows if r['cat'] == 'Fondo')
        total_fc_f     = sum(r['fc']     for r in detail_rows if r['cat'] == 'Fondo')
        total_actual   = sum(r['actual'] for r in detail_rows)
        total_fc       = sum(r['fc']     for r in detail_rows)
        n_covered      = sum(1 for r in detail_rows if r['covered'])
        pct_covered    = round(n_covered / len(detail_rows) * 100) if detail_rows else 0

        with col_info:
            c1, c2, c3 = st.columns(3)
            with c1:
                st.metric("Fondos reales", f"{round(total_actual_f):,}",
                          delta=f"{round(total_actual_f - total_fc_f):+,} vs pronóstico")
            with c2:
                st.metric("Total platos reales", f"{round(total_actual):,}",
                          delta=f"{round(total_actual - total_fc):+,} vs pronóstico")
            with c3:
                cov_color = "normal" if pct_covered >= 75 else "inverse"
                st.metric("Platos cubiertos", f"{pct_covered}%",
                          delta=f"{n_covered}/{len(detail_rows)} items")

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        # Tabla detalle por categoría
        for cat in CAT_ORDER:
            cat_rows = sorted(
                [r for r in detail_rows if r['cat'] == cat],
                key=lambda r: -r['actual'],
            )
            if not cat_rows:
                continue

            color = CAT_HEX[cat]
            total_act_cat = round(sum(r['actual'] for r in cat_rows))
            total_fc_cat  = round(sum(r['fc']     for r in cat_rows))
            n_cov_cat     = sum(1 for r in cat_rows if r['covered'])

            st.markdown(f"""<div class="sec-hd" style="margin-top:16px">
              <span class="sec-t" style="color:{color};font-size:18px">{cat}</span>
              <span class="sec-r"></span>
              <span class="sec-c">{len(cat_rows)} platos · real {total_act_cat:,} vs pron {total_fc_cat:,} · {n_cov_cat}/{len(cat_rows)} cubiertos</span>
            </div>""", unsafe_allow_html=True)

            rows_html = ''
            for r in cat_rows:
                diff_v   = round(r['diff'])
                diff_txt = f'+{diff_v}' if diff_v > 0 else str(diff_v)
                if r['covered']:
                    st_icon = f'<span style="color:rgba(100,230,170,.9);font-weight:700">✓</span>'
                    diff_col = 'rgba(100,230,170,.7)'
                else:
                    st_icon = f'<span style="color:rgba(215,75,65,.9);font-weight:700">✗</span>'
                    diff_col = 'rgba(215,75,65,.8)'

                # Mini bar showing actual vs forecast
                max_v  = max(r['actual'], r['fc'], 1)
                act_pct = round(r['actual'] / max_v * 100)
                fc_pct  = round(r['fc']     / max_v * 100)

                rows_html += f"""<tr>
                  <td class="td-name">{r['name']}</td>
                  <td class="td-num" style="color:rgba(255,255,255,.85);font-weight:600">{round(r['actual'])}</td>
                  <td class="td-num" style="color:{GOLD}">{round(r['fc'])}</td>
                  <td class="td-num" style="color:{diff_col};font-weight:600">{diff_txt}</td>
                  <td class="td-num" style="color:rgba(255,255,255,.4);font-size:10px">{round(r['mu'])} ±{round(r['std'])}</td>
                  <td class="td-bar" style="padding-right:12px">
                    <div style="position:relative;height:20px;display:flex;align-items:center;gap:3px">
                      <div style="width:{act_pct}%;max-width:100%;height:8px;background:{'rgba(100,230,170,.6)' if r['covered'] else 'rgba(215,75,65,.6)'};border-radius:3px;transition:width .4s"></div>
                      <div style="position:absolute;left:0;width:{fc_pct}%;max-width:100%;height:3px;background:{GOLD};border-radius:1px;opacity:.7;top:50%;transform:translateY(6px)"></div>
                    </div>
                  </td>
                  <td style="text-align:center;width:32px">{st_icon}</td>
                </tr>"""

            st.markdown(
                f'<div class="prod-wrap"><table class="prod-table">'
                f'<thead><tr>'
                f'<th>Plato</th>'
                f'<th class="r">Real</th>'
                f'<th class="r" style="color:{GOLD}">Pron. (k={k_factor:.1f})</th>'
                f'<th class="r">Dif.</th>'
                f'<th class="r">μ ±σ</th>'
                f'<th class="bh">Real / Pron.</th>'
                f'<th style="text-align:center"></th>'
                f'</tr></thead>'
                f'<tbody>{rows_html}</tbody>'
                f'</table></div>',
                unsafe_allow_html=True,
            )

        # Platos reales sin modelo (vendidos ese día pero no en el modelo del tipo de día)
        extras = []
        for (_, name, cat), qty in actual_data.items():
            if (name, cat) not in dm and cat in CAT_ORDER:
                name_up = name.upper()
                if not any(name_up.endswith(s.upper()) for s in _EXCLUDE_SUFFIXES):
                    extras.append((name, cat, qty))
        if extras:
            st.markdown(f"""<div class="sec-hd" style="margin-top:16px">
              <span class="sec-t" style="color:rgba(245,158,11,.8);font-size:16px">Sin modelo propio este día</span>
              <span class="sec-r"></span>
              <span class="sec-c">Platos vendidos sin historial para {sel_dt}</span>
            </div>""", unsafe_allow_html=True)
            extras_html = ''.join(
                f'<tr><td class="td-name">{n}</td>'
                f'<td class="td-num" style="color:rgba(255,255,255,.85)">{round(q)}</td>'
                f'<td class="td-num" style="color:rgba(255,255,255,.35)">—</td>'
                f'<td class="td-num" style="color:rgba(245,158,11,.7)">sin datos</td>'
                f'<td></td><td></td><td></td></tr>'
                for n, c, q in sorted(extras, key=lambda x: -x[2])
            )
            st.markdown(
                f'<div class="prod-wrap"><table class="prod-table">'
                f'<thead><tr><th>Plato</th><th class="r">Real</th>'
                f'<th class="r" style="color:{GOLD}">Pron.</th>'
                f'<th class="r">Dif.</th><th></th><th></th><th></th></tr></thead>'
                f'<tbody>{extras_html}</tbody></table></div>',
                unsafe_allow_html=True,
            )

# ──────────────────────────────────────────────────────────────────────────────
# TAB 6: COMPRAS
# ──────────────────────────────────────────────────────────────────────────────

with tab_compras:

    # Aggregate ingredients across all forecast days
    _ing = defaultdict(lambda: {
        'categoria': '', 'temperatura': 'ambiente',
        'dias_despacho': [], 'total_g': 0.0, 'platos': set()
    })
    _sin_receta = {}  # {(name, cat): total_fc}

    for _d, _dt, _dm, _fb in day_fc:
        for (_dname, _dcat), _stats in _dm.items():
            _fc = dish_fc(_stats, k_factor)
            if _fc <= 0:
                continue
            _sku = name_to_sku.get(_dname)
            if _sku and _sku in RECETAS:
                for _ri in RECETAS[_sku]['ingredientes']:
                    _k = _ri['nombre']
                    _ing[_k]['categoria'] = _ri['categoria']
                    _ing[_k]['temperatura'] = _ri['temperatura']
                    # Union of delivery days (ingredient can arrive via any of them)
                    _dias = _ri.get('dias_despacho', [])
                    _ing[_k]['dias_despacho'] = sorted(
                        set(_ing[_k]['dias_despacho']) | set(_dias),
                        key=lambda x: ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado'].index(x)
                        if x in ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado'] else 9
                    )
                    _ing[_k]['total_g'] += _ri['cantidad'] * _fc
                    _ing[_k]['platos'].add(_dname)
            else:
                _sr_key = (_dname, _dcat)
                _sin_receta[_sr_key] = _sin_receta.get(_sr_key, 0) + _fc

    # Category config
    _CAT_ORDER = [
        "Proteínas del mar", "Mariscos", "Carnes", "Aves",
        "Frutas", "Lácteos", "Huevos",
        "Pasta y cereales", "Repostería", "Helados", "Aceites y condimentos",
    ]
    _CAT_ICON = {
        "Proteínas del mar": "🐟", "Mariscos": "🦐", "Carnes": "🥩", "Aves": "🍗",
        "Frutas": "🍋", "Lácteos": "🧀", "Huevos": "🥚",
        "Pasta y cereales": "🌾", "Repostería": "🍫", "Helados": "🍨",
        "Aceites y condimentos": "🫒",
    }

    # Próximo día de despacho a partir de hoy (excl. domingo)
    _DNUM = {'Lunes':0,'Martes':1,'Miércoles':2,'Jueves':3,'Viernes':4,'Sábado':5}
    _NDIA = {v: k for k, v in _DNUM.items()}

    def _prox_despacho(dias):
        """Devuelve (días_hasta_despacho, nombre_día). Domingo = sin despacho."""
        if not dias:
            return 99, '—'
        _hoy_wd = datetime.now().weekday()   # 0=lun … 5=sab, 6=dom
        if _hoy_wd == 6:                     # domingo: contar desde lunes
            _hoy_wd = 7
        _wds = sorted(_DNUM[d] for d in dias if d in _DNUM)
        if not _wds:
            return 99, '—'
        for _wd in _wds:
            if _wd >= _hoy_wd:
                return _wd - _hoy_wd, _NDIA[_wd]
        # Próxima semana
        return 7 - _hoy_wd + _wds[0], _NDIA[_wds[0]]

    def _urgency_desp(dias):
        _n, _d = _prox_despacho(dias)
        if _n == 0:   return f"⚡ LLEGA HOY ({_d})",       "#A5D6A7"
        if _n == 1:   return f"📋 LLEGA MAÑANA ({_d})",    "#FBBF24"
        if _n <= 3:   return f"📦 En {_n} días ({_d})",    "#7EB8F7"
        if _n < 99:   return f"🗓 En {_n} días ({_d})",    "rgba(255,255,255,.42)"
        return "— Sin días configurados", "rgba(255,255,255,.22)"

    def _fmt_g(g):
        if g >= 1000:
            return f"{g/1000:.1f} kg"
        return f"{round(g):,} g"

    # Group by category
    _by_cat = defaultdict(list)
    for _iname, _idata in _ing.items():
        _by_cat[_idata['categoria']].append((_iname, _idata))
    for _c in _by_cat:
        _by_cat[_c].sort(key=lambda x: -x[1]['total_g'])

    # Header
    st.markdown(
        f'<div class="section-title">LISTA DE COMPRAS · {horizon} DÍA{"S" if horizon > 1 else ""} · '
        f'{total_all:,} PLATOS · {len(_ing)} INGREDIENTES</div>',
        unsafe_allow_html=True,
    )

    if not RECETAS:
        st.warning("recetas.json no encontrado en la carpeta del proyecto. Asegúrate de que el archivo exista.")
    elif not _ing:
        st.info("No se calcularon ingredientes. Verifica que el pronóstico tenga platos con receta configurada.")
    else:
        _all_rows_excel = []

        for _cat in _CAT_ORDER:
            if _cat not in _by_cat:
                continue
            _items = _by_cat[_cat]
            _max_g = _items[0][1]['total_g'] if _items else 1
            # Urgencia de la categoría = próximo despacho más temprano de sus ingredientes
            _min_dias_cat = min((_prox_despacho(i[1]['dias_despacho'])[0] for i in _items), default=99)
            _best_dias = next((i[1]['dias_despacho'] for i in _items
                               if _prox_despacho(i[1]['dias_despacho'])[0] == _min_dias_cat), [])
            _urg_txt, _urg_c = _urgency_desp(_best_dias)
            _icon = _CAT_ICON.get(_cat, '·')

            _rows_html = ''
            for _iname, _idata in _items:
                _qty_str = _fmt_g(_idata['total_g'])
                _bar_w = max(4, round((_idata['total_g'] / _max_g) * 130))
                _np = len(_idata['platos'])
                _i_dias = _idata.get('dias_despacho', [])
                _i_urg_txt, _i_urg_c = _urgency_desp(_i_dias)
                _dias_str = ' · '.join(d[:3] for d in _i_dias) if _i_dias else '—'
                _rows_html += (
                    f'<tr>'
                    f'<td class="ci-name">{_iname}</td>'
                    f'<td class="ci-qty" style="color:{_i_urg_c}">{_qty_str}</td>'
                    f'<td class="ci-bar"><div class="ci-bar-bg">'
                    f'<div class="ci-bar-fill" style="width:{_bar_w}px;background:{_i_urg_c}"></div>'
                    f'</div></td>'
                    f'<td class="ci-meta" style="color:{_i_urg_c}">{_dias_str}</td>'
                    f'<td class="ci-platos">{_np}p</td>'
                    f'</tr>'
                )
                _n_prox, _d_prox = _prox_despacho(_i_dias)
                _all_rows_excel.append({
                    'Categoría': _cat,
                    'Ingrediente': _iname,
                    'Cantidad': round(_idata['total_g'] / 1000, 3) if _idata['total_g'] >= 1000 else round(_idata['total_g']),
                    'Unidad': 'kg' if _idata['total_g'] >= 1000 else 'g',
                    'Temperatura': _idata['temperatura'],
                    'Días de despacho': ', '.join(_i_dias),
                    'Próximo despacho': f'En {_n_prox} días ({_d_prox})' if _n_prox < 99 else '—',
                    'N° platos pronóstico': _np,
                })

            st.markdown(f"""
<div style="margin-bottom:20px">
  <div style="display:flex;align-items:center;gap:10px;margin-bottom:8px;padding:0 2px">
    <span style="font-size:18px">{_icon}</span>
    <span style="font-family:var(--font-serif);font-size:15px;color:var(--t1);font-weight:300">{_cat}</span>
    <div style="flex:1;height:1px;background:rgba(255,255,255,.06);margin:0 8px"></div>
    <span style="font-size:9px;font-weight:600;letter-spacing:.1em;color:{_urg_c}">{_urg_txt}</span>
  </div>
  <div style="background:rgba(255,255,255,.025);border:1px solid rgba(255,255,255,.06);border-radius:10px;overflow:hidden">
    <table class="ci-table" style="width:100%;border-collapse:collapse">
      <tbody>{_rows_html}</tbody>
    </table>
  </div>
</div>""", unsafe_allow_html=True)

        # Excel export
        if _all_rows_excel:
            _df_comp = pd.DataFrame(_all_rows_excel)
            _buf = io.BytesIO()
            with pd.ExcelWriter(_buf, engine='openpyxl') as _xwr:
                _df_comp.to_excel(_xwr, sheet_name='Compras', index=False)
                _ws_x = _xwr.sheets['Compras']
                for _col in _ws_x.columns:
                    _ws_x.column_dimensions[_col[0].column_letter].width = 24
            _buf.seek(0)
            st.download_button(
                "📥  Descargar lista de compras Excel",
                data=_buf,
                file_name=f"compras_{datetime.now().strftime('%Y%m%d')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    # Dishes without recipe
    if _sin_receta:
        _sr_sorted = sorted(_sin_receta.items(), key=lambda x: -x[1])
        with st.expander(f"⚠  {len(_sin_receta)} platos sin receta configurada"):
            st.markdown(
                '<div style="font-size:11px;color:rgba(255,255,255,.45);margin-bottom:10px">'
                'Estos platos están en el pronóstico pero no tienen ingredientes en recetas.json. '
                'Sus materias primas NO están incluidas en la lista de compras.</div>',
                unsafe_allow_html=True,
            )
            for (_dname, _dcat), _fc in _sr_sorted:
                st.markdown(
                    f'<div style="font-size:12px;padding:3px 0;color:rgba(255,255,255,.55)">'
                    f'<span style="color:rgba(201,169,122,.6)">{_dcat}</span> · {_dname} '
                    f'<span style="color:rgba(255,255,255,.28)">({round(_fc):,} porciones)</span></div>',
                    unsafe_allow_html=True,
                )

# ──────────────────────────────────────────────────────────────────────────────
# TAB 7: RECETAS — Editor de ingredientes por plato
# ──────────────────────────────────────────────────────────────────────────────

with tab_recetas:

    _CAT_OPTS_ED  = [
        "Proteínas del mar", "Mariscos", "Carnes", "Aves",
        "Frutas", "Lácteos", "Huevos",
        "Pasta y cereales", "Repostería", "Helados", "Aceites y condimentos",
    ]
    _TEMP_OPTS   = ["refrigerado", "congelado", "ambiente"]
    _UNID_OPTS   = ["g", "ml", "unidades"]
    _DIAS_OPTS   = ["Lunes", "Martes", "Miércoles", "Jueves", "Viernes", "Sábado"]

    # Build complete dish list from history: {label: (sku, name, cat)}
    _ed_dishes = {}
    for _hdata in history.values():
        for (_hcode, _hname, _hcat) in _hdata:
            _lbl = f"{_hcat} · {_hname}"
            if _lbl not in _ed_dishes:
                _ed_dishes[_lbl] = (_hcode, _hname, _hcat)
    _ed_opts = sorted(_ed_dishes.keys())

    # Summary bar: how many dishes have recipes
    _n_total  = len(_ed_dishes)
    _n_config = sum(1 for _, (sku, _, _) in _ed_dishes.items() if sku in RECETAS)
    _pct_conf = round(_n_config / _n_total * 100) if _n_total else 0

    st.markdown(f"""
<div style="display:flex;align-items:center;gap:20px;margin-bottom:20px;
            background:rgba(255,255,255,.025);border:1px solid rgba(255,255,255,.06);
            border-radius:10px;padding:14px 20px">
  <div>
    <div style="font-size:22px;font-weight:700;color:{GOLD}">{_n_config}</div>
    <div style="font-size:10px;color:rgba(255,255,255,.4);letter-spacing:.08em">CON RECETA</div>
  </div>
  <div style="width:1px;height:32px;background:rgba(255,255,255,.08)"></div>
  <div>
    <div style="font-size:22px;font-weight:700;color:rgba(255,255,255,.55)">{_n_total - _n_config}</div>
    <div style="font-size:10px;color:rgba(255,255,255,.4);letter-spacing:.08em">SIN RECETA</div>
  </div>
  <div style="flex:1;margin:0 12px">
    <div style="height:6px;background:rgba(255,255,255,.07);border-radius:3px;overflow:hidden">
      <div style="height:6px;width:{_pct_conf}%;background:{GOLD};border-radius:3px;transition:width .6s"></div>
    </div>
    <div style="font-size:10px;color:rgba(255,255,255,.35);margin-top:4px">{_pct_conf}% configurado</div>
  </div>
  <div style="font-size:11px;color:rgba(255,255,255,.35);max-width:220px;line-height:1.5">
    Ingresa los ingredientes reales de cada plato.<br>
    Las cantidades son <strong>por porción individual</strong>.
  </div>
</div>""", unsafe_allow_html=True)

    # Dish selector + editor
    col_ed_left, col_ed_right = st.columns([1, 2])

    with col_ed_left:
        st.markdown("##### Seleccionar plato")
        _ed_sel_label = st.selectbox(
            "Plato", _ed_opts, key="ed_dish_sel",
            label_visibility="collapsed",
        )
        _ed_sku, _ed_name, _ed_cat = _ed_dishes[_ed_sel_label]

        # Status badge
        _ed_has = _ed_sku in RECETAS
        _ed_n   = len(RECETAS.get(_ed_sku, {}).get('ingredientes', []))
        if _ed_has:
            st.markdown(
                f'<div style="margin:6px 0 12px;font-size:11px;color:#A5D6A7">'
                f'✓ {_ed_n} ingrediente{"s" if _ed_n!=1 else ""} configurado{"s" if _ed_n!=1 else ""}</div>',
                unsafe_allow_html=True,
            )
        else:
            st.markdown(
                '<div style="margin:6px 0 12px;font-size:11px;color:#F87171">'
                '⚠ Sin receta · excluido de lista de compras</div>',
                unsafe_allow_html=True,
            )

        # Quick stats: dishes by category that still need recipes
        st.divider()
        st.markdown("##### Estado por categoría")
        _by_cat_status = defaultdict(lambda: [0, 0])  # {cat: [con_receta, total]}
        for _lbl2, (_sku2, _nm2, _ct2) in _ed_dishes.items():
            _by_cat_status[_ct2][1] += 1
            if _sku2 in RECETAS:
                _by_cat_status[_ct2][0] += 1
        for _cat2 in CAT_ORDER:
            if _cat2 not in _by_cat_status:
                continue
            _c_ok, _c_tot = _by_cat_status[_cat2]
            _c_pct = round(_c_ok / _c_tot * 100) if _c_tot else 0
            _c_col = '#A5D6A7' if _c_pct == 100 else ('#FBBF24' if _c_pct >= 50 else '#F87171')
            st.markdown(
                f'<div style="display:flex;justify-content:space-between;align-items:center;'
                f'font-size:11px;padding:3px 0;border-bottom:1px solid rgba(255,255,255,.04)">'
                f'<span style="color:rgba(255,255,255,.65)">{_cat2}</span>'
                f'<span style="color:{_c_col}">{_c_ok}/{_c_tot}</span></div>',
                unsafe_allow_html=True,
            )

    with col_ed_right:
        import uuid as _uuid

        st.markdown(f"##### Receta: {_ed_name}")
        st.markdown(
            f'<div style="font-size:10px;color:rgba(255,255,255,.35);margin-bottom:12px;'
            f'letter-spacing:.06em">{_ed_cat.upper()} · SKU {_ed_sku}</div>',
            unsafe_allow_html=True,
        )

        # Session-state key for this dish's ingredient list
        _sk = f'ed_ings_{_ed_sku}'

        # Reset state when user switches to a different dish
        if st.session_state.get('_ed_last_sku') != _ed_sku:
            st.session_state['_ed_last_sku'] = _ed_sku
            _base = RECETAS.get(_ed_sku, {}).get('ingredientes', [])
            st.session_state[_sk] = [
                {**i, '_id': _uuid.uuid4().hex[:8]} for i in _base
            ]

        _ings_ss = st.session_state[_sk]

        # Column headers (sin lead_time; días de despacho van en sub-fila)
        _hcols = st.columns([3.5, 1.5, 1, 2.5, 2, 0.4])
        for _hc, _hl in zip(_hcols, ['Ingrediente','Cantidad','Unidad',
                                      'Categoría','Temperatura','']):
            _hc.markdown(
                f'<div style="font-size:9px;color:rgba(255,255,255,.38);'
                f'letter-spacing:.1em;padding-bottom:5px">{_hl.upper()}</div>',
                unsafe_allow_html=True,
            )

        # One ingredient per "card" (dos filas: datos + días despacho)
        _del_id = None
        for _ing in _ings_ss:
            _id = _ing['_id']

            # Fila 1: campos principales
            _c0,_c1,_c2,_c3,_c4,_c5 = st.columns([3.5, 1.5, 1, 2.5, 2, 0.4])
            _c0.text_input('n', key=f'n_{_id}',
                           value=_ing.get('nombre',''),
                           placeholder='Nombre del ingrediente',
                           label_visibility='collapsed')
            _c1.number_input('q', key=f'q_{_id}',
                             value=float(_ing.get('cantidad', 100)),
                             min_value=0.0, step=5.0, format='%.0f',
                             label_visibility='collapsed')
            _u_i = _UNID_OPTS.index(_ing['unidad']) if _ing.get('unidad') in _UNID_OPTS else 0
            _c2.selectbox('u', _UNID_OPTS, index=_u_i, key=f'u_{_id}',
                          label_visibility='collapsed')
            _c_i = _CAT_OPTS_ED.index(_ing['categoria']) if _ing.get('categoria') in _CAT_OPTS_ED else 0
            _c3.selectbox('c', _CAT_OPTS_ED, index=_c_i, key=f'c_{_id}',
                          label_visibility='collapsed')
            _t_i = _TEMP_OPTS.index(_ing['temperatura']) if _ing.get('temperatura') in _TEMP_OPTS else 0
            _c4.selectbox('t', _TEMP_OPTS, index=_t_i, key=f't_{_id}',
                          label_visibility='collapsed')
            if _c5.button('✕', key=f'x_{_id}', help='Eliminar ingrediente'):
                _del_id = _id

            # Fila 2: días de despacho (ancho completo, indentado)
            _lbl_col, _day_col = st.columns([1.2, 7.8])
            _lbl_col.markdown(
                '<div style="font-size:9px;color:rgba(255,255,255,.32);'
                'letter-spacing:.08em;padding-top:10px;text-align:right">DESPACHO</div>',
                unsafe_allow_html=True,
            )
            _default_dias = [d for d in _ing.get('dias_despacho', []) if d in _DIAS_OPTS]
            _day_col.multiselect(
                'd', _DIAS_OPTS, default=_default_dias,
                key=f'd_{_id}', label_visibility='collapsed',
                placeholder='Sin días configurados (no llega)',
            )
            st.markdown(
                '<hr style="border:none;border-top:1px solid rgba(255,255,255,.05);margin:2px 0 8px">',
                unsafe_allow_html=True,
            )

        # Process delete (outside the loop to avoid mid-loop rerun issues)
        if _del_id:
            st.session_state[_sk] = [i for i in _ings_ss if i['_id'] != _del_id]
            st.rerun()

        # Action buttons
        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)
        _ba, _bs, _bd, _ = st.columns([1.1, 1, 1, 1.3])

        if _ba.button('＋ Ingrediente', use_container_width=True, key='btn_add_ing'):
            st.session_state[_sk].append({
                'nombre': '', 'cantidad': 100.0, 'unidad': 'g',
                'categoria': _CAT_OPTS_ED[0], 'temperatura': 'refrigerado',
                'lead_time': 1, '_id': _uuid.uuid4().hex[:8],
            })
            st.rerun()

        if _bs.button('💾 Guardar', use_container_width=True,
                      type='primary', key='btn_save_rec'):
            _final_ings = []
            for _ing in st.session_state[_sk]:
                _id = _ing['_id']
                _nm = str(st.session_state.get(f'n_{_id}', '') or '').strip()
                if not _nm:
                    continue
                _final_ings.append({
                    'nombre':        _nm,
                    'cantidad':      float(st.session_state.get(f'q_{_id}', 0) or 0),
                    'unidad':        str(st.session_state.get(f'u_{_id}', 'g') or 'g'),
                    'categoria':     str(st.session_state.get(f'c_{_id}', _CAT_OPTS_ED[0]) or _CAT_OPTS_ED[0]),
                    'temperatura':   str(st.session_state.get(f't_{_id}', 'refrigerado') or 'refrigerado'),
                    'dias_despacho': list(st.session_state.get(f'd_{_id}', []) or []),
                })
            _recetas_raw[_ed_sku] = {'nombre': _ed_name, 'ingredientes': _final_ings}
            RECETAS[_ed_sku]      = {'nombre': _ed_name, 'ingredientes': _final_ings}
            st.session_state[_sk] = [
                {**i, '_id': _uuid.uuid4().hex[:8]} for i in _final_ings
            ]
            try:
                with open(_RECETAS_PATH, 'w', encoding='utf-8') as _wf:
                    json.dump(_recetas_raw, _wf, ensure_ascii=False, indent=2)
                st.success(f'✓ Guardado · {len(_final_ings)} ingredientes')
            except Exception as _se:
                st.error(f'Error al guardar: {_se}')

        if _ed_has:
            if _bd.button('🗑 Borrar', use_container_width=True, key='btn_del_rec'):
                _recetas_raw.pop(_ed_sku, None)
                RECETAS.pop(_ed_sku, None)
                st.session_state[_sk] = []
                try:
                    with open(_RECETAS_PATH, 'w', encoding='utf-8') as _wf:
                        json.dump(_recetas_raw, _wf, ensure_ascii=False, indent=2)
                    st.warning(f"Receta de '{_ed_name}' eliminada.")
                    st.rerun()
                except Exception as _de:
                    st.error(f'Error: {_de}')

# ──────────────────────────────────────────────────────────────────────────────
# TAB 8: VENTAS — Análisis financiero YoY
# ──────────────────────────────────────────────────────────────────────────────

with tab_ventas:

    if _vdf.empty:
        st.warning("No se encontraron archivos XLS en la carpeta `Ventas/`.")
    else:
        # ── Aggregations ──────────────────────────────────────────────────────
        _vm = (
            _vdf.groupby(['año', 'mes'])['bruto'].sum()
            .reset_index().rename(columns={'bruto': 'total'})
        )
        _vc = (
            _vdf.groupby(['año', 'mes', 'categoria'])['bruto'].sum()
            .reset_index()
        )
        _vq = (
            _vdf.groupby(['año', 'mes', 'categoria'])['cantidad'].sum()
            .reset_index()
        )
        _vd = (
            _vdf.groupby(['año', 'mes'])['descuento'].sum()
            .reset_index()
        )

        _v25 = _vm[_vm['año'] == 2025].set_index('mes')['total']
        _v26 = _vm[_vm['año'] == 2026].set_index('mes')['total']
        _common = sorted(set(_v25.index) & set(_v26.index))

        _tot25 = _v25[_common].sum()
        _tot26 = _v26[_common].sum()
        _delta_pct = (_tot26 - _tot25) / _tot25 * 100 if _tot25 else 0

        _yoy = {m: (_v26[m] - _v25[m]) / _v25[m] * 100 for m in _common}
        _best_m  = max(_yoy, key=_yoy.get)
        _worst_m = min(_yoy, key=_yoy.get)
        _n_pos   = sum(1 for v in _yoy.values() if v >= 0)

        # ── KPI Cards ─────────────────────────────────────────────────────────
        _arr, _acol = _delta_arrow(_delta_pct)
        _kc1, _kc2, _kc3, _kc4 = st.columns(4)

        for _col, _title, _val, _sub, _scol in [
            (_kc1, "VENTAS 2026", _fmt_clp(_tot26), f"Ene–{_MESES_NOM_V[max(_common)]} ({len(_common)} meses)", GOLD),
            (_kc2, "vs 2025 (mismo período)", f"{_arr} {abs(_delta_pct):.1f}%",
             f"2025: {_fmt_clp(_tot25)}", _acol),
            (_kc3, f"MEJOR MES  {_MESES_NOM_V[_best_m]} 26",
             f"+{_yoy[_best_m]:.1f}%", f"vs {_MESES_NOM_V[_best_m]} 25: {_fmt_clp(_v25[_best_m])} → {_fmt_clp(_v26[_best_m])}", "#5CE8D4"),
            (_kc4, f"PEOR MES  {_MESES_NOM_V[_worst_m]} 26",
             f"{_yoy[_worst_m]:.1f}%", f"vs {_MESES_NOM_V[_worst_m]} 25: {_fmt_clp(_v25[_worst_m])} → {_fmt_clp(_v26[_worst_m])}", "#F7A8D0"),
        ]:
            _col.markdown(f"""
            <div style="background:{BG_2};border:1px solid {BORDER};border-radius:10px;
                        padding:18px 20px;margin-bottom:4px">
              <div style="font-size:10px;letter-spacing:.1em;color:{SUBTLE};
                          font-family:var(--font-sans);margin-bottom:6px">{_title}</div>
              <div style="font-size:26px;font-family:var(--font-serif);
                          color:{_scol};font-weight:600;line-height:1.1">{_val}</div>
              <div style="font-size:11px;color:{MUTED};margin-top:6px">{_sub}</div>
            </div>""", unsafe_allow_html=True)

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        # ── Monthly Comparison Chart ──────────────────────────────────────────
        _all_mes_2025 = sorted(_v25.index)
        _all_mes_2026 = sorted(_v26.index)

        _fig_main = go.Figure()
        _fig_main.add_trace(go.Bar(
            name='2025',
            x=[_MESES_NOM_V[m] for m in _all_mes_2025],
            y=[_v25[m]/1e6 for m in _all_mes_2025],
            marker_color='rgba(201,169,122,0.35)',
            marker_line_color='rgba(201,169,122,0.5)',
            marker_line_width=1,
            hovertemplate='%{x} 2025<br><b>%{customdata}</b><extra></extra>',
            customdata=[_fmt_clp(_v25[m]) for m in _all_mes_2025],
        ))
        _fig_main.add_trace(go.Bar(
            name='2026',
            x=[_MESES_NOM_V[m] for m in _all_mes_2026],
            y=[_v26[m]/1e6 for m in _all_mes_2026],
            marker_color=GOLD,
            marker_line_color=GOLD_2,
            marker_line_width=1,
            hovertemplate='%{x} 2026<br><b>%{customdata}</b><extra></extra>',
            customdata=[_fmt_clp(_v26[m]) for m in _all_mes_2026],
        ))

        # YoY % overlay on secondary axis
        _fig_main.add_trace(go.Scatter(
            name='Var. % YoY',
            x=[_MESES_NOM_V[m] for m in _common],
            y=[_yoy[m] for m in _common],
            yaxis='y2',
            mode='lines+markers+text',
            line=dict(color='rgba(255,255,255,0.55)', width=1.5, dash='dot'),
            marker=dict(
                size=9,
                color=[('#5CE8D4' if _yoy[m] >= 0 else '#F7A8D0') for m in _common],
                line=dict(color=BG, width=1.5),
            ),
            text=[f"{_yoy[m]:+.0f}%" for m in _common],
            textposition='top center',
            textfont=dict(size=10, color='rgba(255,255,255,0.75)'),
            hovertemplate='%{x}<br>Variación YoY: <b>%{y:.1f}%</b><extra></extra>',
        ))

        _ylayout = {**PLOTLY_BASE}
        _ylayout.update(
            title=dict(text='Ventas Mensuales — 2025 vs 2026', **PLOTLY_BASE['title']),
            barmode='group',
            bargap=0.25,
            bargroupgap=0.05,
            yaxis=dict(**PLOTLY_BASE['yaxis'],
                        title=dict(text='Ventas (millones CLP)', font=dict(size=11, color=MUTED)),
                        tickformat=',.0f', ticksuffix='M'),
            yaxis2=dict(overlaying='y', side='right',
                        title=dict(text='Variación % YoY', font=dict(size=11, color=MUTED)),
                        tickformat='+.0f', ticksuffix='%',
                        gridcolor='rgba(0,0,0,0)',
                        zerolinecolor='rgba(255,255,255,0.08)',
                        zeroline=True,
                        tickfont=dict(size=10)),
            legend=dict(**PLOTLY_BASE['legend'], orientation='h', x=0, y=1.12),
            margin=dict(l=16, r=64, t=64, b=16),
            height=400,
        )
        _fig_main.update_layout(**_ylayout)

        st.plotly_chart(_fig_main, use_container_width=True)

        # ── Second Row: Category Mix + Monthly table ──────────────────────────
        _col_mix, _col_tbl = st.columns([1, 1])

        # Category Mix Donut comparison
        with _col_mix:
            _CATS_MAIN = ['Alimentos', 'Bebidas S/Alcohol', 'Bebidas C/Alcohol', 'Vinos']
            _CAT_COLORS = [GOLD, '#7EB8F7', '#5CE8D4', '#C4B5FD']

            _vals25 = _cat_totals(_vc, _CATS_MAIN, 2025, _common)
            _vals26 = _cat_totals(_vc, _CATS_MAIN, 2026, _common)

            _fig_mix = go.Figure()
            _fig_mix.add_trace(go.Pie(
                labels=_CATS_MAIN, values=_vals25,
                name='2025',
                domain=dict(x=[0, 0.46]),
                hole=0.55,
                marker_colors=_CAT_COLORS,
                textinfo='none',
                hovertemplate='%{label}<br>2025: <b>%{value:,.0f}</b> (%{percent})<extra></extra>',
            ))
            _fig_mix.add_trace(go.Pie(
                labels=_CATS_MAIN, values=_vals26,
                name='2026',
                domain=dict(x=[0.54, 1.0]),
                hole=0.55,
                marker_colors=_CAT_COLORS,
                textinfo='none',
                hovertemplate='%{label}<br>2026: <b>%{value:,.0f}</b> (%{percent})<extra></extra>',
            ))
            _fig_mix.add_annotation(text='2025', x=0.23, y=0.5, font=dict(size=13, color=MUTED), showarrow=False)
            _fig_mix.add_annotation(text='2026', x=0.77, y=0.5, font=dict(size=13, color=GOLD), showarrow=False)
            _mix_layout = {**PLOTLY_BASE}
            _mix_layout.update(
                title=dict(text='Mix por categoría', **PLOTLY_BASE['title']),
                height=340,
                showlegend=True,
                legend=dict(**PLOTLY_BASE['legend'], orientation='v', x=1.0, y=0.5,
                            xanchor='left', yanchor='middle'),
                margin=dict(l=0, r=120, t=48, b=0),
            )
            _fig_mix.update_layout(**_mix_layout)
            st.plotly_chart(_fig_mix, use_container_width=True)

        # Monthly detail table
        with _col_tbl:
            st.markdown(f"""
            <div style="font-family:var(--font-serif);font-size:15px;
                        color:{TEXT};margin-bottom:12px;margin-top:8px">
              Detalle mensual
            </div>""", unsafe_allow_html=True)

            _rows_html = ""
            for _m in _common:
                _a25 = _v25[_m]; _a26 = _v26[_m]
                _pct = (_a26 - _a25) / _a25 * 100 if _a25 else 0
                _bar_color = '#5CE8D4' if _pct >= 0 else '#F7A8D0'
                _bar_w = min(abs(_pct) / 40 * 100, 100)
                _sign = '+' if _pct >= 0 else ''
                _rows_html += f"""
                <tr>
                  <td style="color:{MUTED};font-size:12px;padding:7px 8px;white-space:nowrap">
                    {_MESES_NOM_V[_m]}</td>
                  <td style="color:{SUBTLE};font-size:11px;text-align:right;padding:7px 8px">
                    {_fmt_clp(_a25)}</td>
                  <td style="color:{GOLD};font-size:11px;text-align:right;padding:7px 8px">
                    {_fmt_clp(_a26)}</td>
                  <td style="padding:7px 8px;min-width:120px">
                    <div style="display:flex;align-items:center;gap:6px">
                      <div style="flex:1;height:5px;border-radius:3px;
                                  background:rgba(255,255,255,0.07);overflow:hidden">
                        <div style="width:{_bar_w:.0f}%;height:100%;background:{_bar_color};
                                    border-radius:3px"></div>
                      </div>
                      <span style="font-size:11px;color:{_bar_color};
                                   min-width:44px;text-align:right">
                        {_sign}{_pct:.1f}%</span>
                    </div>
                  </td>
                </tr>"""

            # 2025-only months (no comparable)
            _only25 = sorted(set(_v25.index) - set(_common))
            for _m in _only25:
                _rows_html += f"""
                <tr>
                  <td style="color:{MUTED};font-size:12px;padding:7px 8px">{_MESES_NOM_V[_m]}</td>
                  <td style="color:{SUBTLE};font-size:11px;text-align:right;padding:7px 8px">
                    {_fmt_clp(_v25[_m])}</td>
                  <td style="color:{SUBTLE};font-size:11px;text-align:right;padding:7px 8px">—</td>
                  <td style="padding:7px 8px">
                    <span style="font-size:10px;color:{SUBTLE}">sin dato 2026</span></td>
                </tr>"""

            st.markdown(f"""
            <table style="width:100%;border-collapse:collapse;font-family:var(--font-sans)">
              <thead>
                <tr style="border-bottom:1px solid {BORDER}">
                  <th style="color:{SUBTLE};font-size:10px;letter-spacing:.08em;
                             font-weight:500;text-align:left;padding:4px 8px">MES</th>
                  <th style="color:{SUBTLE};font-size:10px;letter-spacing:.08em;
                             font-weight:500;text-align:right;padding:4px 8px">2025</th>
                  <th style="color:{SUBTLE};font-size:10px;letter-spacing:.08em;
                             font-weight:500;text-align:right;padding:4px 8px">2026</th>
                  <th style="color:{SUBTLE};font-size:10px;letter-spacing:.08em;
                             font-weight:500;padding:4px 8px">VARIACIÓN</th>
                </tr>
              </thead>
              <tbody>{_rows_html}</tbody>
            </table>""", unsafe_allow_html=True)

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        # ── Tercera fila: Variación por categoría ─────────────────────────────
        _CATS_ANA    = ['Alimentos', 'Bebidas S/Alcohol', 'Bebidas C/Alcohol', 'Vinos']
        _CATS_LABEL  = ['Alimentos\n(platos)', 'Bebidas\nS/Alcohol', 'Bebidas\nC/Alcohol', 'Vinos']
        _CAT_CLR     = [GOLD, '#7EB8F7', '#5CE8D4', '#C4B5FD']

        _cat_rows = []
        for _cat, _clr in zip(_CATS_ANA, _CAT_CLR):
            _r25 = _csum(_vc, _common, 2025, _cat, 'bruto')
            _r26 = _csum(_vc, _common, 2026, _cat, 'bruto')
            _q25 = _csum(_vq, _common, 2025, _cat, 'cantidad')
            _q26 = _csum(_vq, _common, 2026, _cat, 'cantidad')
            _rv  = (_r26-_r25)/_r25*100 if _r25 else 0
            _qv  = (_q26-_q25)/_q25*100 if _q25 else 0
            _a25 = _r25/_q25 if _q25 else 0
            _a26 = _r26/_q26 if _q26 else 0
            _av  = (_a26-_a25)/_a25*100 if _a25 else 0
            _cat_rows.append(dict(cat=_cat, color=_clr,
                r25=_r25, r26=_r26, rv=_rv,
                q25=_q25, q26=_q26, qv=_qv,
                a25=_a25, a26=_a26, av=_av))

        # Chart: ingresos% y unidades% side by side per category
        _fig_cat = go.Figure()
        _fig_cat.add_trace(go.Bar(
            name='Ingresos YoY %',
            x=_CATS_LABEL,
            y=[d['rv'] for d in _cat_rows],
            marker_color=[GOLD if d['rv'] >= 0 else '#F7A8D0' for d in _cat_rows],
            marker_line_width=0, opacity=0.9,
            text=[f"{d['rv']:+.1f}%" for d in _cat_rows],
            textposition='outside', textfont=dict(size=11, color='rgba(255,255,255,0.8)'),
            hovertemplate='%{x}<br>Ingresos: <b>%{y:+.1f}%</b><extra></extra>',
        ))
        _fig_cat.add_trace(go.Bar(
            name='Unidades vendidas YoY %',
            x=_CATS_LABEL,
            y=[d['qv'] for d in _cat_rows],
            marker_color=['#5CE8D4' if d['qv'] >= 0 else '#9F7AEA' for d in _cat_rows],
            marker_line_width=0, opacity=0.85,
            text=[f"{d['qv']:+.1f}%" for d in _cat_rows],
            textposition='outside', textfont=dict(size=11, color='rgba(255,255,255,0.7)'),
            hovertemplate='%{x}<br>Unidades: <b>%{y:+.1f}%</b><extra></extra>',
        ))
        _fig_cat.add_hline(y=0, line_color='rgba(255,255,255,0.12)', line_width=1)
        _cl = {**PLOTLY_BASE}
        _cl.update(
            title=dict(text='Variación por categoría — Ingresos vs Unidades vendidas (YoY %)',
                       **PLOTLY_BASE['title']),
            barmode='group', bargap=0.28, bargroupgap=0.06,
            yaxis=dict(**PLOTLY_BASE['yaxis'], tickformat='+.0f', ticksuffix='%',
                       title=dict(text='Variación %', font=dict(size=11, color=MUTED))),
            legend=dict(**PLOTLY_BASE['legend'], orientation='h', x=0, y=1.12),
            height=380, margin=dict(l=16, r=16, t=72, b=48),
        )
        _fig_cat.update_layout(**_cl)
        st.plotly_chart(_fig_cat, use_container_width=True)

        # ── Tabla detalle por categoría ───────────────────────────────────────
        _tbl_rows = ""
        for d in _cat_rows:
            _tbl_rows += (
                f'<tr style="border-bottom:1px solid rgba(255,255,255,0.04)">'
                f'<td style="padding:8px 10px;font-size:12px;color:{TEXT};white-space:nowrap">'
                f'<span style="display:inline-block;width:8px;height:8px;border-radius:50%;'
                f'background:{d["color"]};margin-right:7px;vertical-align:middle"></span>{d["cat"]}</td>'
                f'<td style="padding:8px 10px;font-size:11px;color:{MUTED};text-align:right">{_fmt_clp(d["r25"])}</td>'
                f'<td style="padding:8px 10px;font-size:11px;color:{GOLD};text-align:right">{_fmt_clp(d["r26"])}</td>'
                f'<td style="padding:8px 10px;text-align:center">{_pct_badge(d["rv"])}</td>'
                f'<td style="padding:8px 10px;font-size:11px;color:{MUTED};text-align:right">{_fmt_q(d["q25"])}</td>'
                f'<td style="padding:8px 10px;font-size:11px;color:{GOLD};text-align:right">{_fmt_q(d["q26"])}</td>'
                f'<td style="padding:8px 10px;text-align:center">{_pct_badge(d["qv"])}</td>'
                f'<td style="padding:8px 10px;font-size:11px;color:{MUTED};text-align:right">{_fmt_clp(d["a25"])}</td>'
                f'<td style="padding:8px 10px;font-size:11px;color:{GOLD};text-align:right">{_fmt_clp(d["a26"])}</td>'
                f'<td style="padding:8px 10px;text-align:center">{_pct_badge(d["av"])}</td>'
                f'</tr>'
            )

        _th = (f'color:{SUBTLE};font-size:10px;letter-spacing:.07em;font-weight:500;'
               f'padding:6px 10px;white-space:nowrap')
        st.markdown(
            f'<table style="width:100%;border-collapse:collapse;font-family:sans-serif">'
            f'<thead><tr style="border-bottom:1px solid rgba(255,255,255,0.1)">'
            f'<th style="{_th};text-align:left">CATEGORÍA</th>'
            f'<th style="{_th};text-align:right">ING. 2025</th>'
            f'<th style="{_th};text-align:right">ING. 2026</th>'
            f'<th style="{_th};text-align:center">VAR. ING.</th>'
            f'<th style="{_th};text-align:right">UNID. 2025</th>'
            f'<th style="{_th};text-align:right">UNID. 2026</th>'
            f'<th style="{_th};text-align:center">VAR. UNID.</th>'
            f'<th style="{_th};text-align:right">P.PROM 2025</th>'
            f'<th style="{_th};text-align:right">P.PROM 2026</th>'
            f'<th style="{_th};text-align:center">VAR. PRECIO</th>'
            f'</tr></thead>'
            f'<tbody>{_tbl_rows}</tbody>'
            f'</table>',
            unsafe_allow_html=True)

        st.markdown("<div style='height:20px'></div>", unsafe_allow_html=True)

        # ── Categorías mes a mes ───────────────────────────────────────────────
        _modo_cat = st.radio(
            "Métrica por categoría",
            ["Ingresos", "Unidades vendidas"],
            horizontal=True, label_visibility="collapsed",
            key="radio_cat_mes")
        _campo_cat = 'bruto' if _modo_cat == "Ingresos" else 'cantidad'
        _src_cat   = _vc if _campo_cat == 'bruto' else _vq
        _sufijo    = 'M CLP' if _campo_cat == 'bruto' else 'unidades'
        _div       = 1e6    if _campo_cat == 'bruto' else 1.0

        from plotly.subplots import make_subplots as _msp
        _fig_mm = _msp(
            rows=2, cols=2,
            subplot_titles=[f'<b>{c}</b>' for c in _CATS_ANA],
            vertical_spacing=0.14, horizontal_spacing=0.08,
            specs=[[{'secondary_y': True}, {'secondary_y': True}],
                   [{'secondary_y': True}, {'secondary_y': True}]],
        )

        _todos_meses_25 = sorted(_v25.index)
        _todos_meses_26 = sorted(_v26.index)

        for _ci, (_cat, _clr) in enumerate(zip(_CATS_ANA, _CAT_CLR)):
            _row = _ci // 2 + 1
            _col = _ci %  2 + 1

            _d25 = _mv_cat(_src_cat, _campo_cat, _cat, 2025, _todos_meses_25)
            _d26 = _mv_cat(_src_cat, _campo_cat, _cat, 2026, _todos_meses_26)
            _xs25 = [_MESES_NOM_V[m] for m in _todos_meses_25]
            _xs26 = [_MESES_NOM_V[m] for m in _todos_meses_26]
            _ys25 = [_d25.get(m, 0) / _div for m in _todos_meses_25]
            _ys26 = [_d26.get(m, 0) / _div for m in _todos_meses_26]

            _show_leg = (_ci == 0)

            _fig_mm.add_trace(go.Bar(
                name='2025', x=_xs25, y=_ys25,
                marker_color='rgba(201,169,122,0.30)',
                marker_line_color='rgba(201,169,122,0.45)', marker_line_width=1,
                showlegend=_show_leg,
                hovertemplate=f'%{{x}} 2025 — {_cat}<br><b>%{{y:.2f}} {_sufijo}</b><extra></extra>',
            ), row=_row, col=_col, secondary_y=False)

            _fig_mm.add_trace(go.Bar(
                name='2026', x=_xs26, y=_ys26,
                marker_color=_clr, marker_line_width=0, opacity=0.9,
                showlegend=_show_leg,
                hovertemplate=f'%{{x}} 2026 — {_cat}<br><b>%{{y:.2f}} {_sufijo}</b><extra></extra>',
            ), row=_row, col=_col, secondary_y=False)

            # Línea YoY % para los meses comparables
            _yoy_xs, _yoy_ys, _yoy_clrs, _yoy_txts = [], [], [], []
            for _m in _common:
                _a25v = _d25.get(_m, 0); _a26v = _d26.get(_m, 0)
                if _a25v:
                    _pv = (_a26v - _a25v) / _a25v * 100
                    _yoy_xs.append(_MESES_NOM_V[_m])
                    _yoy_ys.append(_pv)
                    _yoy_clrs.append('#5CE8D4' if _pv >= 0 else '#F7A8D0')
                    _yoy_txts.append(f'{_pv:+.0f}%')

            _fig_mm.add_trace(go.Scatter(
                name='Var. % YoY', x=_yoy_xs, y=_yoy_ys,
                mode='lines+markers+text',
                line=dict(color='rgba(255,255,255,0.45)', width=1.2, dash='dot'),
                marker=dict(size=7, color=_yoy_clrs, line=dict(color=BG, width=1)),
                text=_yoy_txts, textposition='top center',
                textfont=dict(size=9, color='rgba(255,255,255,0.65)'),
                showlegend=_show_leg,
                hovertemplate='%{x}<br>YoY: <b>%{y:+.1f}%</b><extra></extra>',
            ), row=_row, col=_col, secondary_y=True)

        # Layout global del subplot
        _fig_mm.update_layout(
            **{k: v for k, v in PLOTLY_BASE.items()
               if k not in ('xaxis', 'yaxis', 'title', 'margin', 'legend')},
            barmode='group', bargap=0.22, bargroupgap=0.06,
            height=620,
            margin=dict(l=16, r=48, t=56, b=16),
            legend=dict(**PLOTLY_BASE['legend'], orientation='h', x=0, y=1.04),
            title=dict(
                text=f'Categorías mes a mes — {_modo_cat} 2025 vs 2026',
                font=dict(family="'Palatino Linotype','Palatino',Georgia,serif",
                          size=18, color='rgba(255,255,255,0.95)'),
                x=0, xref='paper', pad=dict(b=8)),
        )
        # Ejes secundarios (YoY %) — solo estilo, make_subplots ya fijó overlaying
        for _ax in ['yaxis2', 'yaxis4', 'yaxis6', 'yaxis8']:
            _fig_mm.update_layout(**{
                _ax: dict(tickformat='+.0f', ticksuffix='%',
                          tickfont=dict(size=9, color='rgba(255,255,255,0.35)'),
                          gridcolor='rgba(0,0,0,0)',
                          zeroline=True, zerolinecolor='rgba(255,255,255,0.07)',
                          showgrid=False)
            })
        # Ejes x
        for _ax in ['xaxis', 'xaxis2', 'xaxis3', 'xaxis4']:
            _fig_mm.update_layout(**{
                _ax: dict(gridcolor='rgba(255,255,255,0.04)',
                          tickfont=dict(size=10))
            })
        _fig_mm.update_annotations(font=dict(
            family="'Palatino Linotype',Georgia,serif",
            size=13, color='rgba(255,255,255,0.85)'))
        st.plotly_chart(_fig_mm, use_container_width=True)

        st.markdown("<div style='height:8px'></div>", unsafe_allow_html=True)

        # ── Cuarta fila: Acumulado + Diagnóstico ─────────────────────────────
        _col_acum, _col_diag = st.columns([3, 2])

        with _col_acum:
            _fig_acum = go.Figure()

            _cx25, _cy25, _acc = [], [], 0
            for _m in sorted(_v25.index):
                _acc += _v25[_m]
                _cx25.append(_MESES_NOM_V[_m]); _cy25.append(_acc / 1e6)
            _cx26, _cy26, _acc = [], [], 0
            for _m in sorted(_v26.index):
                _acc += _v26[_m]
                _cx26.append(_MESES_NOM_V[_m]); _cy26.append(_acc / 1e6)

            _fig_acum.add_trace(go.Scatter(
                x=_cx25, y=_cy25, name='Acum. 2025',
                mode='lines+markers',
                line=dict(color='rgba(201,169,122,0.40)', width=2, dash='dot'),
                marker=dict(size=6, color='rgba(201,169,122,0.40)'),
                fill='tozeroy', fillcolor='rgba(201,169,122,0.04)',
                hovertemplate='%{x} 2025 (acum.)<br><b>$%{y:.1f}M</b><extra></extra>',
            ))
            _fig_acum.add_trace(go.Scatter(
                x=_cx26, y=_cy26, name='Acum. 2026',
                mode='lines+markers',
                line=dict(color=GOLD, width=2.5),
                marker=dict(size=7, color=GOLD, line=dict(color=BG, width=1.5)),
                fill='tozeroy', fillcolor='rgba(201,169,122,0.10)',
                hovertemplate='%{x} 2026 (acum.)<br><b>$%{y:.1f}M</b><extra></extra>',
            ))
            _al = {**PLOTLY_BASE}
            _al.update(
                title=dict(text='Ventas acumuladas en el año', **PLOTLY_BASE['title']),
                height=320,
                yaxis=dict(**PLOTLY_BASE['yaxis'],
                           title=dict(text='Millones CLP', font=dict(size=11, color=MUTED)),
                           tickformat=',.0f', ticksuffix='M'),
                legend=dict(**PLOTLY_BASE['legend'], orientation='h', x=0, y=1.15),
                margin=dict(l=16, r=16, t=64, b=16),
            )
            _fig_acum.update_layout(**_al)
            st.plotly_chart(_fig_acum, use_container_width=True)

        with _col_diag:
            # ── Diagnóstico (individual st.markdown por item) ─────────────────
            _ult3 = sorted(_common)[-3:]
            _tend3 = float(np.mean([_yoy[m] for m in _ult3]))
            _tend_up = _tend3 >= 0

            _desc25 = float(_vd[(_vd['año']==2025) & (_vd['mes'].isin(_common))]['descuento'].sum())
            _desc26 = float(_vd[(_vd['año']==2026) & (_vd['mes'].isin(_common))]['descuento'].sum())
            _dpct25 = _desc25/_tot25*100 if _tot25 else 0
            _dpct26 = _desc26/_tot26*100 if _tot26 else 0

            _proy_anual = None
            if len(_v26) >= 3:
                _mp = float(np.mean([_v26[m] for m in sorted(_v26.index)]))
                _mr = 12 - int(max(_v26.index))
                _proy_anual = float(_tot26) + _mp * _mr

            # worst category by revenue
            _worst_cat = min(_cat_rows, key=lambda d: d['rv'])
            _best_cat  = max(_cat_rows, key=lambda d: d['rv'])

            st.markdown(
                f'<div style="font-size:15px;font-family:Georgia,serif;'
                f'color:rgba(255,255,255,0.95);margin-bottom:14px;margin-top:4px">'
                f'Diagnóstico</div>',
                unsafe_allow_html=True)


            _tc = '#5CE8D4' if _tend_up else '#F7A8D0'
            _diag_card(
                'TENDENCIA RECIENTE',
                f'Últimos {len(_ult3)} meses: {"crecimiento" if _tend_up else "caída"} promedio de {abs(_tend3):.1f}%',
                _tc)

            _gc = '#5CE8D4' if _n_pos >= len(_common)//2 else '#F7A8D0'
            _diag_card(
                'MESES EN VERDE',
                f'{_n_pos} de {len(_common)} meses superan a 2025',
                _gc)

            _diag_card(
                'CATEGORIA MAS AFECTADA',
                f'{_worst_cat["cat"]}: ingresos {_worst_cat["rv"]:+.1f}%, unidades {_worst_cat["qv"]:+.1f}%',
                '#F7A8D0')

            _diag_card(
                'CATEGORIA CON MEJOR DESEMPENO',
                f'{_best_cat["cat"]}: ingresos {_best_cat["rv"]:+.1f}%, unidades {_best_cat["qv"]:+.1f}%',
                '#5CE8D4')

            _diag_card(
                'DESCUENTOS SOBRE VENTAS',
                f'{_dpct25:.1f}% (2025) vs {_dpct26:.1f}% (2026)',
                GOLD)

            if _proy_anual:
                _diag_card(
                    'PROYECCION ANUAL 2026',
                    f'~{_fmt_clp(_proy_anual)} si se mantiene el ritmo actual',
                    GOLD_2)

        # ── Impacto AMEX Domingos ─────────────────────────────────────────────
        st.markdown("<div style='height:12px'></div>", unsafe_allow_html=True)
        st.markdown(
            f'<div style="font-size:10px;letter-spacing:.09em;color:rgba(255,255,255,.32);'
            f'font-family:sans-serif;margin-bottom:10px">IMPACTO PROMO AMEX DOMINGOS</div>',
            unsafe_allow_html=True)

        _amex_meses = sorted(_vdf[(_vdf['año']==2026) & (_vdf['mes']>=6)]['mes'].unique())
        if not _amex_meses:
            st.markdown(
                f'<div style="background:rgba(201,169,122,0.06);border:1px solid rgba(201,169,122,.18);'
                f'border-radius:8px;padding:16px 18px;display:flex;align-items:center;gap:14px">'
                f'<div style="font-size:22px">⭐</div>'
                f'<div>'
                f'<div style="font-size:13px;color:rgba(255,255,255,.85)">Promo activa desde el 7 Jun 2026</div>'
                f'<div style="font-size:11px;color:rgba(255,255,255,.4);margin-top:4px">'
                f'Los datos de impacto aparecerán aquí automáticamente cuando se agregue <b style="color:rgba(201,169,122,.7)">Jun26.xls</b> a la carpeta Ventas/</div>'
                f'</div></div>',
                unsafe_allow_html=True)
        else:
            # Hay datos de Jun 2026 en adelante — comparar Dom vs resto
            _amex_data = _vdf[(_vdf['año']==2026) & (_vdf['mes'].isin(_amex_meses))]
            # Para comparar necesitaríamos día de semana por transacción, que no tenemos en los XLS mensuales
            # Mostramos lo que hay: ingreso mensual Jun+ 2026 vs 2025
            _amex_comp = []
            for _am in _amex_meses:
                _a26v = _amex_data[_amex_data['mes']==_am]['bruto'].sum()
                _a25v = _vdf[(_vdf['año']==2025) & (_vdf['mes']==_am)]['bruto'].sum()
                _amex_comp.append((_am, _a25v, _a26v))
            _NOM = ['','Ene','Feb','Mar','Abr','May','Jun','Jul','Ago','Sep','Oct','Nov','Dic']
            _amex_html = ''
            for _am, _a25v, _a26v in _amex_comp:
                _pv = (_a26v-_a25v)/_a25v*100 if _a25v else 0
                _pc = '#5CE8D4' if _pv >= 0 else '#F7A8D0'
                _amex_html += (
                    f'<div style="display:flex;align-items:center;gap:10px;padding:8px 0;'
                    f'border-bottom:1px solid rgba(255,255,255,.04)">'
                    f'<div style="min-width:32px;font-size:12px;color:rgba(255,255,255,.6)">{_NOM[_am]}</div>'
                    f'<div style="font-size:11px;color:rgba(255,255,255,.4);min-width:70px">2025: {_fmt_clp(_a25v)}</div>'
                    f'<div style="font-size:11px;color:{GOLD};min-width:70px">2026: {_fmt_clp(_a26v)}</div>'
                    f'<div style="font-size:12px;font-weight:600;color:{_pc}">{"+" if _pv>=0 else ""}{_pv:.1f}%</div>'
                    f'</div>')
            st.markdown(
                f'<div style="background:rgba(201,169,122,0.06);border:1px solid rgba(201,169,122,.18);'
                f'border-radius:8px;padding:14px 18px">'
                f'<div style="font-size:12px;color:rgba(201,169,122,.8);margin-bottom:10px">'
                f'⭐ Meses con promo activa (Jun 2026 en adelante)</div>'
                f'{_amex_html}</div>',
                unsafe_allow_html=True)

# ── Footer ─────────────────────────────────────────────────────────────────────

st.markdown(f"""
<div style="margin-top:48px;padding:20px 0;border-top:1px solid {BORDER};
            display:flex;justify-content:space-between;align-items:center;
            font-size:11px;color:{SUBTLE};letter-spacing:.06em">
  <span><span style="color:{GOLD};font-weight:600">Margo · Nelí</span> · Sistema de Pronóstico de Producción</span>
  <span>{_fecha_es(datetime.now())} · {len(history)} días en modelo · μ + {k_factor}σ</span>
</div>
""", unsafe_allow_html=True)
