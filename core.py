#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
core.py — Lógica de negocio pura · Margo / Nelí
Sin dependencias de Streamlit.  Importado por app_qt.py (PyQt6) y app.py (Streamlit).
"""

import os, io, json, html as _html, warnings
from datetime import datetime, timedelta
from collections import defaultdict
from functools import lru_cache

warnings.filterwarnings("ignore")

import numpy as np
import pandas as pd
import plotly.graph_objects as go
import plotly.io as _pio

# Deshabilitar el template default de Plotly — inyecta colores blancos/azules
# que Plotly JS 3.x puede usar como fallback sobreescribiendo nuestros colores warm.
_pio.templates.default = 'none'

try:
    import xlrd
    HAS_XLRD = True
except ImportError:
    HAS_XLRD = False

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

# ── Paths ──────────────────────────────────────────────────────────────────────
# Cuando corre como .exe (PyInstaller), los datos de usuario viven junto al .exe.
# sys._MEIPASS apunta a los recursos empaquetados (assets/), no a los datos.

import sys as _sys

if getattr(_sys, 'frozen', False):
    BASE_DIR = os.path.dirname(_sys.executable)   # carpeta donde está el .exe
else:
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))

CONFIG_PATH    = os.path.join(BASE_DIR, "config.json")
DEFAULT_FOLDER = os.path.join(BASE_DIR, "Reportes")

# ── Design constants (usados por chart builders y HTML) ────────────────────────

GOLD     = "#C9A97A"
GOLD_2   = "#E0C090"
GOLD_DIM = "rgba(201,169,122,0.12)"
BG       = "#090806"
BG_1     = "#0F0C09"
BG_2     = "#16120D"
BG_3     = "#1E1912"
BG_4     = "#272018"
BORDER   = "rgba(250,235,200,0.07)"
TEXT     = "rgba(242,234,224,0.95)"
MUTED    = "rgba(242,234,224,0.64)"
SUBTLE   = "rgba(242,234,224,0.32)"

CAT_HEX = {
    "Fondo":    "#C9A97A",
    "Entrada":  "#7EB8F7",
    "Ensalada": "#5CE8D4",
    "Postre":   "#F7A8D0",
    "Niños":    "#C4B5FD",
    "Barra":    "#A78BFA",
}

# Colores de tipografía específicos para charts Plotly
# TEXT/MUTED son near-white y se ven blancos en fondo oscuro.
# Usamos la paleta gold para que el texto sea visiblemente cálido.
_CH_TITLE  = "#E0C090"                  # títulos principales — gold claro
_CH_LEGEND = "rgba(201,169,122,0.90)"   # leyenda — gold
_CH_TICK   = "rgba(201,169,122,0.70)"   # tick labels ejes — gold medio
_CH_AXIS   = "rgba(201,169,122,0.55)"   # títulos de ejes — gold suave
_CH_FONT   = "rgba(201,169,122,0.85)"   # fuente global chart — gold base
_SANS      = "'Segoe UI','Segoe UI Light',Calibri,system-ui,sans-serif"
_SERIF     = "'Palatino Linotype','Book Antiqua',Palatino,Georgia,serif"

PLOTLY_BASE = dict(
    paper_bgcolor="rgba(0,0,0,0)",
    plot_bgcolor="rgba(16,12,9,0.45)",
    font=dict(family=_SANS, color=_CH_FONT, size=12),
    margin=dict(l=16, r=16, t=48, b=16),
    xaxis=dict(gridcolor="rgba(250,235,200,0.04)", linecolor="rgba(250,235,200,0.07)",
               zerolinecolor="rgba(250,235,200,0.04)", tickfont=dict(size=11, color=_CH_TICK)),
    yaxis=dict(gridcolor="rgba(250,235,200,0.04)", linecolor="rgba(250,235,200,0.07)",
               zerolinecolor="rgba(250,235,200,0.04)", tickfont=dict(size=11, color=_CH_TICK)),
    hoverlabel=dict(bgcolor=BG_3, bordercolor="rgba(201,169,122,0.45)",
                    font=dict(family=_SANS, color=_CH_FONT, size=12)),
    title=dict(font=dict(family=_SERIF, size=20, color=_CH_TITLE),
               x=0, xref="paper", pad=dict(b=12)),
    legend=dict(bgcolor="rgba(0,0,0,0)", bordercolor=BORDER,
                borderwidth=1, font=dict(size=11, color=_CH_LEGEND)),
    colorway=[GOLD, "#7EB8F7", "#5CE8D4", "#F7A8D0", "#C4B5FD", "#E0C090", "#A5D6A7"],
)

def _theme(fig: "go.Figure") -> "go.Figure":
    """Fuerza paleta gold en TODOS los elementos de texto de Plotly.
    Plotly JS 3.x no hereda layout.font.color a axis-title/tick/legend."""
    # Solo impone COLOR — no sobreescribe tamaños específicos por chart (ej: fig_mm size=9)
    fig.update_xaxes(
        title_font_color=_CH_AXIS,
        tickfont_color=_CH_TICK,
    )
    fig.update_yaxes(
        title_font_color=_CH_AXIS,
        tickfont_color=_CH_TICK,
    )
    fig.update_layout(
        font_color=_CH_FONT,
        legend_font_color=_CH_LEGEND,
    )
    return fig

# ── Utility helpers ────────────────────────────────────────────────────────────

def _fmt_clp(v):
    if v is None or v != v: return "$0K"   # None o NaN
    if v >= 1e9: return f"${v/1e9:.2f}MM"
    if v >= 1e6: return f"${v/1e6:.1f}M"
    return f"${v/1e3:.0f}K"

def _delta_arrow(v):
    return ("▲" if v >= 0 else "▼"), ("#5CE8D4" if v >= 0 else "#F7A8D0")

def _pct_badge(v):
    c = '#5CE8D4' if v >= 0 else '#F7A8D0'
    s = '+' if v >= 0 else ''
    return f'<span style="color:{c};font-weight:600">{s}{v:.1f}%</span>'

def _fmt_q(v):
    if v >= 1000: return f"{v/1000:.1f}k"
    return f"{v:.0f}"

def cov_color(pct, is_current=False):
    border = f'border:2px solid {GOLD};' if is_current else 'border:1px solid rgba(242,234,224,.06);'
    if pct >= 90:   bg, color = 'rgba(100,230,170,.12)', 'rgba(100,230,170,.9)'
    elif pct >= 75: bg, color = 'rgba(245,158,11,.10)',  '#F59E0B'
    else:           bg, color = 'rgba(215,75,65,.10)',   'rgba(215,75,65,.9)'
    return (f'background:{bg};color:{color};{border}'
            f'border-radius:6px;padding:5px 8px;text-align:center;'
            f'font-family:monospace;font-size:13px;font-weight:600')

def _csum(df, common, anio, cat, campo):
    m = df[(df['año']==anio) & (df['mes'].isin(common)) & (df['categoria']==cat)]
    return float(m[campo].sum())

def _cat_totals(vc, cats, anio, meses):
    sub = vc[(vc['año'] == anio) & (vc['mes'].isin(meses))]
    agg = sub.groupby('categoria')['bruto'].sum()
    return [agg.get(c, 0) for c in cats]

def _mv_cat(src_cat, campo_cat, cat, anio, meses):
    sub = src_cat[(src_cat['año']==anio) & (src_cat['mes'].isin(meses))
                  & (src_cat['categoria']==cat)]
    return {int(r['mes']): float(r[campo_cat]) for _, r in sub.iterrows()}

# ── Config ─────────────────────────────────────────────────────────────────────

def load_config() -> dict:
    try:
        with open(CONFIG_PATH, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return {}

def save_config(cfg: dict) -> None:
    _atomic_write(CONFIG_PATH, cfg)


def _atomic_write(path: str, data: dict) -> None:
    tmp = path + '.tmp'
    with open(tmp, 'w', encoding='utf-8') as f:
        json.dump(data, f, indent=2, ensure_ascii=False)
    os.replace(tmp, path)

def get_folder() -> str:
    folder = load_config().get("reportes_folder", DEFAULT_FOLDER)
    return folder if os.path.isdir(folder) else DEFAULT_FOLDER

def count_reports(folder: str) -> int:
    if not os.path.isdir(folder):
        return 0
    return sum(1 for f in os.listdir(folder)
               if f.startswith("informe") and f.endswith((".xls", ".xlsx")))

# ── Business constants ─────────────────────────────────────────────────────────

PROMO_FROM      = datetime(2026, 6, 7)
K_SAFETY        = 0.7
BUFFER_FALLBACK = 0.20
DEFAULT_HORIZON = 3

EXCLUDED = {
    datetime(2026, 5, 18),
    datetime(2026, 5, 21), datetime(2026, 5, 22),
    datetime(2026, 5, 23), datetime(2026, 5, 24),
}

FERIADOS = {
    datetime(2026,  1,  1): "Año Nuevo",
    datetime(2026,  4,  3): "Viernes Santo",
    datetime(2026,  4,  4): "Sábado Santo",
    datetime(2026,  5,  1): "Día del Trabajo",
    datetime(2026,  5, 21): "Glorias Navales",
    datetime(2026,  6, 29): "San Pedro y San Pablo",
    datetime(2026,  7, 16): "Virgen del Carmen",
    datetime(2026,  8, 15): "Asunción de la Virgen",
    datetime(2026,  9, 18): "Fiestas Patrias",
    datetime(2026,  9, 19): "Glorias del Ejército",
    datetime(2026, 10, 12): "Encuentro de Dos Mundos",
    datetime(2026, 10, 31): "Iglesias Evangélicas",
    datetime(2026, 11,  1): "Todos los Santos",
    datetime(2026, 12,  8): "Inmaculada Concepción",
    datetime(2026, 12, 25): "Navidad",
    datetime(2027,  1,  1): "Año Nuevo",
    datetime(2027,  4, 26): "Viernes Santo",
    datetime(2027,  5,  1): "Día del Trabajo",
    datetime(2027,  5, 21): "Glorias Navales",
    datetime(2027,  6, 28): "San Pedro y San Pablo",
    datetime(2027,  7, 16): "Virgen del Carmen",
    datetime(2027,  8, 15): "Asunción de la Virgen",
    datetime(2027,  9, 18): "Fiestas Patrias",
    datetime(2027,  9, 19): "Glorias del Ejército",
    datetime(2027, 10, 12): "Encuentro de Dos Mundos",
    datetime(2027, 10, 31): "Iglesias Evangélicas",
    datetime(2027, 11,  1): "Todos los Santos",
    datetime(2027, 12,  8): "Inmaculada Concepción",
    datetime(2027, 12, 25): "Navidad",
}

PREFIXES = {
    'VECARFON': 'Fondo',   'VECARFJT': 'Fondo',   'VECARFUB': 'Fondo',
    'VECARENT': 'Entrada',  'VECARAJT': 'Entrada',  'VECARAUB': 'Entrada',
    'VECARENS': 'Ensalada', 'VECAREJT': 'Ensalada', 'VECAREUB': 'Ensalada',
    'VECARPOS': 'Postre',   'VECARPUB': 'Postre',
    'VECARNIN': 'Niños',
    'VELICACO': 'Barra',   'VELICASA': 'Barra',
    'VELICATR': 'Barra',   'VELICDES': 'Barra',
    'VCVINCOP': 'Barra',   'VCVINCAR': 'Barra',
    'VCVINESP': 'Barra',   'VCVINSAU': 'Barra',
}

_EXCLUDE_SUFFIXES = (' PY', ' JT')
CAT_ORDER  = ['Fondo', 'Entrada', 'Ensalada', 'Postre', 'Niños', 'Barra']
DAYS_ES    = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado','Domingo']
MONTHS_ES  = ['','enero','febrero','marzo','abril','mayo','junio',
              'julio','agosto','septiembre','octubre','noviembre','diciembre']

# ── Core model functions ───────────────────────────────────────────────────────

def fecha_es(d):
    return f"{d.day} de {MONTHS_ES[d.month]} de {d.year}"

def confidence_label(n):
    if n == 0: return "Sin datos",       "#6B7280"
    if n == 1: return "⚠ 1 obs · baja", "#F59E0B"
    if n < 4:  return f"△ {n} obs · mod", "#7EB8F7"
    return             f"✓ {n} obs",     "#64E6AA"

def day_type(date):
    if date.weekday() == 6 and date >= PROMO_FROM:
        return 'Dom-Promo'
    return DAYS_ES[date.weekday()]

def parse_report(path, errors=None):
    result = {}
    def handle_row(code, name, raw):
        code = str(code or '').strip()
        name = str(name or '').strip()
        if not code or not name:
            return
        if any(name.upper().endswith(s.upper()) for s in _EXCLUDE_SUFFIXES):
            return
        for prefix, cat in PREFIXES.items():
            if code.startswith(prefix):
                try:
                    qty = float(raw or 0)
                    if qty > 0:
                        result[(code, name, cat)] = qty
                except (ValueError, TypeError):
                    pass
                break
    try:
        if path.endswith('.xlsx'):
            if not HAS_OPENPYXL:
                return result
            wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
            for row in wb.active.iter_rows(min_row=22, values_only=True):
                if len(row) > 9:
                    handle_row(row[0], row[1], row[9])
            wb.close()
        else:
            if not HAS_XLRD:
                return result
            wb = xlrd.open_workbook(path)
            ws = wb.sheet_by_index(0)
            if ws.ncols > 9:
                for r in range(22, ws.nrows):
                    handle_row(ws.cell_value(r, 0), ws.cell_value(r, 1), ws.cell_value(r, 9))
    except Exception as e:
        if errors is not None:
            errors.append(f"Error leyendo {os.path.basename(path)}: {type(e).__name__}: {e}")
        return {}
    return result

def load_history(folder: str, errors: list | None = None,
                 excluded_extra: set | None = None) -> dict:
    history = {}
    if not os.path.isdir(folder):
        return history
    _excluded = EXCLUDED | (excluded_extra or set())
    for fname in sorted(os.listdir(folder)):
        if not fname.startswith('informe'):
            continue
        if not (fname.endswith('.xls') or fname.endswith('.xlsx')):
            continue
        raw = fname.replace('informe ', '').replace('.xlsx','').replace('.xls','').strip()
        try:
            d, m, y = raw.split('-')
            date = datetime(int(y), int(m), int(d))
        except ValueError:
            if errors is not None:
                errors.append(f"Nombre inválido: {fname}")
            continue
        if date in _excluded:
            continue
        data = parse_report(os.path.join(folder, fname), errors)
        if data:
            history[date] = data
        elif errors is not None:
            errors.append(f"Sin datos reconocibles: {fname}")
    return history

def build_model(history):
    accum = defaultdict(lambda: defaultdict(list))
    for date, data in history.items():
        dt = day_type(date)
        for (code, name, cat), qty in data.items():
            accum[dt][(name, cat)].append(qty)
    model = {}
    for dt, dishes in accum.items():
        model[dt] = {}
        for dish, vals in dishes.items():
            n    = len(vals)
            mean = sum(vals) / n
            std  = (sum((v - mean)**2 for v in vals) / (n - 1))**0.5 if n > 1 else None
            model[dt][dish] = {'mean': mean, 'std': std, 'n': n}
    return model

def build_name_to_sku(history):
    lookup = {}
    for data in history.values():
        for (code, name, cat) in data:
            if name not in lookup:
                lookup[name] = code
    return lookup

def dish_fc(stats, k=K_SAFETY):
    mean, std = stats['mean'], stats['std']
    return mean + k * std if std is not None else mean * (1 + BUFFER_FALLBACK)

def make_forecast(model, start, horizon, k=K_SAFETY):
    day_details, totals = [], defaultdict(float)
    mu_acc, sg_acc = defaultdict(float), defaultdict(float)
    for i in range(horizon):
        d  = start + timedelta(days=i)
        dt = day_type(d)
        if dt in model:
            dm = model[dt]
        elif dt == 'Dom-Promo':
            dm = model.get('Domingo', {})
        elif dt == 'Domingo':
            dm = model.get('Dom-Promo', {})
        else:
            dm = {}
        day_details.append((d, dt, dm, dt not in model))
        for dish, stats in dm.items():
            mean   = stats['mean']
            buffer = k * stats['std'] if stats['std'] is not None else mean * BUFFER_FALLBACK
            mu_acc[dish] += mean
            sg_acc[dish] += buffer
            totals[dish] += mean + buffer
    dish_stats = {d: {'mu': mu_acc[d], 'sigma': sg_acc[d]} for d in totals}
    return day_details, dict(totals), dish_stats

def build_df(history):
    rows = []
    for date, data in history.items():
        dt = day_type(date)
        for (code, name, cat), qty in data.items():
            rows.append({
                'date': date, 'day_type': dt,
                'weekday': DAYS_ES[date.weekday()],
                'is_promo': date.weekday() == 6 and date >= PROMO_FROM,
                'dish': name, 'category': cat, 'quantity': qty,
            })
    if not rows:
        return pd.DataFrame()
    df = pd.DataFrame(rows)
    df['date'] = pd.to_datetime(df['date'])
    return df

def compute_dish_coverage(history, day_fc, model, k):
    accum = {}
    for _, dt, _, _ in day_fc:
        dm        = model.get(dt, {})
        hist_days = [(d, data) for d, data in history.items() if day_type(d) == dt]
        if not hist_days:
            continue
        for (name, cat), stats in dm.items():
            mean = stats['mean']
            std  = stats['std']
            fc   = mean + k * std if std is not None else mean * (1 + BUFFER_FALLBACK)
            key  = (name, cat)
            if key not in accum:
                accum[key] = {'covered': 0, 'total': 0}
            for _, actual_data in hist_days:
                actual = next(
                    (qty for (_, n, c), qty in actual_data.items() if n == name and c == cat),
                    None,
                )
                if actual is None:
                    continue
                accum[key]['total']   += 1
                if actual <= fc:
                    accum[key]['covered'] += 1
    return {
        key: round(v['covered'] / v['total'] * 100)
        for key, v in accum.items() if v['total'] > 0
    }

def compute_buffer_coverage(history, model):
    K_VALUES = [0.0, 0.5, 0.7, 1.0, 1.5, 2.0]
    results  = {}
    for dt, dm in model.items():
        days = [(d, data) for d, data in history.items() if day_type(d) == dt]
        if len(days) < 2:
            continue
        base_fondos = {name: s for (name, cat), s in dm.items() if cat == 'Fondo'}
        if not base_fondos:
            continue
        covered = {k: 0 for k in K_VALUES}
        excesos = []
        for _, actual_data in days:
            actual_total = sum(qty for (_, name, cat), qty in actual_data.items() if cat == 'Fondo')
            base_f = sum(s['mean'] for s in base_fondos.values())
            if base_f > 0:
                excesos.append((actual_total - base_f) / base_f * 100)
            for k in K_VALUES:
                forecast = sum(
                    s['mean'] + k * (s['std'] if s['std'] is not None else s['mean'] * BUFFER_FALLBACK)
                    for s in base_fondos.values()
                )
                if actual_total <= forecast:
                    covered[k] += 1
        n = len(days)
        results[dt] = {
            'n': n,
            'coverage': {k: round(covered[k] / n * 100) for k in K_VALUES},
            'avg_excess_pct': round(
                sum(e for e in excesos if e > 0) / max(1, sum(1 for e in excesos if e > 0)), 1
            ) if excesos else 0,
        }
    return results

# ── HTML builders (usados por Qt y Streamlit) ──────────────────────────────────

def _cov_badge(pct):
    if pct is None: return ''
    if pct >= 90:   bg, fg, dot = 'rgba(100,230,170,.12)', 'rgba(100,230,170,.9)', '●'
    elif pct >= 75: bg, fg, dot = 'rgba(245,158,11,.12)',  '#F59E0B',              '●'
    else:           bg, fg, dot = 'rgba(215,75,65,.12)',   'rgba(215,75,65,.9)',   '●'
    return (f'<span style="display:inline-flex;align-items:center;gap:3px;'
            f'background:{bg};border:1px solid {fg}44;color:{fg};'
            f'font-size:9px;font-weight:700;padding:1px 6px;border-radius:10px;'
            f'margin-left:7px;vertical-align:middle">{dot} {pct}%</span>')

def prod_table_html(items, dish_stats, color, dish_coverage=None):
    if not items:
        return f'<p style="color:{_CH_AXIS};padding:20px">Sin datos para esta categoría.</p>'
    max_v = max(items.values())
    rows  = ''
    for i, ((name, cat), qty) in enumerate(sorted(items.items(), key=lambda x: -x[1])):
        st_d  = dish_stats.get((name, cat), {'mu': qty, 'sigma': 0})
        mu    = round(st_d['mu'])
        sigma = round(st_d['sigma'])
        pct   = round(qty / max_v * 100)
        rank  = {0: 'r1', 1: 'r2', 2: 'r3'}.get(i, 'dim' if i >= 10 else '')
        total_style = f'color:{color}' if i < 3 else ''
        badge = _cov_badge(dish_coverage.get((name, cat)) if dish_coverage is not None else None)
        rows += f"""<tr>
          <td class="td-name {rank}">{_html.escape(name)}{badge}</td>
          <td class="td-num td-mu">{mu}</td>
          <td class="td-num td-buf">+{sigma}</td>
          <td class="td-num td-total" style="{total_style}">{round(qty)}</td>
          <td class="td-bar"><div class="bar-track">
            <div class="bar-fill" style="width:{pct}%;background:{color}"></div>
          </div></td></tr>"""
    return f"""<div class="prod-wrap"><table class="prod-table">
      <thead><tr><th>Plato</th>
        <th class="r" title="Promedio histórico">Base</th>
        <th class="r" title="Buffer de seguridad estadístico">Buffer</th>
        <th class="r" style="color:{color}" title="Cantidad total a producir">Producir</th>
        <th class="bh"></th></tr></thead>
      <tbody>{rows}</tbody></table></div>"""

def day_card_html(d, dt, ff, pf, max_ff, is_fallback=False, idx=0):
    is_p    = dt == 'Dom-Promo'
    feriado = FERIADOS.get(datetime(d.year, d.month, d.day))
    nc      = GOLD if is_p else '#7EB8F7'
    delay   = f'{idx * 0.10:.2f}s'

    badges = []
    if is_p:        badges.append('<span class="promo-chip">◆ AMEX</span>')
    if feriado:     badges.append(f'<span class="dc-feriado-chip">{feriado}</span>')
    if is_fallback: badges.append('<span class="dc-warn-chip">Sin historial</span>')
    chips = f'<div class="dc-chips">{"".join(badges)}</div>' if badges else ''

    return f"""<div class="day-card{'  promo' if is_p else ''}" style="animation-delay:{delay}">
  <div class="dc-head">
    <span class="dc-dow">{DAYS_ES[d.weekday()]}</span>
    <span class="dc-date">{d.strftime('%d')}<small>{MONTHS_ES[d.month][:3]}</small></span>
  </div>
  {chips}
  <div class="dc-hero-num" style="color:{nc}">{round(ff):,}</div>
  <div class="dc-hero-sub">fondos</div>
  <div class="dc-sep"></div>
  <div class="dc-foot">
    <span class="dc-foot-num">{round(pf):,}</span>
    <span class="dc-foot-lbl">platos</span>
  </div>
</div>"""

# ── Chart builders (devuelven go.Figure) ──────────────────────────────────────

def chart_forecast_bars(totals, dish_stats, cat='Fondo', top_n=20, horizon=1):
    items = sorted(
        [(n, c, q, dish_stats.get((n, c), {'mu': q, 'sigma': 0}))
         for (n, c), q in totals.items() if c == cat and q >= 1],
        key=lambda x: -x[2])[:top_n]
    if not items:
        return go.Figure()
    names  = [x[0] for x in items]
    mu_v   = [round(x[3]['mu'])    for x in items]
    sig_v  = [round(x[3]['sigma']) for x in items]
    total  = [round(x[2])          for x in items]
    color  = CAT_HEX.get(cat, GOLD)
    rgb    = f'{int(color[1:3],16)},{int(color[3:5],16)},{int(color[5:7],16)}'
    dias_lbl = f"{horizon} día{'s' if horizon != 1 else ''}"
    fig    = go.Figure()
    fig.add_trace(go.Bar(name='Base acumulada', y=names, x=mu_v, orientation='h',
        marker=dict(color=f'rgba({rgb},.45)', cornerradius=4),
        hovertemplate=f'<b>%{{y}}</b><br>Base ({dias_lbl}): %{{x}} uds<extra></extra>'))
    fig.add_trace(go.Bar(name='Buffer de seguridad', y=names, x=sig_v, orientation='h',
        marker=dict(color='rgba(110,231,183,.65)', cornerradius=4),
        hovertemplate=f'<b>%{{y}}</b><br>Buffer ({dias_lbl}): +%{{x}} uds<extra></extra>'))
    for name, tot in zip(names, total):
        fig.add_annotation(x=tot, y=name, text=f'<b>{tot}</b>', showarrow=False,
            xanchor='left', xshift=8, font=dict(size=11, color=color))
    layout = {**PLOTLY_BASE}
    layout.update(barmode='stack', height=max(380, len(names)*32),
        margin=dict(l=220, r=64, t=48, b=16),
        title=dict(text=f'Forecast — {cat} · {dias_lbl}', **PLOTLY_BASE['title']),
        yaxis={**PLOTLY_BASE['yaxis'], 'autorange': 'reversed', 'automargin': True},
        xaxis={**PLOTLY_BASE['xaxis'], 'title': f'Unidades totales a producir ({dias_lbl})'})
    fig.update_layout(**layout)
    return _theme(fig)

def chart_heatmap(df, top_n=8):
    """¿En qué días se vende más cada plato? — Heatmap por tipo de día."""
    if df.empty: return go.Figure()
    fondo_df = df[df['category'] == 'Fondo']
    top_dishes = (fondo_df.groupby('dish')['quantity'].mean()
                  .sort_values(ascending=False).head(top_n).index.tolist())
    pivot = (fondo_df[fondo_df['dish'].isin(top_dishes)]
             .groupby(['day_type','dish'])['quantity'].mean().unstack(fill_value=0))
    order = [d for d in DAYS_ES + ['Dom-Promo'] if d in pivot.index]
    pivot = pivot.reindex(order).reindex(columns=top_dishes, fill_value=0)
    n_rows, n_cols = pivot.shape
    full_names  = list(pivot.columns)
    short_names = [n[:24] for n in full_names]
    customdata  = [[full_names[xi] for xi in range(n_cols)] for _ in range(n_rows)]
    z_vals = pivot.values.astype(float)

    fig = go.Figure(go.Heatmap(
        z=z_vals, x=short_names, y=list(pivot.index),
        customdata=customdata,
        colorscale=[[0,'#0B1428'],[0.35,'#3A2A00'],[0.70,'#8B6B1A'],[1,'#C9A97A']],
        hovertemplate=(
            '<b>%{customdata}</b><br>'
            '%{y}: <b>%{z:.0f}</b> uds promedio<br>'
            '<i>Más dorado = más demanda</i><extra></extra>'),
        showscale=True,
        text=np.where(z_vals > 0, np.round(z_vals, 0).astype(int), ''),
        texttemplate='<b>%{text}</b>',
        textfont=dict(size=11, color='rgba(242,234,224,0.90)'),
        colorbar=dict(
            title=dict(text='Uds<br>promedio', font=dict(color=_CH_TICK, size=10)),
            bgcolor='rgba(0,0,0,0)', bordercolor=BORDER,
            tickfont=dict(color=_CH_TICK, size=10), thickness=14, len=0.85)))
    layout = {**PLOTLY_BASE}
    layout.update(
        title=dict(
            text='¿En qué días se vende más cada fondo?'
                 '<br><sup style="font-size:12px;color:rgba(201,169,122,.55)">'
                 'Promedio histórico de unidades · más dorado = más demanda ese día</sup>',
            **PLOTLY_BASE['title']),
        height=420, margin=dict(l=16, r=80, t=80, b=110),
        xaxis={**PLOTLY_BASE['xaxis'], 'tickangle': -40, 'automargin': True,
               'title': dict(text='Plato', font=dict(color=_CH_AXIS, size=11))},
        yaxis={**PLOTLY_BASE['yaxis'],
               'title': dict(text='Día', font=dict(color=_CH_AXIS, size=11))})
    fig.update_layout(**layout)
    return _theme(fig)

def chart_trend(df, dishes):
    """¿Los platos más vendidos están subiendo o bajando? — Líneas de tendencia."""
    if df.empty or not dishes: return go.Figure()
    colors = [GOLD, '#2DD4BF', '#F472B6', '#A78BFA', '#4F96FF', '#E8C878', '#F97316']
    fig = go.Figure()

    # Franjas de fin de semana — contexto de alta demanda esperada
    if not df.empty:
        for d in pd.date_range(df['date'].min(), df['date'].max(), freq='D'):
            if d.weekday() in (5, 6):
                fig.add_vrect(x0=d - pd.Timedelta(hours=12),
                              x1=d + pd.Timedelta(hours=12),
                              fillcolor='rgba(201,169,122,0.07)',
                              line_width=0, layer='below')

    annotations = []
    for i, dish in enumerate(dishes[:5]):   # máximo 5 platos para no saturar
        data = (df[df['dish'] == dish].groupby('date')['quantity']
                .sum().reset_index().sort_values('date'))
        if data.empty: continue
        data['roll'] = data['quantity'].rolling(4, min_periods=1).mean()
        c = colors[i % len(colors)]

        # Puntos crudos: visibles pero no en hover (contexto de variabilidad diaria)
        fig.add_trace(go.Scatter(x=data['date'], y=data['quantity'],
            mode='markers', marker=dict(size=5, color=c, opacity=0.25),
            showlegend=False, name=dish, hoverinfo='skip'))
        # Línea de tendencia (media 4 días — más reactiva con 38 días de datos)
        fig.add_trace(go.Scatter(x=data['date'], y=data['roll'], name=dish,
            mode='lines', line=dict(width=2.5, color=c),
            hovertemplate=f'<b>{dish[:28]}</b><br>%{{x|%d/%m}}: <b>%{{y:.0f}}</b> uds<extra></extra>'))

        # Label directo al final de la línea — sin necesidad de buscar en la leyenda
        if len(data) > 0:
            last_x  = data['date'].iloc[-1]
            last_y  = float(data['roll'].iloc[-1])
            # Tendencia: comparar última semana vs primera semana
            first_w = float(data['roll'].iloc[:4].mean()) if len(data) >= 4 else last_y
            trend_icon = '↑' if last_y > first_w * 1.05 else ('↓' if last_y < first_w * 0.95 else '→')
            short_name = dish.split(' ')[0]   # primera palabra del plato
            annotations.append(dict(
                x=last_x, y=last_y,
                text=f'  {trend_icon} {short_name}',
                showarrow=False, xanchor='left',
                font=dict(size=10, color=c)))

    layout = {**PLOTLY_BASE}
    layout.update(
        title=dict(
            text='¿Los platos más vendidos están subiendo o bajando?'
                 '<br><sup style="font-size:11px;color:rgba(201,169,122,.55)">'
                 'Línea = tendencia reciente · punto = venta real del día'
                 ' · sombreado = fin de semana (alta demanda normal)</sup>',
            **PLOTLY_BASE['title']),
        height=460, hovermode='x unified',
        annotations=annotations,
        margin=dict(l=16, r=120, t=88, b=16),
        xaxis=dict(**PLOTLY_BASE['xaxis'], tickformat='%d %b'),
        yaxis=dict(**PLOTLY_BASE['yaxis'], title='Unidades producidas'))
    fig.update_layout(**layout)
    return _theme(fig)

def chart_variability(model, cat='Fondo', top_n=18):
    # Agrega ponderando por n (observaciones por tipo de día)
    # mean_n / n_total  →  media ponderada global
    # std_n  / n_std    →  std ponderada solo sobre tipos-de-día con std disponible
    accum: dict = {}
    for dt, dishes in model.items():
        for (name, c), stats in dishes.items():
            if c != cat:
                continue
            n = stats.get('n', 1)
            if name not in accum:
                accum[name] = {'mean_n': 0.0, 'std_n': 0.0, 'n': 0, 'n_std': 0}
            accum[name]['mean_n'] += stats['mean'] * n
            accum[name]['n']      += n
            if stats['std'] is not None:
                accum[name]['std_n']  += stats['std'] * n
                accum[name]['n_std']  += n        # solo cuenta donde existe std

    if not accum:
        return go.Figure()

    rows = []
    for name, a in accum.items():
        if a['n'] == 0:
            continue
        mean_w = a['mean_n'] / a['n']
        # Denominador correcto: solo las observaciones donde se pudo calcular std
        std_w  = a['std_n'] / a['n_std'] if a['n_std'] > 0 else 0
        cv     = std_w / mean_w if mean_w > 0 and std_w > 0 else 0
        rows.append({'dish': name, 'mean': mean_w, 'std': std_w, 'cv': cv})

    if not rows:
        return go.Figure()

    df_v = pd.DataFrame(rows)
    top  = df_v.nlargest(top_n, 'mean')['dish']
    agg  = df_v[df_v['dish'].isin(top)].sort_values('mean', ascending=False)

    has_std = agg['std'].gt(0).any()
    error_x = dict(type='data', array=agg['std'].tolist(),
                   color='rgba(110,231,183,.6)', thickness=1.5, width=6) if has_std else None
    # customdata lleva el std por plato para mostrarlo en el hover (error_x.array no es interpolable)
    hover = '<b>%{y}</b><br>Promedio: %{x:.1f}' + (
        ' ± %{customdata:.1f}' if has_std else '') + '<extra></extra>'

    fig = go.Figure()
    fig.add_trace(go.Bar(
        y=agg['dish'], x=agg['mean'], name='Promedio', orientation='h',
        customdata=agg['std'].tolist(),
        error_x=error_x,
        marker=dict(
            color=[f'rgba(201,169,122,{max(.25, 1 - row.cv * .8):.2f})' for _, row in agg.iterrows()],
            cornerradius=4),
        hovertemplate=hover))

    subtitle = ' (promedio ponderado por observaciones)'
    layout = {**PLOTLY_BASE}
    layout.update(
        title=dict(text=f'Variabilidad — {cat}{subtitle}', **PLOTLY_BASE['title']),
        height=max(380, len(agg) * 30), showlegend=False,
        margin=dict(l=220, r=64, t=48, b=16),
        yaxis=dict(**PLOTLY_BASE['yaxis'], autorange='reversed', automargin=True),
        xaxis=dict(**PLOTLY_BASE['xaxis'], title='Unidades promedio por día'))
    fig.update_layout(**layout)
    return _theme(fig)

def chart_promo_impact(model, top_n=15):
    normal_d = model.get('Domingo',   {})
    promo_d  = model.get('Dom-Promo', {})
    if not promo_d: return go.Figure()
    all_promo = [(n, s['mean']) for (n, c), s in promo_d.items() if c == 'Fondo']
    dishes    = sorted(all_promo, key=lambda x: -x[1])[:top_n]
    if not dishes: return go.Figure()
    names    = [d[0] for d in dishes]
    promo_v  = [d[1] for d in dishes]
    normal_v = [next((s['mean'] for (n,c),s in normal_d.items() if n==name and c=='Fondo'), 0)
                for name in names]
    factors  = [p/n if n > 0 else 0 for p, n in zip(promo_v, normal_v)]
    fig = go.Figure()
    # Barras side-by-side (group) — overlay solapaba y ocultaba una barra
    fig.add_trace(go.Bar(
        name='Dom. Normal', y=names, x=normal_v, orientation='h',
        marker=dict(color='rgba(130,105,65,.52)', cornerradius=4),
        hovertemplate='<b>%{y}</b><br>Domingo Normal: <b>%{x:.1f}</b> uds<extra></extra>'))
    fig.add_trace(go.Bar(
        name='Dom-Promo AMEX ⭐', y=names, x=promo_v, orientation='h',
        marker=dict(color=GOLD, opacity=0.88, cornerradius=4),
        hovertemplate='<b>%{y}</b><br>Dom-Promo AMEX: <b>%{x:.1f}</b> uds<extra></extra>'))
    for name, pv, nv, fct in zip(names, promo_v, normal_v, factors):
        if nv > 0 and fct > 0:
            fig.add_annotation(
                x=max(pv, nv), y=name, text=f'  {fct:.1f}×',
                showarrow=False, xanchor='left',
                font=dict(size=10, color=GOLD if fct >= 1.5 else _CH_AXIS))
    layout = {**PLOTLY_BASE}
    layout.update(
        barmode='group', bargap=0.32, bargroupgap=0.06,
        title=dict(text='Impacto AMEX — Fondos (Dom. Normal vs Dom-Promo)', **PLOTLY_BASE['title']),
        height=max(420, len(names)*52),
        margin=dict(l=220, r=80, t=56, b=16),
        yaxis={**PLOTLY_BASE['yaxis'], 'autorange': 'reversed', 'automargin': True},
        xaxis={**PLOTLY_BASE['xaxis'], 'title': 'Unidades promedio por domingo'})
    fig.update_layout(**layout)
    return _theme(fig)

def chart_weekly_total(df):
    """¿Cuánto se produjo cada semana? — Barras por semana."""
    if df.empty: return go.Figure()
    grp = pd.Grouper(key='date', freq='W')
    weekly_sum   = df.groupby(grp)['quantity'].sum().reset_index()
    _nd = df.groupby(grp)['date'].nunique(); _nd.name = 'n_days'
    weekly_days  = _nd.reset_index()
    weekly_promo = df.groupby(grp)['is_promo'].any().reset_index()
    weekly_sum.columns   = ['week', 'total']
    weekly_days.columns  = ['week', 'n_days']
    weekly_promo.columns = ['week', 'has_promo']
    weekly = weekly_sum.merge(weekly_days, on='week').merge(weekly_promo, on='week')
    # Promedio diario normalizado para comparar semanas con distinto número de días activos
    weekly['avg_day'] = (weekly['total'] / weekly['n_days']).round(0)

    # Colores por tipo de semana
    bar_colors = [GOLD if hp else 'rgba(130,105,65,.65)' for hp in weekly['has_promo']]
    labels = [
        f"Sem. {'con ⭐ AMEX' if hp else 'normal'}<br>{w.strftime('%d/%m')}"
        for w, hp in zip(weekly['week'], weekly['has_promo'])
    ]

    fig = go.Figure()
    # Una sola traza de barras con colores individuales — más claro que dos líneas
    fig.add_trace(go.Bar(
        x=[w.strftime('%d/%m') for w in weekly['week']],
        y=weekly['avg_day'],
        text=weekly['avg_day'].astype(int),
        textposition='outside',
        textfont=dict(size=11, color=_CH_FONT),
        cliponaxis=False,
        marker=dict(color=bar_colors, cornerradius=5),
        customdata=list(zip(weekly['n_days'], weekly['has_promo'])),
        hovertemplate=(
            'Semana del %{x}<br>'
            '<b>%{y:.0f}</b> uds/día promedio<br>'
            'Días activos: %{customdata[0]}'
            '<extra>%{marker.color}</extra>'),
        showlegend=False))

    # Línea de promedio general — referencia visual
    avg_total = float(weekly['avg_day'].mean())
    fig.add_hline(
        y=avg_total, line_dash='dot',
        line_color='rgba(201,169,122,.45)', line_width=1.5,
        annotation_text=f'  Promedio: {avg_total:.0f} uds/día',
        annotation_position='right',
        annotation_font=dict(size=10, color=_CH_AXIS))

    # Anotaciones para semanas con días excluidos
    excluded_weeks: dict = {}
    for ex_date in EXCLUDED:
        week_end = ex_date + timedelta(days=(6 - ex_date.weekday()))
        ts = pd.Timestamp(week_end)
        excluded_weeks[ts] = excluded_weeks.get(ts, 0) + 1

    annotations = []
    for ts_key, n_excl in excluded_weeks.items():
        x_label = ts_key.strftime('%d/%m')
        match = weekly[weekly['week'] == ts_key]
        if not match.empty:
            y_val = float(match['avg_day'].iloc[0])
            annotations.append(dict(
                x=x_label, y=y_val,
                text=f'⚠ {n_excl} día{"s" if n_excl > 1 else ""}<br>sin datos',
                showarrow=True, arrowhead=2, arrowcolor=_CH_AXIS,
                arrowsize=1, arrowwidth=1.2, ax=0, ay=-50,
                font=dict(size=9, color=_CH_AXIS),
                bgcolor='rgba(9,8,12,.80)', bordercolor=_CH_AXIS, borderwidth=1,
                borderpad=4))

    # Leyenda manual como annotation — más clara que leyenda de Plotly
    annotations += [
        dict(x=1.01, y=1.0, xref='paper', yref='paper',
             text='<b style="color:#C9A97A">■</b> Semana con AMEX ⭐',
             showarrow=False, xanchor='left', font=dict(size=10, color=_CH_FONT)),
        dict(x=1.01, y=0.92, xref='paper', yref='paper',
             text='<b style="color:rgba(130,105,65,.65)">■</b> Semana normal',
             showarrow=False, xanchor='left', font=dict(size=10, color=_CH_FONT)),
    ]

    layout = {**PLOTLY_BASE}
    layout.update(
        title=dict(
            text='¿Cuánto se produjo cada semana?'
                 '<br><sup style="font-size:11px;color:rgba(201,169,122,.55)">'
                 'Unidades por día promedio · dorado = semana con domingo AMEX</sup>',
            **PLOTLY_BASE['title']),
        height=380, annotations=annotations,
        margin=dict(l=16, r=160, t=80, b=48),
        xaxis=dict(**PLOTLY_BASE['xaxis'], title='Semana'),
        yaxis=dict(**PLOTLY_BASE['yaxis'], title='Unidades / día'))
    fig.update_layout(**layout)
    return _theme(fig)

# ── Ventas financieras ─────────────────────────────────────────────────────────

VENTAS_DIR  = os.path.join(BASE_DIR, "Ventas")
MESES_MAP_V = {'Ene':1,'Feb':2,'Mar':3,'Abr':4,'May':5,'Jun':6,
               'Jul':7,'Ago':8,'Sep':9,'Oct':10,'Nov':11,'Dic':12}
MESES_NOM_V = ['','Ene','Feb','Mar','Abr','May','Jun',
               'Jul','Ago','Sep','Oct','Nov','Dic']
GRUPOS_VTA  = {'Alimentos','Bebidas S/Alcohol','Bebidas C/Alcohol','Vinos'}

def load_ventas_df() -> pd.DataFrame:
    rows = []
    if not os.path.isdir(VENTAS_DIR):
        return pd.DataFrame()
    for fname in sorted(os.listdir(VENTAS_DIR)):
        if not fname.endswith('.xls'): continue
        mes_str = fname[:3]
        try:
            anio = int('20' + fname[3:5])
            mes  = MESES_MAP_V.get(mes_str, 0)
            if mes == 0: continue
        except Exception:
            continue
        try:
            if not HAS_XLRD: continue
            wb = xlrd.open_workbook(os.path.join(VENTAS_DIR, fname))
            sh = wb.sheets()[0]
            for r in range(sh.nrows):
                try:
                    nombre = str(sh.cell_value(r, 7)).strip()
                    if nombre in GRUPOS_VTA:
                        rows.append({'año': anio, 'mes': mes, 'categoria': nombre,
                                     'bruto':    float(sh.cell_value(r, 21) or 0),
                                     'cantidad': float(sh.cell_value(r, 16) or 0),
                                     'descuento':float(sh.cell_value(r, 28) or 0)})
                except Exception:
                    pass
        except Exception:
            pass
    return pd.DataFrame(rows) if rows else pd.DataFrame()

# ── Recetas ────────────────────────────────────────────────────────────────────

RECETAS_PATH = os.path.join(BASE_DIR, "recetas.json")

def load_recetas():
    try:
        with open(RECETAS_PATH, encoding='utf-8') as f:
            raw = json.load(f)
        return raw, {k: v for k, v in raw.items() if not k.startswith('_')}
    except Exception:
        return {}, {}

def save_recetas(raw: dict):
    _atomic_write(RECETAS_PATH, raw)

def calc_costos(recetas_raw: dict, model: dict, df=None) -> list[dict]:
    """Calcula food cost por plato. Retorna lista de dicts ordenada por nombre."""
    # Popularidad: media real de unidades/día desde el historial completo.
    # Usando df evita el sesgo de acumular medias por tipo de día (algunos platos
    # solo aparecen en ciertos días, distorsionando la comparación entre platos).
    pop_avg: dict[str, float] = {}
    if df is not None and not df.empty:
        grp = df.groupby('dish')['quantity'].mean()
        pop_avg = grp.to_dict()
    else:
        # Fallback: media ponderada por N observaciones históricas de cada tipo de día
        pop_sum: dict[str, float] = {}
        pop_n:   dict[str, int]   = {}
        for dt, dishes in model.items():
            if dt == 'Dom-Promo':
                continue
            for (name, cat), stats in dishes.items():
                n = stats.get('n', 1)
                pop_sum[name] = pop_sum.get(name, 0) + stats['mean'] * n
                pop_n[name]   = pop_n.get(name, 0) + n
        pop_avg = {n: pop_sum[n] / pop_n[n] for n in pop_sum if pop_n.get(n, 0) > 0}

    rows = []
    for sku, rec in recetas_raw.items():
        if sku.startswith('_'):
            continue
        nombre       = rec.get('nombre', sku)
        precio_venta = float(rec.get('precio_venta', 0) or 0)
        ings         = rec.get('ingredientes', [])

        # Costo total = suma(cantidad × precio_unitario por ingrediente)
        costo = sum(
            float(i.get('cantidad', 0) or 0) * float(i.get('precio_unitario', 0) or 0)
            for i in ings
        )

        food_cost_pct = (costo / precio_venta * 100) if precio_venta > 0 else None
        margen        = (precio_venta - costo)        if precio_venta > 0 else None
        popularidad   = pop_avg.get(nombre)

        rows.append({
            'sku':           sku,
            'nombre':        nombre,
            'precio_venta':  precio_venta,
            'costo':         costo,
            'food_cost_pct': food_cost_pct,
            'margen':        margen,
            'popularidad':   popularidad,
            'n_ings':        len(ings),
        })

    rows.sort(key=lambda r: (r['nombre']))
    return rows

# ── Excel export ───────────────────────────────────────────────────────────────

def make_excel_bytes(day_details, totals, dish_stats, history_len, k_factor=K_SAFETY) -> bytes | None:
    if not HAS_OPENPYXL:
        return None
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Producción'
    NAVY   = PatternFill('solid', fgColor='1F3460')
    GOLD_F = PatternFill('solid', fgColor='C9A84C')
    GRAY   = PatternFill('solid', fgColor='F2F4F8')
    thin   = Side(style='thin', color='D0D5E0')
    brd    = Border(left=thin, right=thin, top=thin, bottom=thin)
    def c(r, col, val, bold=False, fill=None, clr=None, align='left'):
        cell = ws.cell(row=r, column=col, value=val)
        cell.font      = Font(bold=bold, color=clr or ('FFFFFF' if fill in (NAVY, GOLD_F) else '111827'))
        cell.fill      = fill or PatternFill()
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border    = brd
    ws.merge_cells('A1:E1')
    ws['A1'] = 'PRONÓSTICO DE PRODUCCIÓN — Margo / Nelí'
    ws['A1'].font      = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill      = NAVY
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 24
    period = '  +  '.join(
        f"{DAYS_ES[d.weekday()]} {d.strftime('%d/%m/%Y')} [{dt}]"
        for d, dt, _, _ in day_details)
    ws.merge_cells('A2:E2'); ws['A2'] = f'Período: {period}'
    ws.merge_cells('A3:E3')
    ws['A3'] = f'Generado: {fecha_es(datetime.now())}   Buffer: μ + {k_factor}σ   Modelo: {history_len} días'
    ws['A3'].font = Font(italic=True, color='6B7280')
    row = 5
    for cat in CAT_ORDER:
        items = {k: v for k, v in totals.items() if k[1] == cat and v >= 1}
        if not items: continue
        ws.merge_cells(f'A{row}:E{row}')
        ws[f'A{row}'] = cat.upper()
        ws[f'A{row}'].font      = Font(bold=True, color='FFFFFF')
        ws[f'A{row}'].fill      = GOLD_F
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1
        for lbl, col in [('Plato',1),('Promedio μ',2),('Buffer σ',3),('Total',4)]:
            c(row, col, lbl, bold=True, fill=GRAY, align='center')
        row += 1
        for (name, _), qty in sorted(items.items(), key=lambda x: -x[1]):
            st_d = dish_stats.get((name, _), {'mu': qty, 'sigma': 0})
            c(row, 1, name)
            c(row, 2, round(st_d['mu']),    align='center')
            c(row, 3, round(st_d['sigma']), align='center')
            c(row, 4, round(qty),           align='center', bold=True)
            row += 1
        row += 1
    ws.column_dimensions['A'].width = 40
    for col in ['B','C','D']:
        ws.column_dimensions[col].width = 14
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Inventario ─────────────────────────────────────────────────────────────────

INVENTARIO_PATH = os.path.join(BASE_DIR, "inventario.json")
_INV_DAYS = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado']
_INV_DNUM = {d: i for i, d in enumerate(_INV_DAYS)}


def build_all_ingredients(recetas: dict) -> dict:
    """Enumera todos los ingredientes únicos de todas las recetas.
    Retorna {key: {nombre, unidad, categoria, temperatura, precio_unitario, proveedor, dias_despacho}}
    donde key = 'nombre||unidad'.
    """
    result: dict = {}
    for sku, rec in recetas.items():
        for ing in rec.get('ingredientes', []):
            nombre = ing.get('nombre', '').strip()
            if not nombre:
                continue
            key = f"{nombre}||{ing.get('unidad', 'g')}"
            if key not in result:
                result[key] = {
                    'nombre':          nombre,
                    'unidad':          ing.get('unidad', 'g'),
                    'categoria':       ing.get('categoria', ''),
                    'temperatura':     ing.get('temperatura', ''),
                    'precio_unitario': float(ing.get('precio_unitario', 0) or 0),
                    'proveedor':       ing.get('proveedor', ''),
                    'dias_despacho':   list(ing.get('dias_despacho', [])),
                }
            else:
                # Actualizar campos no-nulos con datos más recientes
                if ing.get('precio_unitario'):
                    result[key]['precio_unitario'] = float(ing['precio_unitario'])
                if ing.get('proveedor'):
                    result[key]['proveedor'] = ing['proveedor']
    return result


def build_needed_ingredients(model: dict, start, horizon: int, k_factor: float,
                              recetas: dict, name_to_sku: dict) -> dict:
    """Agrega los ingredientes necesarios según el pronóstico.
    Retorna {key: {nombre, unidad, categoria, total, ...}}
    donde key = 'nombre||unidad'.
    """
    _, totals, _ = make_forecast(model, start, horizon, k_factor)
    result: dict = {}
    for (dname, _), total_qty in totals.items():
        sku = name_to_sku.get(dname)
        if not sku or sku not in recetas:
            continue
        for ing in recetas[sku].get('ingredientes', []):
            nombre = ing.get('nombre', '').strip()
            if not nombre:
                continue
            key = f"{nombre}||{ing.get('unidad', 'g')}"
            pu  = float(ing.get('precio_unitario', 0) or 0)
            if key not in result:
                result[key] = {
                    'nombre':          nombre,
                    'unidad':          ing.get('unidad', 'g'),
                    'categoria':       ing.get('categoria', ''),
                    'temperatura':     ing.get('temperatura', ''),
                    'total':           0.0,
                    'precio_unitario': pu,
                    'proveedor':       ing.get('proveedor', ''),
                    'dias_despacho':   list(ing.get('dias_despacho', [])),
                }
            result[key]['total'] += float(ing.get('cantidad', 0) or 0) * total_qty
            if pu:
                result[key]['precio_unitario'] = pu
            dd = ing.get('dias_despacho', [])
            result[key]['dias_despacho'] = sorted(
                set(result[key]['dias_despacho']) | set(dd),
                key=lambda x: _INV_DNUM.get(x, 9))
    return result


def load_inventario() -> dict:
    try:
        with open(INVENTARIO_PATH, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def save_inventario(data: dict):
    _atomic_write(INVENTARIO_PATH, data)


# ── Odoo integration ───────────────────────────────────────────────────────────

ODOO_MAPPING_PATH = os.path.join(BASE_DIR, "odoo_mapping.json")


def load_odoo_mapping() -> dict:
    try:
        with open(ODOO_MAPPING_PATH, encoding='utf-8') as f:
            return json.load(f)
    except Exception:
        return {}


def save_odoo_mapping(data: dict) -> None:
    _atomic_write(ODOO_MAPPING_PATH, data)


def sync_prices_from_odoo_mapping(mapping: dict) -> int:
    """
    Actualiza precio_unitario en recetas.json desde el mapping Odoo.
    Solo actualiza ingredientes con price > 0 en el mapping.
    Retorna cantidad de ingredientes actualizados.
    """
    raw, _ = load_recetas()
    count = 0
    for sku, receta in raw.items():
        if sku.startswith('_'):
            continue
        for ing in receta.get('ingredientes', []):
            nombre = ing.get('nombre', '').strip()
            if not nombre:
                continue
            key = f"{nombre}||{ing.get('unidad', 'g')}"
            entry = mapping.get(key, {})
            price = float(entry.get('price', 0) or 0)
            if price > 0:
                ing['precio_unitario'] = price
                if entry.get('supplier_name') and not ing.get('proveedor'):
                    ing['proveedor'] = entry['supplier_name']
                count += 1
    save_recetas(raw)
    return count
