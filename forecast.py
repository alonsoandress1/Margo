#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
forecast.py — Pronóstico de producción diaria
Margo / Nelí

Uso:
  python forecast.py                  → próximos 3 días desde mañana
  python forecast.py 14-06-2026       → 3 días desde esa fecha
  python forecast.py 14-06-2026 1     → solo ese día
"""

import os, sys, io, json
from datetime import datetime, timedelta
from collections import defaultdict

if sys.stdout.encoding and sys.stdout.encoding.lower() != 'utf-8':
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8', errors='replace')
    sys.stdin  = io.TextIOWrapper(sys.stdin.buffer,  encoding='utf-8', errors='replace')

try:
    import xlrd
except ImportError:
    sys.exit("Falta xlrd. Ejecuta:  pip install xlrd==1.2.0")

try:
    import openpyxl
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

try:
    from rich.console import Console  as _RC
    from rich.table   import Table    as _RT
    from rich.panel   import Panel    as _RP
    from rich.text    import Text     as _RTX
    from rich.columns import Columns  as _RCO
    from rich.align   import Align    as _RA
    from rich.rule    import Rule     as _RR
    from rich.padding import Padding  as _RPD
    from rich.layout  import Layout   as _RL
    from rich         import box      as _rbox
    HAS_RICH = True
except ImportError:
    HAS_RICH = False

# ── Configuración ──────────────────────────────────────────────────────────────

_SCRIPT_DIR  = os.path.dirname(os.path.abspath(__file__))
_CONFIG_PATH = os.path.join(_SCRIPT_DIR, "config.json")

# Carpeta de informes: lee config.json (igual que app.py) o usa Reportes/ por defecto
try:
    with open(_CONFIG_PATH, encoding="utf-8") as _f:
        _cfg = json.load(_f)
    _reports_folder = _cfg.get("reportes_folder", os.path.join(_SCRIPT_DIR, "Reportes"))
except Exception:
    _reports_folder = os.path.join(_SCRIPT_DIR, "Reportes")

FOLDER          = _SCRIPT_DIR      # directorio base para exportar HTML/Excel
PROMO_FROM      = datetime(2026, 6, 7)
K_SAFETY        = 0.7
BUFFER_FALLBACK = 0.20
DEFAULT_HORIZON = 3

EXCLUDED = {
    datetime(2026, 5, 18),
    datetime(2026, 5, 21),
    datetime(2026, 5, 22),
    datetime(2026, 5, 23),
    datetime(2026, 5, 24),
}
# EXCLUDED.add(datetime(2026, 9, 18))  ← agrega feriados futuros aquí

PREFIXES = {
    'VECARFON': 'Fondo',   'VECARFJT': 'Fondo',   'VECARFUB': 'Fondo',
    'VECARENT': 'Entrada',  'VECARAJT': 'Entrada',  'VECARAUB': 'Entrada',
    'VECARENS': 'Ensalada', 'VECAREJT': 'Ensalada', 'VECAREUB': 'Ensalada',
    'VECARPOS': 'Postre',   'VECARPUB': 'Postre',
    'VECARNIN': 'Niños',
    'VELICACO': 'Barra',   'VELICASA': 'Barra',
    'VELICATR': 'Barra',   'VELICDES': 'Barra',
    'VCVINCOP': 'Barra',   'VCVINCAR': 'Barra',
    'VCVINESP': 'Barra',   'VCVINSAU': 'Barra',
}

_EXCLUDE_SUFFIXES = (' PY', ' JT')

FERIADOS = {
    datetime(2026,  1,  1): "Año Nuevo",
    datetime(2026,  4,  3): "Viernes Santo",
    datetime(2026,  4,  4): "Sábado Santo",
    datetime(2026,  5,  1): "Día del Trabajo",
    datetime(2026,  5, 21): "Glorias Navales",
    datetime(2026,  6, 29): "San Pedro y San Pablo",
    datetime(2026,  7, 16): "Virgen del Carmen",
    datetime(2026,  8, 15): "Asunción de la Virgen",
    datetime(2026,  9, 18): "Fiestas Patrias",
    datetime(2026,  9, 19): "Glorias del Ejército",
    datetime(2026, 10, 12): "Encuentro de Dos Mundos",
    datetime(2026, 10, 31): "Iglesias Evangélicas",
    datetime(2026, 11,  1): "Todos los Santos",
    datetime(2026, 12,  8): "Inmaculada Concepción",
    datetime(2026, 12, 25): "Navidad",
    datetime(2027,  1,  1): "Año Nuevo",
}

CAT_ORDER  = ['Fondo', 'Entrada', 'Ensalada', 'Postre', 'Niños', 'Barra']
DAYS_ES    = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
MONTHS_ES  = ['','enero','febrero','marzo','abril','mayo','junio',
               'julio','agosto','septiembre','octubre','noviembre','diciembre']

def _fecha_es(d):
    return f"{d.day} de {MONTHS_ES[d.month]} de {d.year}"
CAT_COLORS = {
    'Fondo':    '#D4AF37',
    'Entrada':  '#60A5FA',
    'Ensalada': '#34D399',
    'Postre':   '#F472B6',
    'Niños':    '#A78BFA',
}
CAT_HEX = {
    'Fondo':    'D4AF37',
    'Entrada':  '60A5FA',
    'Ensalada': '34D399',
    'Postre':   'F472B6',
    'Niños':    'A78BFA',
}

# ── Lógica de negocio ──────────────────────────────────────────────────────────

def day_type(date):
    if date.weekday() == 6 and date >= PROMO_FROM:
        return 'Dom-Promo'
    return DAYS_ES[date.weekday()]


def parse_report(path):
    result = {}
    def handle_row(code, name, raw):
        code = str(code or '').strip()
        name = str(name or '').strip()
        if not code or not name:
            return
        if any(name.upper().endswith(s.upper()) for s in _EXCLUDE_SUFFIXES):
            return
        for prefix, cat in PREFIXES.items():
            if code.startswith(prefix):
                try:
                    qty = float(raw or 0)
                    if qty > 0:
                        result[(code, name, cat)] = qty
                except (ValueError, TypeError):
                    pass
                break

    if path.endswith('.xlsx'):
        if not HAS_OPENPYXL:
            return result
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
        for row in wb.active.iter_rows(min_row=22, values_only=True):
            if len(row) > 9:
                handle_row(row[0], row[1], row[9])
        wb.close()
    else:
        wb = xlrd.open_workbook(path)
        ws = wb.sheet_by_index(0)
        if ws.ncols > 9:
            for r in range(22, ws.nrows):
                handle_row(ws.cell_value(r, 0), ws.cell_value(r, 1), ws.cell_value(r, 9))
    return result


def load_history(folder):
    history = {}
    if not os.path.isdir(folder):
        return history
    for fname in sorted(os.listdir(folder)):
        if not fname.startswith('informe'):
            continue
        if not (fname.endswith('.xls') or fname.endswith('.xlsx')):
            continue
        raw = fname.replace('informe ', '').replace('.xlsx', '').replace('.xls', '').strip()
        try:
            d, m, y = raw.split('-')
            date = datetime(int(y), int(m), int(d))
        except ValueError:
            continue
        if date in EXCLUDED:
            continue
        data = parse_report(os.path.join(folder, fname))
        if data:
            history[date] = data
    return history


def build_model(history):
    accum = defaultdict(lambda: defaultdict(list))
    for date, data in history.items():
        dt = day_type(date)
        for (code, name, cat), qty in data.items():
            accum[dt][(name, cat)].append(qty)
    model = {}
    for dt, dishes in accum.items():
        model[dt] = {}
        for dish, vals in dishes.items():
            n    = len(vals)
            mean = sum(vals) / n
            std  = (sum((v - mean)**2 for v in vals) / n)**0.5 if n > 1 else None
            model[dt][dish] = {'mean': mean, 'std': std, 'n': n}
    return model


def dish_forecast(stats):
    mean, std = stats['mean'], stats['std']
    return mean + K_SAFETY * std if std is not None else mean * (1 + BUFFER_FALLBACK)


def make_forecast(model, start, horizon):
    day_details = []
    totals = defaultdict(float)
    mu_acc = defaultdict(float)
    sg_acc = defaultdict(float)

    for i in range(horizon):
        d  = start + timedelta(days=i)
        dt = day_type(d)
        day_model = model.get(dt)
        fallback  = day_model is None
        if fallback:
            day_model = model.get('Domingo', {})
        day_details.append((d, dt, day_model, fallback))
        for dish, stats in day_model.items():
            mean   = stats['mean']
            buffer = K_SAFETY * stats['std'] if stats['std'] is not None else mean * BUFFER_FALLBACK
            mu_acc[dish] += mean
            sg_acc[dish] += buffer
            totals[dish] += mean + buffer

    dish_stats = {d: {'mu': mu_acc[d], 'sigma': sg_acc[d]} for d in totals}
    return day_details, dict(totals), dish_stats


def _confidence(n):
    if n == 0: return '— sin datos'
    if n == 1: return '⚠ muy baja (1 obs)'
    if n < 4:  return f'△ moderada ({n} obs)'
    return          f'✓ buena ({n} obs)'


# ── Display en terminal (plain) ────────────────────────────────────────────────

def _bar(value, max_val, width=12):
    filled = round(value / max_val * width) if max_val else 0
    return '█' * filled + '░' * (width - filled)


def display_plain(day_details, totals, dish_stats, history, model):
    W = 74
    n_promo = sum(1 for d in history if day_type(d) == 'Dom-Promo')
    has_promo = any(dt == 'Dom-Promo' for _, dt, _, _ in day_details)
    print()
    print('═' * W)
    print('  PRONÓSTICO DE PRODUCCIÓN  —  Margo / Nelí'.center(W))
    print('═' * W)
    period = '  +  '.join(
        f"{DAYS_ES[d.weekday()][:3]} {d.strftime('%d/%m')} [{dt}{'*' if fb else ''}]"
        for d, dt, _, fb in day_details
    )
    print(f"\n  Período : {period}")
    print(f"  Buffer  : promedio + {K_SAFETY}σ por plato  |  Días en modelo: {len(history)}")
    if has_promo:
        print(f"  Dom-Promo: {_confidence(n_promo)}")
        if n_promo <= 2:
            print(f"  * Promo con < 2 obs — fallback {int(BUFFER_FALLBACK*100)}%")
    print(f"\n  {'─'*(W-2)}")
    print(f"  {'Día':<38} {'Fondos':>6}  {'Total platos':>12}")
    print(f"  {'─'*(W-2)}")
    for d, dt, dm, _ in day_details:
        label = f"{DAYS_ES[d.weekday()]} {d.strftime('%d/%m')} [{dt}]"
        f_fc  = sum(dish_forecast(s) for (_, c), s in dm.items() if c == 'Fondo')
        p_fc  = sum(dish_forecast(s) for s in dm.values())
        print(f"  {label:<38} {round(f_fc):>6}  {round(p_fc):>12}")
    print()
    for cat in CAT_ORDER:
        items = {k: v for k, v in totals.items() if k[1] == cat and v >= 1.0}
        if not items:
            continue
        max_val = max(items.values())
        print(f"  {'─'*4} {cat.upper()} {'─'*(W-8-len(cat))}")
        print(f"  {'Plato':<32}  {'Prom':>5}  {'Buffer':>6}  {'Total':>5}")
        for (name, _), qty in sorted(items.items(), key=lambda x: -x[1]):
            st = dish_stats.get((name, _), {'mu': qty, 'sigma': 0})
            print(f"  {name:<32}  {round(st['mu']):>5}  {round(st['sigma']):>+6}  {round(qty):>5}  {_bar(qty, max_val)}")
        print()
    tf = round(sum(v for (_, c), v in totals.items() if c == 'Fondo'))
    ta = round(sum(totals.values()))
    print(f"  {'─'*(W-2)}")
    print(f"  TOTAL FONDOS  {tf:>5}   TOTAL PLATOS  {ta:>5}")
    print('═' * W)
    print()


# ── Display en terminal (rich) ─────────────────────────────────────────────────

_BLOCKS = ' ▏▎▍▌▋▊▉█'

def _sbar(value, max_val, width=22, color='#D4AF37'):
    """Smooth sub-character bar. No trailing spaces to avoid Rich ellipsis."""
    ratio         = min(value / max_val, 1.0) if max_val else 0
    total_eighths = round(ratio * width * 8)
    full          = total_eighths // 8
    partial       = total_eighths % 8
    filled_str    = _BLOCKS[8] * full + (_BLOCKS[partial] if partial else '')
    return f'[bold {color}]{filled_str}[/]'


def _cat_items_sorted(cat, items, dish_stats):
    """Return sorted list of (name, qty, name_style, total_style) for a category."""
    clr  = CAT_COLORS[cat]
    rows = []
    for i, ((name, _), qty) in enumerate(sorted(items.items(), key=lambda x: -x[1])):
        if   i == 0: ns, ts = f'bold {clr}', f'bold {clr}'
        elif i == 1: ns, ts = clr,            clr
        elif i < 5:  ns, ts = 'white',        'white'
        else:        ns, ts = '#3A4A60',       '#3A4A60'
        rows.append((name, round(qty), ns, ts))
    return rows


def display_rich(day_details, totals, dish_stats, history, model):
    console      = _RC()
    W            = console.width
    n_promo      = sum(1 for d in history if day_type(d) == 'Dom-Promo')
    has_promo    = any(dt == 'Dom-Promo' for _, dt, _, _ in day_details)
    total_fondos = round(sum(v for (_, c), v in totals.items() if c == 'Fondo'))
    total_all    = round(sum(totals.values()))
    fecha        = _fecha_es(day_details[0][0])

    # ── 1. HEADER ─────────────────────────────────────────────────────────────
    # Left: brand + title   Right: date
    hdr_grid = _RT.grid(expand=True, padding=(0, 1))
    hdr_grid.add_column(ratio=3)
    hdr_grid.add_column(ratio=1, justify='right')

    brand = _RTX()
    brand.append('MARGO · NELÍ  ', style='bold #D4AF37')
    brand.append('Pronóstico de Producción', style='bold white')

    period_parts = []
    for d, dt, _, _ in day_details:
        tag = ' [bold #D4AF37 on #1A1A00] AMEX [/]' if dt == 'Dom-Promo' else ''
        period_parts.append(
            f'[{"bold #D4AF37" if dt=="Dom-Promo" else "white"}]'
            f'{DAYS_ES[d.weekday()][:3]} {d.strftime("%d/%m")}[/]{tag}'
        )
    sub = _RTX.from_markup('  [#3A4A60]·[/]  '.join(period_parts))

    left = _RTX()
    left.append_text(brand)
    left.append('\n')
    left.append_text(sub)

    right = _RTX()
    right.append(fecha, style='dim #3A4A60')
    right.append(f'\n{len(history)} días · μ+{K_SAFETY}σ', style='dim #2A3A50')

    hdr_grid.add_row(left, right)
    console.print()
    console.print(_RP(hdr_grid, border_style='#D4AF37', padding=(0, 2),
                      style='on #080E1E'))
    console.print()

    # ── 2. KPI STRIP — single table, always 1 row, always aligned ───────────
    conf_c = '#F59E0B' if n_promo <= 1 else '#60A5FA' if n_promo < 4 else '#22C55E'
    conf_s = '⚠' if n_promo <= 1 else '△' if n_promo < 4 else '✓'

    kpi_t = _RT(box=_rbox.SIMPLE_HEAVY, expand=True,
                border_style='#1E2E48', padding=(0, 3),
                show_header=False)
    for _ in range(4):
        kpi_t.add_column(justify='center', ratio=1)

    def _kv(v, l, sub='', c='#D4AF37'):
        t = _RTX()
        t.append(f'{v:,}', style=f'bold {c}')
        t.append(f'\n{l}', style='bold #3A4A60')
        if sub:
            t.append(f'\n{sub}', style=f'dim {c}')
        return _RA.center(t)

    kpi_t.add_row(
        _kv(total_fondos, 'FONDOS',       f'{len(day_details)} días'),
        _kv(total_all,    'TOTAL PLATOS',  ''),
        _kv(len(history), 'DÍAS MODELO',   f'μ + {K_SAFETY}σ / plato'),
        _kv(n_promo,      'DOM-PROMO',     _confidence(n_promo), conf_c),
    )
    console.print(kpi_t)
    console.print()

    # ── 3. DAY ROW — single table, N columns, guaranteed same row ────────────
    all_ff = [sum(dish_forecast(s) for (_, c), s in dm.items() if c == 'Fondo')
              for _, _, dm, _ in day_details]
    all_pf = [sum(dish_forecast(s) for s in dm.values())
              for _, _, dm, _ in day_details]
    max_ff = max(all_ff) if all_ff else 1

    day_t = _RT(box=_rbox.SIMPLE_HEAVY, expand=True,
                border_style='#1E2E48', padding=(0, 2),
                show_header=False)
    for _ in day_details:
        day_t.add_column(ratio=1)

    cells = []
    for (d, dt, dm, _), ff, pf in zip(day_details, all_ff, all_pf):
        is_p  = dt == 'Dom-Promo'
        nc    = '#D4AF37' if is_p else 'white'
        bar_c = '#D4AF37' if is_p else '#4F96FF'
        t = _RTX()
        t.append(f'{DAYS_ES[d.weekday()]}', style=f'bold {nc}')
        t.append(f'  {d.strftime("%d/%m")}', style='dim #3A4A60')
        if is_p:
            t.append('  ◆ AMEX', style='bold #D4AF37')
        t.append(f'\n{round(ff):,}', style=f'bold {nc}')
        t.append(' fondos', style='dim #3A4A60')
        t.append('\n')
        t.append_text(_RTX.from_markup(_sbar(ff, max_ff, 20, bar_c)))
        t.append(f'\n{round(pf):,} platos', style='dim #3A4A60')
        cells.append(t)

    day_t.add_row(*cells)
    console.print(day_t)
    console.print()

    # ── 4. FONDOS (full width — primary category) ─────────────────────────────
    fondos = {k: v for k, v in totals.items() if k[1] == 'Fondo' and v >= 1.0}
    if fondos:
        max_f  = max(fondos.values())
        clr    = CAT_COLORS['Fondo']
        # bar width: terminal width minus fixed columns (~48 chars)
        # terminal - name(34) - μ(6) - +σ(6) - ▶(6) - padding(12) - margin(8)
        bar_w  = max(10, min(22, W - 72))

        tbl = _RT(
            box=_rbox.SIMPLE, show_edge=False, expand=False,
            title=f'[bold {clr}]FONDOS[/]  '
                  f'[dim #3A4A60]{len(fondos)} platos · '
                  f'{round(sum(fondos.values()))} unidades[/]',
            title_justify='left',
            header_style='dim #3A4A60', padding=(0, 1),
        )
        tbl.add_column('Plato',  min_width=28, max_width=34, no_wrap=True)
        tbl.add_column('μ',      justify='right', width=6, style='#3A4A60')
        tbl.add_column('+σ',     justify='right', width=6, style='#6EE7B7')
        tbl.add_column('▶',      justify='right', width=6)
        tbl.add_column('',       width=bar_w, no_wrap=True)

        for i, ((name, _), qty) in enumerate(sorted(fondos.items(), key=lambda x: -x[1])):
            st = dish_stats.get((name, _), {'mu': qty, 'sigma': 0})
            if   i == 0: ns, ts = f'bold {clr}', f'bold {clr}'
            elif i == 1: ns, ts = clr,             clr
            elif i < 4:  ns, ts = 'white',         'white'
            else:        ns, ts = '#4A5870',        '#4A5870'
            tbl.add_row(
                f'[{ns}]{name}[/]',
                str(round(st['mu'])),
                f'+{round(st["sigma"])}',
                f'[{ts}]{round(qty)}[/]',
                _sbar(qty, max_f, bar_w, clr),
            )

        console.print(tbl)
        console.print()

    # ── 5. OTHER CATEGORIES — true 2-column table, one render call ───────────
    # Build as a SINGLE table with 6 columns (L:name,total,bar  R:name,total,bar)
    # This guarantees equal space without Rich grid inconsistencies.
    others = ['Entrada', 'Ensalada', 'Postre', 'Niños']
    pairs  = [(others[i], others[i+1]) for i in range(0, len(others)-1, 2)]
    # Layout budget per side: name + total(4) + bar + padding(6) = W/2 - 3
    # Solve for name: name = W//2 - 3 - 4 - bar_w - 6
    side_bar  = 12
    name_w    = max(16, W // 2 - side_bar - 16)   # 16 = total+padding+spacer overhead

    for lc, rc in pairs:
        li = {k: v for k, v in totals.items() if k[1] == lc and v >= 1.0}
        ri = {k: v for k, v in totals.items() if k[1] == rc and v >= 1.0}
        if not li and not ri:
            continue

        lclr = CAT_COLORS[lc]; rclr = CAT_COLORS[rc]
        lrows = _cat_items_sorted(lc, li, dish_stats) if li else []
        rrows = _cat_items_sorted(rc, ri, dish_stats) if ri else []
        lmax  = max((q for _, q, _, _ in lrows), default=1)
        rmax  = max((q for _, q, _, _ in rrows), default=1)
        ltot  = sum(q for _, q, _, _ in lrows)
        rtot  = sum(q for _, q, _, _ in rrows)

        lhdr = f'[bold {lclr}]{lc.upper()}[/]  [dim #3A4A60]{len(li)} · {ltot} uds[/]'
        rhdr = f'[bold {rclr}]{rc.upper()}[/]  [dim #3A4A60]{len(ri)} · {rtot} uds[/]'

        tbl = _RT(
            box=_rbox.SIMPLE, show_edge=False, expand=False,
            padding=(0, 1), header_style='dim #3A4A60',
        )
        tbl.add_column(lhdr, width=name_w, no_wrap=True)
        tbl.add_column('▶',  justify='right', width=4)
        tbl.add_column('',   width=side_bar, no_wrap=True)
        tbl.add_column(' ',  width=1)
        tbl.add_column(rhdr, width=name_w, no_wrap=True)
        tbl.add_column('▶',  justify='right', width=4)
        tbl.add_column('',   width=side_bar, no_wrap=True)

        n_rows = max(len(lrows), len(rrows))
        for idx in range(n_rows):
            ln, lq, lns, lts = lrows[idx] if idx < len(lrows) else ('', '', '', '')
            rn, rq, rns, rts = rrows[idx] if idx < len(rrows) else ('', '', '', '')
            tbl.add_row(
                f'[{lns}]{ln}[/]' if ln else '',
                f'[{lts}]{lq}[/]' if ln else '',
                _sbar(lq, lmax, side_bar, lclr) if ln else '',
                '',
                f'[{rns}]{rn}[/]' if rn else '',
                f'[{rts}]{rq}[/]' if rn else '',
                _sbar(rq, rmax, side_bar, rclr) if rn else '',
            )

        console.print(tbl)
        console.print()

    # ── 6. FOOTER — single rule line, never wraps ─────────────────────────────
    console.print()
    console.rule(
        f'[bold #3A4A60]FONDOS[/] [bold #D4AF37]{total_fondos:,}[/]  '
        f'[dim #3A4A60]·[/]  [bold #3A4A60]PLATOS[/] [bold #D4AF37]{total_all:,}[/]  '
        f'[dim #3A4A60]·  {fecha}  ·  {len(history)} días  ·  μ+{K_SAFETY}σ[/]',
        style='#1E2E48',
    )

    if has_promo and n_promo <= 2:
        console.print()
        console.print(_RP(
            f'[bold #F59E0B]⚠  Dom-Promo AMEX[/]  ·  {_confidence(n_promo)}  ·  '
            '[dim #3A4A60]El modelo mejora con cada domingo nuevo[/]',
            border_style='#7C3E00', padding=(0, 2),
        ))
    console.print()


def display(day_details, totals, dish_stats, history, model):
    if HAS_RICH:
        display_rich(day_details, totals, dish_stats, history, model)
    else:
        display_plain(day_details, totals, dish_stats, history, model)


# ── HTML export ────────────────────────────────────────────────────────────────

_CSS = """<style>
@import url('https://fonts.googleapis.com/css2?family=Cormorant+Garamond:ital,wght@0,300;0,400;0,600;1,300;1,400&family=DM+Sans:wght@300;400;500;600;700&family=Space+Mono:wght@400;700&display=swap');

:root {
  --bg:       #040810;
  --bg-1:     #070E1C;
  --bg-2:     #0B1628;
  --bg-3:     #102030;
  --bg-4:     #172A48;
  --border:   rgba(255,255,255,.06);
  --border-g: rgba(212,175,55,.22);
  --gold:     #D4AF37;
  --gold-2:   #E8CB72;
  --gold-3:   #8B6B1A;
  --gold-glow:rgba(212,175,55,.12);
  --text:     #EDF0F8;
  --text-2:   #A0AEC0;
  --text-3:   #4A5870;
  --ff-h: 'Cormorant Garamond', Georgia, 'Times New Roman', serif;
  --ff-b: 'DM Sans', 'Segoe UI', Helvetica, Arial, sans-serif;
  --ff-m: 'Space Mono', 'Consolas', 'Courier New', monospace;
}

*{box-sizing:border-box;margin:0;padding:0}
html{scroll-behavior:smooth}

::-webkit-scrollbar{width:5px}
::-webkit-scrollbar-track{background:var(--bg)}
::-webkit-scrollbar-thumb{background:var(--gold-3);border-radius:3px}
::-webkit-scrollbar-thumb:hover{background:var(--gold)}

body{
  font-family:var(--ff-b);
  background:var(--bg);
  color:var(--text);
  line-height:1.6;
  min-height:100vh;
  overflow-x:hidden;
}

/* mesh + grain */
body::before{
  content:'';position:fixed;inset:0;z-index:0;pointer-events:none;
  background:
    radial-gradient(ellipse 70% 60% at 15% 15%,rgba(212,175,55,.05) 0%,transparent 65%),
    radial-gradient(ellipse 80% 70% at 85% 85%,rgba(30,60,200,.07) 0%,transparent 65%),
    radial-gradient(ellipse 50% 50% at 55% 5%, rgba(120,60,200,.04) 0%,transparent 55%);
}
body::after{
  content:'';position:fixed;inset:0;z-index:0;pointer-events:none;
  opacity:.03;
  background-image:url("data:image/svg+xml,%3Csvg xmlns='http://www.w3.org/2000/svg' width='300' height='300'%3E%3Cfilter id='n'%3E%3CfeTurbulence type='fractalNoise' baseFrequency='.75' numOctaves='4' stitchTiles='stitch'/%3E%3C/filter%3E%3Crect width='300' height='300' filter='url(%23n)'/%3E%3C/svg%3E");
  background-size:300px 300px;
}

.wrap{position:relative;z-index:1}

/* ── HEADER ───────────────────────────────────────────────── */
.hdr{
  position:relative;
  padding:64px 80px 52px;
  overflow:hidden;
  border-bottom:1px solid var(--border-g);
}

.hdr-bg{
  position:absolute;inset:0;
  background:linear-gradient(150deg,#02050E 0%,#060F25 35%,#0A183A 65%,#050B1A 100%);
}

.hdr-glow{
  position:absolute;z-index:1;pointer-events:none;
  top:-40%;left:25%;width:90%;height:200%;
  background:radial-gradient(ellipse,rgba(212,175,55,.07) 0%,transparent 55%);
}

.hdr-grid{
  position:absolute;inset:0;z-index:1;pointer-events:none;
  background-image:
    linear-gradient(rgba(212,175,55,.03) 1px,transparent 1px),
    linear-gradient(90deg,rgba(212,175,55,.03) 1px,transparent 1px);
  background-size:60px 60px;
  mask-image:linear-gradient(to bottom,transparent,rgba(0,0,0,.4) 30%,rgba(0,0,0,.4) 70%,transparent);
}

.hdr-wm{
  position:absolute;z-index:1;
  right:-10px;bottom:-20px;
  font-family:var(--ff-h);
  font-size:200px;font-weight:700;
  color:rgba(212,175,55,.028);
  letter-spacing:-8px;line-height:1;
  user-select:none;white-space:nowrap;
  pointer-events:none;
}

.hdr-inner{
  position:relative;z-index:2;
  max-width:1100px;margin:0 auto;
  display:flex;justify-content:space-between;align-items:flex-end;gap:40px;
}

.hdr-brand{
  font-family:var(--ff-b);
  font-size:10px;font-weight:600;
  letter-spacing:.4em;text-transform:uppercase;
  color:var(--gold);
  margin-bottom:16px;
  display:flex;align-items:center;gap:10px;
}
.hdr-brand-line{width:28px;height:1px;background:var(--gold);display:inline-block}

.hdr-title{
  font-family:var(--ff-h);
  font-size:58px;font-weight:300;
  line-height:1.03;letter-spacing:-1.5px;
  color:var(--text);
}
.hdr-title em{font-style:italic;color:var(--gold-2);font-weight:400}

.hdr-period{
  margin-top:18px;
  font-size:12px;color:var(--text-3);letter-spacing:.07em;
  display:flex;align-items:center;gap:8px;flex-wrap:wrap;
}
.hdr-sep{color:var(--gold-3)}
.chip{
  display:inline-flex;align-items:center;
  background:var(--gold-glow);border:1px solid var(--border-g);
  color:var(--gold);font-size:9px;font-weight:700;
  letter-spacing:.15em;text-transform:uppercase;
  padding:2px 9px;border-radius:20px;
}

.hdr-right{text-align:right;flex-shrink:0}
.hdr-date{font-size:12px;color:var(--text-3);letter-spacing:.05em}
.hdr-model{font-size:11px;color:var(--text-3);margin-top:5px;opacity:.55}

/* ── CONTAINER ──────────────────────────────────────────────── */
.ct{max-width:1100px;margin:0 auto;padding:56px 80px}

/* ── KPI GRID ──────────────────────────────────────────────── */
.kpi-g{display:grid;grid-template-columns:repeat(4,1fr);gap:14px;margin-bottom:64px}

.kpi{
  position:relative;
  background:var(--bg-2);
  border:1px solid var(--border);
  border-radius:16px;
  padding:30px 24px 24px;
  overflow:hidden;
  transition:border-color .3s,transform .3s,box-shadow .3s;
  opacity:0;transform:translateY(20px);
}
.kpi.vis{animation:slideUp .6s cubic-bezier(.16,1,.3,1) forwards}

.kpi:hover{
  border-color:var(--border-g);
  transform:translateY(-3px);
  box-shadow:0 20px 40px rgba(0,0,0,.4),0 0 0 1px var(--border-g) inset;
}

.kpi::before{
  content:'';position:absolute;top:0;left:0;right:0;height:1px;
  background:linear-gradient(90deg,transparent 0%,var(--gold) 50%,transparent 100%);
  opacity:.6;
}
.kpi-promo{border-color:var(--border-g) !important;background:linear-gradient(135deg,#0B1A34,#112040)}
.kpi-promo::before{opacity:1;background:linear-gradient(90deg,var(--gold-3),var(--gold),var(--gold-2),var(--gold),var(--gold-3))}
.kpi-promo::after{
  content:'';position:absolute;inset:0;pointer-events:none;
  background:radial-gradient(ellipse 90% 70% at 50% -10%,rgba(212,175,55,.07) 0%,transparent 60%);
}

.kpi-v{
  font-family:var(--ff-m);
  font-size:50px;line-height:1;
  color:var(--gold);
  letter-spacing:-2px;
  margin-bottom:12px;
  text-shadow:0 0 30px rgba(212,175,55,.25);
}

.kpi-l{font-size:9px;font-weight:700;letter-spacing:.25em;text-transform:uppercase;color:var(--text-3)}
.kpi-s{margin-top:7px;font-size:11px;color:var(--text-3)}

.badge{
  display:inline-flex;align-items:center;gap:4px;
  padding:3px 10px;border-radius:20px;font-size:10px;font-weight:600;letter-spacing:.04em;
}
.b-warn{background:rgba(245,158,11,.1);color:#F59E0B;border:1px solid rgba(245,158,11,.22)}
.b-ok  {background:rgba(34,197,94,.08);color:#22C55E;border:1px solid rgba(34,197,94,.18)}
.b-mod {background:rgba(96,165,250,.08);color:#60A5FA;border:1px solid rgba(96,165,250,.18)}
.b-none{background:rgba(148,163,184,.07);color:var(--text-3);border:1px solid rgba(148,163,184,.12)}

/* ── SECTION ──────────────────────────────────────────────── */
.sec{margin-bottom:56px;opacity:0;transform:translateY(16px)}
.sec.vis{animation:slideUp .55s cubic-bezier(.16,1,.3,1) forwards}

.sec-hd{display:flex;align-items:baseline;gap:14px;margin-bottom:24px;padding-bottom:16px;border-bottom:1px solid var(--border)}

.sec-num{
  font-family:var(--ff-m);font-size:11px;font-weight:700;
  color:var(--text-3);letter-spacing:.1em;min-width:30px;
}

.sec-title{font-family:var(--ff-h);font-size:24px;font-weight:400;letter-spacing:.02em}
.sec-rule{flex:1;height:1px;background:linear-gradient(90deg,var(--border),transparent);margin-left:4px}
.sec-count{font-size:11px;color:var(--text-3);white-space:nowrap}

/* ── DAY CARDS ──────────────────────────────────────────────── */
.day-g{display:grid;grid-template-columns:repeat(auto-fit,minmax(160px,1fr));gap:12px}

.day-c{
  background:var(--bg-2);border:1px solid var(--border);border-radius:12px;
  padding:22px 18px;
  transition:border-color .25s,transform .25s;
  opacity:0;transform:translateY(14px);
}
.day-c.vis{animation:slideUp .5s cubic-bezier(.16,1,.3,1) forwards}
.day-c:hover{border-color:var(--border-g);transform:translateY(-2px)}

.day-c.promo{
  border-color:var(--border-g);
  background:linear-gradient(145deg,#0C1E3A,#132B52);
  position:relative;overflow:hidden;
}
.day-c.promo::after{
  content:'';position:absolute;inset:0;
  background:radial-gradient(ellipse 80% 60% at 50% 0%,rgba(212,175,55,.07),transparent 65%);
  pointer-events:none;
}

.day-dow{font-size:10px;font-weight:600;letter-spacing:.2em;text-transform:uppercase;color:var(--text-3);margin-bottom:4px}
.day-dn{
  font-family:var(--ff-h);font-size:46px;font-weight:300;
  color:var(--text);line-height:1;margin-bottom:16px;
}
.day-stat{font-size:12px;color:var(--text-3);margin-bottom:2px}
.day-stat b{color:var(--gold);font-family:var(--ff-m);font-weight:400}

/* ── TABLE ──────────────────────────────────────────────────── */
.tw{
  background:var(--bg-2);border:1px solid var(--border);border-radius:14px;
  overflow:hidden;
  opacity:0;transform:translateY(14px);
}
.tw.vis{animation:slideUp .5s cubic-bezier(.16,1,.3,1) forwards}

.dt{width:100%;border-collapse:collapse}

.dt thead tr{background:var(--bg-3)}

.dt th{
  padding:13px 16px;
  font-size:9px;font-weight:700;letter-spacing:.2em;text-transform:uppercase;
  color:var(--text-3);text-align:left;
  border-bottom:1px solid var(--border);
  white-space:nowrap;
}
.dt th.r{text-align:right}
.dt th.bh{width:180px}

.dt tbody tr{border-bottom:1px solid rgba(255,255,255,.025);transition:background .15s}
.dt tbody tr:last-child{border-bottom:none}
.dt tbody tr:hover{background:rgba(255,255,255,.02)}

.dt td{padding:12px 16px;font-size:13px;vertical-align:middle}

.dn{color:var(--text)}
.dn.r1{color:var(--gold);font-weight:600}
.dn.r2{color:var(--gold-2);font-weight:500}
.dn.r3{color:#C4B08A}

.num{text-align:right;font-family:var(--ff-m);font-size:12px;white-space:nowrap}
.np{color:var(--text-3)}
.nb{color:#6EE7B7}
.nt{font-weight:700;font-size:14px}

/* ── BAR ────────────────────────────────────────────────────── */
.bw{padding-right:20px;width:180px}

.bt{background:rgba(255,255,255,.04);border-radius:10px;height:8px;overflow:hidden;position:relative}

.bf{
  height:100%;border-radius:10px;
  transform-origin:left;transform:scaleX(0);
  transition:transform .7s cubic-bezier(.16,1,.3,1);
  position:relative;
}

/* glow on the leading edge */
.bf::after{
  content:'';position:absolute;
  top:-3px;right:-2px;width:6px;height:14px;
  background:inherit;border-radius:3px;
  filter:blur(3px) brightness(1.5);
  opacity:0;transition:opacity .3s;
}
.dt tr:hover .bf::after{opacity:.8}

/* ── ALERT ──────────────────────────────────────────────────── */
.alert{
  background:rgba(212,175,55,.05);
  border:1px solid rgba(212,175,55,.18);
  border-left:3px solid var(--gold);
  border-radius:0 10px 10px 0;
  padding:14px 20px;margin-bottom:36px;
  font-size:13px;color:var(--text-2);
}
.alert strong{color:var(--gold)}

/* ── TOTALS ─────────────────────────────────────────────────── */
.ttl{
  display:flex;
  background:var(--bg-2);border:1px solid var(--border);border-radius:16px;
  overflow:hidden;margin-top:60px;
  opacity:0;transform:translateY(14px);
}
.ttl.vis{animation:slideUp .5s ease forwards}

.ti{
  flex:1;padding:30px 36px;border-right:1px solid var(--border);
  transition:background .2s;
}
.ti:last-child{border-right:none}
.ti:hover{background:rgba(212,175,55,.03)}
.ti dt{font-size:9px;font-weight:700;letter-spacing:.22em;text-transform:uppercase;color:var(--text-3);margin-bottom:8px}
.ti dd{font-family:var(--ff-m);font-size:44px;color:var(--gold);letter-spacing:-2px;line-height:1;text-shadow:0 0 30px rgba(212,175,55,.2)}

/* ── FOOTER ─────────────────────────────────────────────────── */
.ftr{
  padding:24px 80px;
  border-top:1px solid var(--border);
  display:flex;align-items:center;justify-content:space-between;
  font-size:11px;color:var(--text-3);letter-spacing:.06em;
}
.ftr-b{color:var(--gold);font-weight:600}

/* ── ANIMATIONS ─────────────────────────────────────────────── */
@keyframes slideUp{
  from{opacity:0;transform:translateY(20px)}
  to  {opacity:1;transform:translateY(0)}
}

/* ── PRINT ──────────────────────────────────────────────────── */
@media print{
  *{animation:none!important;transition:none!important}
  body,body::before,body::after{background:#fff!important;color:#111!important}
  body::after{display:none}
  .hdr,.hdr-bg{background:#1A2C5B!important}
  .hdr-wm,.hdr-grid{display:none}
  .kpi,.day-c,.tw,.ttl{background:#f5f6fa!important;border-color:#d0d4e0!important;opacity:1!important;transform:none!important}
  .kpi-v,.day-dn,.ti dd{color:#1A2C5B!important;text-shadow:none!important}
  .bf{transform:scaleX(1)!important}
  .ftr{border-top-color:#ddd}
}
</style>"""

_JS = """<script>
(function(){
  function easeOutQuart(t){return 1-Math.pow(1-t,4)}

  function animCount(el,target,dur){
    var s=performance.now();
    (function tick(now){
      var p=Math.min((now-s)/dur,1),v=Math.round(easeOutQuart(p)*target);
      el.textContent=v.toLocaleString('es-CL');
      if(p<1) requestAnimationFrame(tick);
    })(s);
  }

  /* Intersection observer — trigger animations when elements enter viewport */
  var io=new IntersectionObserver(function(entries){
    entries.forEach(function(e){
      if(!e.isIntersecting) return;
      var el=e.target;
      el.classList.add('vis');
      /* animate bars inside this element */
      el.querySelectorAll('.bf').forEach(function(b,i){
        setTimeout(function(){b.style.transform='scaleX(1)'},i*32);
      });
      io.unobserve(el);
    });
  },{threshold:.08});

  document.addEventListener('DOMContentLoaded',function(){
    /* observe animatable elements */
    document.querySelectorAll('.kpi,.day-c,.sec,.tw,.ttl').forEach(function(el){
      io.observe(el);
    });

    /* stagger KPI animation delays */
    document.querySelectorAll('.kpi').forEach(function(el,i){
      el.style.animationDelay=(0.08+i*0.08)+'s';
    });

    /* count up numbers */
    setTimeout(function(){
      document.querySelectorAll('[data-n]').forEach(function(el,i){
        setTimeout(function(){animCount(el,parseInt(el.dataset.n),1300)},i*110);
      });
    },350);
  });
})();
</script>"""


_BAR_GRAD = {
    'Fondo':    'linear-gradient(90deg,#7A5C10,#D4AF37,#EDD070)',
    'Entrada':  'linear-gradient(90deg,#1440A0,#4F96FF,#9EC8FF)',
    'Ensalada': 'linear-gradient(90deg,#0A6E65,#2DD4BF,#7EEEE7)',
    'Postre':   'linear-gradient(90deg,#901040,#FB7185,#FDB0BC)',
    'Niños':    'linear-gradient(90deg,#4B1EA8,#A78BFA,#CFC0FD)',
}
_CAT_CLR = {
    'Fondo':'#D4AF37','Entrada':'#4F96FF','Ensalada':'#2DD4BF','Postre':'#FB7185','Niños':'#A78BFA',
}


def export_html(day_details, totals, dish_stats, history):
    first        = day_details[0][0]
    fname        = f"produccion_{first.strftime('%Y%m%d')}.html"
    n_promo      = sum(1 for d in history if day_type(d) == 'Dom-Promo')
    has_promo    = any(dt == 'Dom-Promo' for _, dt, _, _ in day_details)
    total_fondos = round(sum(v for (_, c), v in totals.items() if c == 'Fondo'))
    total_all    = round(sum(totals.values()))
    now          = datetime.now()
    gen_date     = f"{_fecha_es(now)}, {now.strftime('%H:%M')}"

    # ── Period header ──────────────────────────────────────────
    period_parts = []
    for d, dt, _, _ in day_details:
        tag = ' <span class="chip">AMEX</span>' if dt == 'Dom-Promo' else ''
        period_parts.append(
            f'<span>{DAYS_ES[d.weekday()]} {d.strftime("%d/%m")}{tag}</span>'
        )
    period_html = '<span class="hdr-sep"> · </span>'.join(period_parts)

    # ── KPI cards ─────────────────────────────────────────────
    def badge(n):
        if n == 0: return '<span class="badge b-none">Sin datos</span>'
        if n == 1: return '<span class="badge b-warn">⚠ 1 obs</span>'
        if n < 4:  return f'<span class="badge b-mod">△ {n} obs</span>'
        return         f'<span class="badge b-ok">✓ {n} obs</span>'

    kpis = f"""
<div class="kpi-g">
  <div class="kpi"><div class="kpi-v" data-n="{total_fondos}">0</div>
    <div class="kpi-l">Fondos a producir</div>
    <div class="kpi-s">{len(day_details)} días de horizonte</div></div>
  <div class="kpi"><div class="kpi-v" data-n="{total_all}">0</div>
    <div class="kpi-l">Total platos</div></div>
  <div class="kpi"><div class="kpi-v" data-n="{len(history)}">0</div>
    <div class="kpi-l">Días en modelo</div>
    <div class="kpi-s">buffer μ + {K_SAFETY}σ por plato</div></div>
  <div class="kpi kpi-promo"><div class="kpi-v" data-n="{n_promo}">0</div>
    <div class="kpi-l">Dom-Promo AMEX</div>
    <div class="kpi-s">{badge(n_promo)}</div></div>
</div>"""

    # ── Day cards ─────────────────────────────────────────────
    day_cards = '<div class="day-g">'
    for d, dt, dm, _ in day_details:
        f_fc = round(sum(dish_forecast(s) for (_, c), s in dm.items() if c == 'Fondo'))
        p_fc = round(sum(dish_forecast(s) for s in dm.values()))
        promo_cls = ' promo' if dt == 'Dom-Promo' else ''
        promo_chip = '<div class="chip" style="margin-top:8px;font-size:9px">AMEX PROMO</div>' if dt == 'Dom-Promo' else ''
        day_cards += f"""
  <div class="day-c{promo_cls}">
    <div class="day-dow">{DAYS_ES[d.weekday()]}</div>
    <div class="day-dn">{d.strftime('%d')}<small style="font-size:.4em;opacity:.5;margin-left:4px">{d.strftime('%m')}</small></div>
    <div class="day-stat">Fondos <b data-n="{f_fc}">0</b></div>
    <div class="day-stat">Platos <b data-n="{p_fc}">0</b></div>
    {promo_chip}
  </div>"""
    day_cards += '</div>'

    # ── Alert ─────────────────────────────────────────────────
    alert_html = ''
    if has_promo and n_promo <= 2:
        alert_html = (
            f'<div class="alert"><strong>⚠ Dom-Promo AMEX</strong>  ·  '
            f'Confianza {_confidence(n_promo)}  ·  '
            f'El modelo se afina con cada domingo nuevo. Mantén el buffer generoso.</div>'
        )

    # ── Category sections ─────────────────────────────────────
    cat_sections = ''
    for ci, cat in enumerate(CAT_ORDER):
        items = {k: v for k, v in totals.items() if k[1] == cat and v >= 1.0}
        if not items:
            continue
        clr     = _CAT_CLR.get(cat, '#D4AF37')
        grad    = _BAR_GRAD.get(cat, _BAR_GRAD['Fondo'])
        max_val = max(items.values())
        num     = f'0{ci+1}' if ci < 9 else str(ci+1)

        rows = ''
        for i, ((name, _), qty) in enumerate(sorted(items.items(), key=lambda x: -x[1])):
            st    = dish_stats.get((name, _), {'mu': qty, 'sigma': 0})
            mu    = round(st['mu'])
            sigma = round(st['sigma'])
            pct   = round(qty / max_val * 100, 1) if max_val else 0
            rank  = ['r1','r2','r3'][i] if i < 3 else ''
            rows += f"""
      <tr>
        <td class="dn {rank}">{name}</td>
        <td class="num np">{mu}</td>
        <td class="num nb">+{sigma}</td>
        <td class="num nt" style="color:{clr}">{round(qty)}</td>
        <td class="bw"><div class="bt"><div class="bf" style="width:{pct}%;background:{grad}"></div></div></td>
      </tr>"""

        cat_sections += f"""
<div class="sec">
  <div class="sec-hd">
    <span class="sec-num">{num}</span>
    <span class="sec-title" style="color:{clr}">{cat}</span>
    <span class="sec-rule"></span>
    <span class="sec-count">{len(items)} platos</span>
  </div>
  <div class="tw">
    <table class="dt">
      <thead><tr>
        <th>Plato</th>
        <th class="r">Prom</th><th class="r">Buffer σ</th>
        <th class="r" style="color:{clr}">Total</th>
        <th class="bh"></th>
      </tr></thead>
      <tbody>{rows}
      </tbody>
    </table>
  </div>
</div>"""

    # ── Totals strip ──────────────────────────────────────────
    totals_html = f"""
<dl class="ttl">
  <div class="ti"><dt>Total Fondos</dt><dd data-n="{total_fondos}">0</dd></div>
  <div class="ti"><dt>Total Platos</dt><dd data-n="{total_all}">0</dd></div>
  <div class="ti"><dt>Horizonte</dt><dd>{len(day_details)}<small style="font-size:.5em;opacity:.5;margin-left:4px">días</small></dd></div>
  <div class="ti"><dt>Modelo</dt><dd>{len(history)}<small style="font-size:.5em;opacity:.5;margin-left:4px">obs</small></dd></div>
</dl>"""

    # ── Assemble ─────────────────────────────────────────────
    html = f"""<!DOCTYPE html>
<html lang="es">
<head>
<meta charset="UTF-8">
<meta name="viewport" content="width=device-width,initial-scale=1">
<title>Pronóstico · Margo / Nelí</title>
{_CSS}
</head>
<body>
<div class="wrap">

<header class="hdr">
  <div class="hdr-bg"></div>
  <div class="hdr-glow"></div>
  <div class="hdr-grid"></div>
  <div class="hdr-wm">MARGO</div>
  <div class="hdr-inner">
    <div>
      <div class="hdr-brand"><span class="hdr-brand-line"></span>Margo · Nelí · Producción</div>
      <h1 class="hdr-title">Pronóstico de<br><em>Producción</em></h1>
      <div class="hdr-period">{period_html}</div>
    </div>
    <div class="hdr-right">
      <div class="hdr-date">{gen_date}</div>
      <div class="hdr-model">{len(history)} días en modelo · μ + {K_SAFETY}σ</div>
    </div>
  </div>
</header>

<div class="ct">

  {kpis}

  <div class="sec">
    <div class="sec-hd">
      <span class="sec-num">—</span>
      <span class="sec-title" style="color:var(--gold)">Resumen por Día</span>
      <span class="sec-rule"></span>
    </div>
    {day_cards}
  </div>

  {alert_html}
  {cat_sections}
  {totals_html}

</div>

<footer class="ftr">
  <span><span class="ftr-b">Margo / Nelí</span> · forecast.py</span>
  <span>{gen_date}</span>
</footer>

</div>
{_JS}
</body>
</html>"""

    with open(fname, 'w', encoding='utf-8') as f:
        f.write(html)
    print(f"  Reporte guardado: {fname}")
    return fname


# ── Excel export ────────────────────────────────────────────────────────────────

def export_excel(day_details, totals, dish_stats, history):
    if not HAS_OPENPYXL:
        print("  openpyxl no instalado. Ejecuta: pip install openpyxl")
        return
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    first = day_details[0][0]
    fname = f"produccion_{first.strftime('%Y%m%d')}.xlsx"
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Producción'

    GRAY = PatternFill('solid', fgColor='F2F2F2')
    NAVY = PatternFill('solid', fgColor='1F4E79')
    GOLD = PatternFill('solid', fgColor='C9A84C')
    thin = Side(style='thin', color='CCCCCC')
    brd  = Border(left=thin, right=thin, top=thin, bottom=thin)

    def cell(r, c, value, bold=False, fill=None, clr=None, align='left'):
        cl = ws.cell(row=r, column=c, value=value)
        cl.font      = Font(bold=bold, color=clr or ('FFFFFF' if fill in (NAVY, GOLD) else '000000'))
        cl.fill      = fill or PatternFill()
        cl.alignment = Alignment(horizontal=align, vertical='center')
        cl.border    = brd
        return cl

    ws.merge_cells('A1:D1')
    ws['A1'] = 'PRONÓSTICO DE PRODUCCIÓN — Margo / Nelí'
    ws['A1'].font      = Font(bold=True, size=13, color='FFFFFF')
    ws['A1'].fill      = NAVY
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 22

    period = '  +  '.join(
        f"{DAYS_ES[d.weekday()]} {d.strftime('%d/%m/%Y')} [{dt}]"
        for d, dt, _, _ in day_details
    )
    ws.merge_cells('A2:D2')
    ws['A2'] = f'Período: {period}'
    ws['A2'].alignment = Alignment(horizontal='left')

    ws.merge_cells('A3:D3')
    ws['A3'] = (f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}   "
                f"Buffer: μ + {K_SAFETY}σ por plato   Días en modelo: {len(history)}")
    ws['A3'].font      = Font(italic=True, color='666666')
    ws['A3'].alignment = Alignment(horizontal='left')

    row = 5
    ws.merge_cells(f'A{row}:D{row}')
    ws[f'A{row}'] = 'RESUMEN POR DÍA'
    ws[f'A{row}'].font      = Font(bold=True, color='FFFFFF')
    ws[f'A{row}'].fill      = GOLD
    ws[f'A{row}'].alignment = Alignment(horizontal='center')
    row += 1

    for lbl, c in [('Día', 1), ('Tipo', 2), ('Fondos', 3), ('Total platos', 4)]:
        cell(row, c, lbl, bold=True, fill=GRAY, align='center')
    row += 1

    for d, dt, dm, _ in day_details:
        f_t = round(sum(dish_forecast(s) for (_, c), s in dm.items() if c == 'Fondo'))
        p_t = round(sum(dish_forecast(s) for s in dm.values()))
        cell(row, 1, f"{DAYS_ES[d.weekday()]} {d.strftime('%d/%m/%Y')}")
        cell(row, 2, dt, align='center')
        cell(row, 3, f_t, align='center')
        cell(row, 4, p_t, align='center')
        row += 1

    row += 1
    for cat in CAT_ORDER:
        items = {k: v for k, v in totals.items() if k[1] == cat and v >= 1.0}
        if not items:
            continue
        ws.merge_cells(f'A{row}:D{row}')
        ws[f'A{row}'] = cat.upper()
        ws[f'A{row}'].font      = Font(bold=True, color='FFFFFF')
        ws[f'A{row}'].fill      = NAVY
        ws[f'A{row}'].alignment = Alignment(horizontal='center')
        row += 1
        for lbl, c in [('Plato', 1), ('Promedio', 2), ('Buffer (σ)', 3), ('Total', 4)]:
            cell(row, c, lbl, bold=True, fill=GRAY, align='center')
        row += 1
        for (name, _), qty in sorted(items.items(), key=lambda x: -x[1]):
            st = dish_stats.get((name, _), {'mu': qty, 'sigma': 0})
            cell(row, 1, name)
            cell(row, 2, round(st['mu']),    align='center')
            cell(row, 3, round(st['sigma']), align='center')
            cell(row, 4, round(qty),         align='center')
            row += 1
        row += 1

    tf = round(sum(v for (_, c), v in totals.items() if c == 'Fondo'))
    ta = round(sum(totals.values()))
    for c in range(1, 5):
        ws.cell(row=row, column=c).fill = GRAY
    cell(row, 1, 'TOTAL FONDOS', bold=True); cell(row, 2, tf, bold=True, align='center')
    row += 1
    cell(row, 1, 'TOTAL PLATOS', bold=True); cell(row, 2, ta, bold=True, align='center')

    ws.column_dimensions['A'].width = 38
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10

    wb.save(fname)
    print(f"  Archivo guardado: {fname}")


# ── Main ───────────────────────────────────────────────────────────────────────

def main():
    args = sys.argv[1:]

    if args:
        try:
            start = datetime.strptime(args[0], '%d-%m-%Y')
        except ValueError:
            sys.exit(f"Fecha inválida '{args[0]}'. Formato: dd-mm-yyyy")
    else:
        start = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0) + timedelta(days=1)

    try:
        horizon = int(args[1]) if len(args) > 1 else DEFAULT_HORIZON
    except (ValueError, IndexError):
        horizon = DEFAULT_HORIZON

    if not HAS_RICH:
        print("\n  Tip: instala 'rich' para el display visual  →  pip install rich\n")

    print(f"\nCargando informes desde {_reports_folder} ...")
    history = load_history(_reports_folder)
    if not history:
        sys.exit("No se encontraron archivos de informe.")

    dates = sorted(history)
    print(f"  {len(history)} días cargados  ({dates[0]:%d/%m/%Y} – {dates[-1]:%d/%m/%Y})")

    model                       = build_model(history)
    day_fc, totals, dish_stats  = make_forecast(model, start, horizon)

    display(day_fc, totals, dish_stats, history, model)

    try:
        resp = input("  Exportar →  [h] HTML  [e] Excel  [a] Ambos  [n] No: ").strip().lower()
    except EOFError:
        resp = 'n'

    if resp in ('h', 'a'):
        fname = export_html(day_fc, totals, dish_stats, history)
        try:
            import webbrowser
            webbrowser.open(os.path.abspath(fname))
        except Exception:
            pass

    if resp in ('e', 'a'):
        export_excel(day_fc, totals, dish_stats, history)


if __name__ == '__main__':
    main()
